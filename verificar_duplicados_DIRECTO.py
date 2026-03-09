"""
VERIFICAR DUPLICADOS - MÉTODO DIRECTO CON OPENPYXL
"""
from openpyxl import load_workbook
from collections import Counter
import os

archivo = r'uploads\dian\Dian_23022026.xlsx'

print("\n" + "="*80)
print("VERIFICANDO DUPLICADOS - MÉTODO DIRECTO")
print("="*80)

if not os.path.exists(archivo):
    print(f"\n❌ Archivo no encontrado")
    exit()

print(f"\n📂 Abriendo Excel con openpyxl (solo lectura)...")

# Abrir Excel
wb = load_workbook(archivo, read_only=True, data_only=True)
ws = wb.active

print(f"✅ Excel abierto")

# Leer header (primera fila)
headers = []
for cell in ws[1]:
    headers.append(str(cell.value).strip().lower())

print(f"✅ Headers leídos: {len(headers)} columnas")

# Buscar columna CUFE
col_cufe_idx = None
for idx, header in enumerate(headers):
    if 'cufe' in header or 'cude' in header:
        col_cufe_idx = idx + 1  # openpyxl usa índice 1-based
        print(f"🔍 Columna CUFE encontrada: '{header}' en posición {col_cufe_idx}")
        break

if not col_cufe_idx:
    print("\n❌ No se encontró columna CUFE")
    wb.close()
    exit()

print(f"\n📖 Leyendo CUFEs de la columna {col_cufe_idx}...")

# Leer todos los CUFEs
cufes = []
fila_num = 0
for row in ws.iter_rows(min_row=2, min_col=col_cufe_idx, max_col=col_cufe_idx):
    fila_num += 1
    cufe = row[0].value
    if cufe:
        cufes.append(str(cufe).strip())
    
    if fila_num % 10000 == 0:
        print(f"   Leídas {fila_num:,} filas...")

wb.close()

print(f"✅ Total CUFEs leídos: {len(cufes):,}")

# Análisis
unicos = len(set(cufes))
duplicados = len(cufes) - unicos

print(f"\n📊 RESULTADO:")
print(f"   Total CUFEs: {len(cufes):,}")
print(f"   CUFEs únicos: {unicos:,}")
print(f"   Duplicados: {duplicados:,}")

if duplicados > 0:
    print(f"\n❌ ¡HAY {duplicados:,} CUFEs DUPLICADOS!")
    
    # Contar repeticiones
    contador = Counter(cufes)
    dups = [(cufe, count) for cufe, count in contador.items() if count > 1]
    dups.sort(key=lambda x: x[1], reverse=True)
    
    print(f"\n🔍 Top 10 CUFEs más repetidos:")
    cufe_error = "929f7761de9ff5fd92865b32d3aabbd4e056589f9cb854c0cfc15a570564f657aba35f6131872b4523c43152f393c793"
    
    for i, (cufe, count) in enumerate(dups[:10]):
        marca = " ⚠️ ERROR!" if cufe == cufe_error else ""
        print(f"   {i+1}. {cufe[:50]}... → {count} veces{marca}")
    
    print("\n❌ CONCLUSIÓN: El Excel tiene duplicados. Debes limpiarlo primero.")
else:
    print(f"\n✅ EXCELENTE: No hay duplicados")
    print("\n🤔 Si no hay duplicados Y las tablas están vacías,")
    print("   el problema está en el CÓDIGO de inserción.")

print("\n")
