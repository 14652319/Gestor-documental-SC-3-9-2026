"""
Routes SAGRILAFT - Sistema de Gestión de Radicados
Adaptado como Flask Blueprint para integración en proyecto principal
"""

from flask import render_template, jsonify, request, send_file, session
from decoradores_permisos import requiere_permiso_html, requiere_permiso
# Usar importación relativa para evitar circular import
from . import sagrilaft_bp
from .models import TerceroPreregistro
# NO importar de app aquí - causaría circular import
# Los modelos se importarán dentro de las funciones (lazy import)
from extensions import db
from datetime import datetime, timedelta
from sqlalchemy import text
import os
from io import BytesIO

# Ruta a documentos (usar ruta del proyecto principal)
RUTA_DOCUMENTOS = r'C:\Users\Usuario\Desktop\Gestor Documental\PAQUETES_TRANSPORTABLES\GESTOR_DOCUMENTAL_TRANSPORTABLE_20251113_204059\documentos_terceros'


@sagrilaft_bp.route('/')
@sagrilaft_bp.route('/radicados')
@requiere_permiso_html('sagrilaft', 'acceder_modulo')
def index():
    """Lista de radicados pendientes"""
    return render_template('lista_radicados.html')


@sagrilaft_bp.route('/api/radicados/pendientes', methods=['GET'])
@requiere_permiso('sagrilaft', 'listar_radicados')
def listar_radicados_pendientes():
    """Lista todos los radicados ordenados del más antiguo arriba"""
    from app import SolicitudRegistro, Tercero  # Lazy import - Tercero de app.py
    try:
        # JOIN con Tercero (tabla 'terceros'), NO con TerceroPreregistro
        solicitudes = db.session.query(SolicitudRegistro, Tercero)\
            .join(Tercero, SolicitudRegistro.tercero_id == Tercero.id)\
            .order_by(SolicitudRegistro.fecha_solicitud.asc())\
            .all()
        
        resultado = []
        fecha_hoy = datetime.now()
        
        # 📋 OBTENER OBSERVACIONES REALES DE LA TABLA
        from sqlalchemy import text
        radicados_list = [s.radicado for s, t in solicitudes]
        
        obs_query = text("""
            SELECT DISTINCT ON (radicado) radicado, observacion
            FROM observaciones_radicado
            WHERE radicado = ANY(:radicados)
            ORDER BY radicado, fecha_registro DESC
        """)
        
        observaciones_map = {}
        if radicados_list:  # Solo consultar si hay radicados
            with db.engine.connect() as conn:
                obs_rows = conn.execute(obs_query, {"radicados": radicados_list}).fetchall()
                observaciones_map = {row[0]: row[1] for row in obs_rows}
        
        for solicitud, tercero in solicitudes:
            try:
                dias_desde = (fecha_hoy - solicitud.fecha_solicitud).days if solicitud.fecha_solicitud else 0
                proxima_actualizacion = solicitud.fecha_solicitud + timedelta(days=360) if solicitud.fecha_solicitud else fecha_hoy
                dias_restantes = (proxima_actualizacion - fecha_hoy).days
                
                # Limpiar caracteres problem áticos
                razon_social = tercero.razon_social or ''
                if isinstance(razon_social, bytes):
                    razon_social = razon_social.decode('utf-8', errors='replace')
                
                # 💡 USAR OBSERVACIÓN REAL si existe, sino mensaje genérico
                estado_actual = solicitud.estado or 'pendiente'
                observacion_real = observaciones_map.get(solicitud.radicado)
                
                if observacion_real:
                    mensaje_estado = observacion_real  # 🎯 Observación real del usuario
                else:
                    # Mensaje genérico si no hay observación registrada
                    mensaje_estado = {
                        'pendiente': 'Radicado en revisión. Pendiente de análisis de documentos SAGRILAFT.',
                        'aprobado': 'Radicado APROBADO. Proveedor cumple con todos los requisitos SAGRILAFT.',
                        'rechazado': 'Radicado RECHAZADO. Proveedor no cumple con requisitos SAGRILAFT.',
                        'aprobado_condicionado': 'Sin observaciones registradas.'
                    }.get(estado_actual, 'Estado sin información adicional.')
                
                resultado.append({
                    'id': solicitud.id,
                    'radicado': solicitud.radicado or '',
                    'fecha_solicitud': solicitud.fecha_solicitud.strftime('%d/%m/%Y') if solicitud.fecha_solicitud else '',
                    'nit': tercero.nit or '',
                    'nombre': razon_social,
                    'dias_desde_radicado': dias_desde,
                    'estado': estado_actual,
                    'proxima_actualizacion': proxima_actualizacion.strftime('%d/%m/%Y'),
                    'dias_para_actualizacion': max(0, dias_restantes),
                    'observacion': mensaje_estado
                })
            except Exception as e:
                # Si un registro falla, continuar con el siguiente
                print(f"⚠️ Error procesando radicado {solicitud.radicado}: {e}")
                continue
        
        return jsonify({
            'success': True,
            'radicados': resultado,
            'total': len(resultado)
        })
        
    except Exception as e:
        print(f"❌ Error en listar_radicados_pendientes: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@sagrilaft_bp.route('/api/radicados/<radicado>/documentos', methods=['GET'])
@requiere_permiso('sagrilaft', 'revisar_documentos')
def obtener_documentos_radicado(radicado):
    """Obtiene documentos de un radicado específico"""
    from app import SolicitudRegistro, DocumentoTercero, Tercero  # Lazy import
    try:
        solicitud = SolicitudRegistro.query.filter_by(radicado=radicado).first()
        
        if not solicitud:
            return jsonify({
                'success': False,
                'error': 'Radicado no encontrado'
            }), 404
        
        tercero = Tercero.query.get(solicitud.tercero_id)
        documentos = DocumentoTercero.query.filter_by(tercero_id=tercero.id).all()
        
        resultado = []
        for doc in documentos:
            # Consultar estado del documento
            estado_query = text("""
                SELECT estado, observacion, usuario_revisor, fecha_revision
                FROM estados_documentos
                WHERE documento_id = :doc_id
            """)
            
            with db.engine.connect() as conn:
                estado_result = conn.execute(estado_query, {"doc_id": doc.id}).fetchone()
            
            if estado_result:
                estado = estado_result[0]
                observacion = estado_result[1] or ''
            else:
                estado = 'pendiente'
                observacion = ''
            
            resultado.append({
                'id': doc.id,
                'tipo_documento': doc.tipo_documento,
                'nombre_archivo': os.path.basename(doc.ruta_archivo),
                'ruta_relativa': doc.ruta_archivo,
                'estado': estado,
                'observacion': observacion
            })
        
        return jsonify({
            'success': True,
            'radicado': radicado,
            'nit': tercero.nit,
            'nombre': tercero.razon_social,
            'documentos': resultado,
            'total': len(resultado),
            'estado_radicado': solicitud.estado  # Estado final del radicado
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@sagrilaft_bp.route('/api/documentos/<int:doc_id>/visualizar', methods=['GET'])
@requiere_permiso('sagrilaft', 'revisar_documentos')
def visualizar_documento(doc_id):
    """Retorna el PDF para visualizar"""
    from app import DocumentoTercero  # Lazy import
    try:
        documento = DocumentoTercero.query.get(doc_id)
        
        if not documento:
            print(f"❌ Documento ID {doc_id} no encontrado en BD")
            return jsonify({
                'success': False,
                'error': 'Documento no encontrado en base de datos'
            }), 404
        
        print(f"📄 Documento ID {doc_id}: {documento.tipo_documento}")
        print(f"📂 Ruta en BD: {documento.ruta_archivo}")
        
        # Normalizar la ruta - QUITAR "documentos_terceros" del inicio si existe
        ruta_db = documento.ruta_archivo.replace('/', os.sep).replace('\\', os.sep)
        
        # Si la ruta en BD empieza con "documentos_terceros", quitarlo
        if ruta_db.startswith('documentos_terceros' + os.sep):
            ruta_db = ruta_db[len('documentos_terceros' + os.sep):]
        
        print(f"📝 Ruta normalizada: {ruta_db}")
        
        # Intentar múltiples formas de construir la ruta
        posibles_rutas = [
            os.path.join(RUTA_DOCUMENTOS, ruta_db),  # Forma 1: RUTA_DOCUMENTOS + ruta relativa
            ruta_db if os.path.isabs(ruta_db) else None,  # Forma 2: Si es ruta absoluta
            os.path.join(RUTA_DOCUMENTOS, os.path.basename(ruta_db))  # Forma 3: Solo nombre del archivo
        ]
        
        posibles_rutas = [r for r in posibles_rutas if r]  # Eliminar None
        
        ruta_completa = None
        for ruta in posibles_rutas:
            ruta_normalizada = os.path.normpath(ruta)
            print(f"🔍 Probando: {ruta_normalizada}")
            if os.path.exists(ruta_normalizada):
                ruta_completa = ruta_normalizada
                print(f"✅ Archivo encontrado: {ruta_completa}")
                break
        
        if not ruta_completa:
            print(f"❌ Archivo no encontrado en ninguna ubicación")
            print(f"   RUTA_DOCUMENTOS: {RUTA_DOCUMENTOS}")
            return jsonify({
                'success': False,
                'error': f'Archivo no encontrado: {ruta_db}'
            }), 404
        
        return send_file(ruta_completa, mimetype='application/pdf')
        
    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@sagrilaft_bp.route('/api/documentos/<int:doc_id>/estado', methods=['POST'])
@requiere_permiso('sagrilaft', 'actualizar_estado_documento')
def actualizar_estado_documento(doc_id):
    """Actualiza el estado de un documento individual"""
    from app import DocumentoTercero  # Lazy import
    try:
        data = request.get_json()
        estado = data.get('estado')
        observacion = data.get('observacion', '')
        usuario = data.get('usuario', 'Administrador')
        
        # Validar que el documento existe
        documento = DocumentoTercero.query.get(doc_id)
        if not documento:
            return jsonify({
                'success': False,
                'error': 'Documento no encontrado'
            }), 404
        
        # UPSERT en tabla estados_documentos
        upsert_query = text("""
            INSERT INTO estados_documentos (documento_id, estado, observacion, usuario_revisor, fecha_revision)
            VALUES (:doc_id, :estado, :observacion, :usuario, :fecha)
            ON CONFLICT (documento_id)
            DO UPDATE SET
                estado = :estado,
                observacion = :observacion,
                usuario_revisor = :usuario,
                fecha_revision = :fecha
        """)
        
        with db.engine.connect() as conn:
            conn.execute(upsert_query, {
                "doc_id": doc_id,
                "estado": estado,
                "observacion": observacion,
                "usuario": usuario,
                "fecha": datetime.now()
            })
            conn.commit()
        
        return jsonify({
            'success': True,
            'message': f'Estado actualizado a: {estado}'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@sagrilaft_bp.route('/api/radicados/<radicado>/estado', methods=['POST'])
@requiere_permiso('sagrilaft', 'actualizar_estado_radicado')
def actualizar_estado_radicado(radicado):
    """Actualiza el estado global de un radicado"""
    from app import SolicitudRegistro, Tercero, mail  # Lazy import
    from .email_sagrilaft import enviar_correo_decision_radicado
    try:
        data = request.get_json()
        estado = data.get('estado')
        observacion = data.get('observacion', '')
        enviar_correo = data.get('enviar_correo', False)
        usuario = data.get('usuario', 'Sistema')
        
        solicitud = SolicitudRegistro.query.filter_by(radicado=radicado).first()
        
        if not solicitud:
            return jsonify({
                'success': False,
                'error': 'Radicado no encontrado'
            }), 404
        
        # ⚠️ ACTUALIZAR SOLO CAMPOS QUE EXISTEN en SolicitudRegistro
        # Campos reales: id, tercero_id, radicado, estado, fecha_solicitud, fecha_actualizacion
        solicitud.estado = estado
        solicitud.fecha_actualizacion = datetime.now()
        
        # � GUARDAR OBSERVACIÓN EN TABLA OBSERVACIONES_RADICADO
        if observacion and observacion.strip():
            from sqlalchemy import text
            insert_observacion = text("""
                INSERT INTO observaciones_radicado 
                (radicado, estado, observacion, usuario, fecha_registro)
                VALUES (:radicado, :estado, :observacion, :usuario, :fecha)
            """)
            
            with db.engine.connect() as conn:
                conn.execute(insert_observacion, {
                    "radicado": radicado,
                    "estado": estado,
                    "observacion": observacion.strip(),
                    "usuario": usuario,
                    "fecha": datetime.now()
                })
                conn.commit()
            
            print(f"✅ Observación guardada para {radicado}: {observacion[:50]}...")
        
        db.session.commit()
        
        # ✅ Envío de correo implementado (OPCIONAL - solo si está configurado)
        correo_enviado = False
        mensaje_correo = ""
        
        if enviar_correo:
            try:
                # Obtener datos del tercero
                tercero = Tercero.query.get(solicitud.tercero_id)
                
                # Verificar si tercero tiene email (campo puede no existir en modelo antiguo)
                email_tercero = getattr(tercero, 'email', None) if tercero else None
                
                if email_tercero:
                    # Importar función de envío
                    from .email_sagrilaft import enviar_correo_decision_radicado
                    
                    # Enviar correo de decisión
                    success, msg = enviar_correo_decision_radicado(
                        mail=mail,
                        destinatario=email_tercero,
                        nit=tercero.nit,
                        razon_social=tercero.razon_social,
                        radicado=radicado,
                        estado=estado,
                        observacion=observacion
                    )
                    
                    if success:
                        correo_enviado = True
                        mensaje_correo = f"Correo enviado a {email_tercero}"
                        print(f"✅ {mensaje_correo}")
                    else:
                        mensaje_correo = f"No se pudo enviar correo: {msg}"
                        print(f"⚠️ {mensaje_correo}")
                else:
                    mensaje_correo = "Tercero no tiene email registrado"
                    print(f"⚠️ {mensaje_correo}")
            except Exception as e:
                # Si falla el envío de correo, no bloquear el cambio de estado
                mensaje_correo = f"Error al intentar enviar correo: {str(e)}"
                print(f"⚠️ {mensaje_correo}")
        
        return jsonify({
            'success': True,
            'message': f'Radicado {estado} correctamente',
            'radicado': radicado,
            'estado': estado,
            'correo_enviado': correo_enviado,
            'mensaje_correo': mensaje_correo
        })
        
    except Exception as e:
        print(f"❌ Error actualizando radicado: {e}")
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@sagrilaft_bp.route('/revisar/<radicado>')
@requiere_permiso_html('sagrilaft', 'revisar_documentos')
def revisar_documentos(radicado):
    """Página de revisión de documentos"""
    return render_template('revisar_documentos.html', radicado=radicado)


@sagrilaft_bp.route('/api/radicados/exportar', methods=['POST'])
@requiere_permiso('sagrilaft', 'exportar_radicados')
def exportar_radicados_excel():
    """Exporta radicados seleccionados a Excel"""
    from app import SolicitudRegistro, Tercero  # Lazy import
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        data = request.get_json()
        radicados_selec = data.get('radicados', [])
        
        if not radicados_selec:
            return jsonify({'success': False, 'error': 'No se seleccionaron radicados'}), 400
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Radicados"
        
        # Headers con estilo
        headers = ['Fecha', 'RAD', 'NIT', 'Nombre/Razón Social', 'Días Desde Radicado', 
                  'Estado', 'Próxima Actualización', 'Días Restantes', 'Observaciones']
        
        header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        row = 2
        fecha_hoy = datetime.now().date()  # Fecha actual como date (sin hora)
        
        # 📋 OBTENER OBSERVACIONES REALES DE LA TABLA
        observaciones_map = {}
        try:
            obs_query = text("""
                SELECT DISTINCT ON (radicado) radicado, observacion
                FROM observaciones_radicado
                WHERE radicado = ANY(:radicados)
                ORDER BY radicado, fecha_registro DESC
            """)
            
            if radicados_selec:
                with db.engine.connect() as conn:
                    obs_rows = conn.execute(obs_query, {"radicados": radicados_selec}).fetchall()
                    observaciones_map = {row[0]: row[1] for row in obs_rows}
        except Exception as e:
            print(f"⚠️ Error obteniendo observaciones: {e}")
            # Continuar sin observaciones si falla la consulta
        
        for rad_str in radicados_selec:
            solicitud = SolicitudRegistro.query.filter_by(radicado=rad_str).first()
            if not solicitud:
                continue
            
            tercero = Tercero.query.get(solicitud.tercero_id)
            if not tercero:
                continue
            
            # Calcular días - asegurar que todo sea date (no datetime)
            fecha_rad = solicitud.fecha_solicitud
            if fecha_rad:
                # Si es datetime, convertir a date
                if hasattr(fecha_rad, 'date'):
                    fecha_rad = fecha_rad.date()
                dias_desde = (fecha_hoy - fecha_rad).days
            else:
                dias_desde = 0
            
            # Próxima actualización
            fecha_prox = solicitud.fecha_actualizacion if solicitud.fecha_actualizacion else solicitud.fecha_solicitud
            if fecha_prox:
                # Si es datetime, convertir a date
                if hasattr(fecha_prox, 'date'):
                    fecha_prox = fecha_prox.date()
                proxima = fecha_prox + timedelta(days=365)
                dias_restantes = (proxima - fecha_hoy).days
                proxima_str = proxima.strftime('%d/%m/%Y')
            else:
                dias_restantes = 0
                proxima_str = ''
            
            # Obtener observación real de la tabla
            observacion_excel = observaciones_map.get(rad_str, '')
            
            # Escribir datos en Excel
            ws.cell(row=row, column=1, value=solicitud.fecha_solicitud.strftime('%d/%m/%Y') if solicitud.fecha_solicitud else '')
            ws.cell(row=row, column=2, value=rad_str)
            ws.cell(row=row, column=3, value=tercero.nit)
            ws.cell(row=row, column=4, value=tercero.razon_social)
            ws.cell(row=row, column=5, value=dias_desde)
            ws.cell(row=row, column=6, value=solicitud.estado or 'pendiente')
            ws.cell(row=row, column=7, value=proxima_str)
            ws.cell(row=row, column=8, value=max(0, dias_restantes))
            ws.cell(row=row, column=9, value=observacion_excel)  # 🎯 Observación real
            row += 1
        
        # Ajustar anchos
        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(
            output,
            as_attachment=True,
            download_name=f'Radicados_{timestamp}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        }), 500


@sagrilaft_bp.route('/api/radicados/<radicado>/descargar_documentos')
@requiere_permiso('sagrilaft', 'descargar_documentos')
def descargar_documentos_radicado(radicado):
    """Descarga todos los documentos de un radicado en ZIP"""
    from app import SolicitudRegistro, DocumentoTercero, Tercero  # Lazy import
    try:
        import zipfile
        
        solicitud = SolicitudRegistro.query.filter_by(radicado=radicado).first()
        
        if not solicitud:
            return jsonify({
                'success': False,
                'error': 'Radicado no encontrado'
            }), 404
        
        tercero = Tercero.query.get(solicitud.tercero_id)
        documentos = DocumentoTercero.query.filter_by(tercero_id=tercero.id).all()
        
        if not documentos:
            return jsonify({
                'success': False,
                'error': 'No hay documentos para este radicado'
            }), 404
        
        # Crear ZIP en memoria
        zip_buffer = BytesIO()
        archivos_agregados = 0
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for doc in documentos:
                # Normalizar ruta del documento
                ruta_db = doc.ruta_archivo.replace('/', os.sep).replace('\\', os.sep)
                
                # La ruta completa es RUTA_DOCUMENTOS + nombre del archivo
                if 'documentos_terceros' in ruta_db:
                    # Si la ruta ya incluye documentos_terceros, usar directamente
                    ruta_completa = os.path.join(
                        r'C:\Users\Usuario\Desktop\Gestor Documental\PAQUETES_TRANSPORTABLES\GESTOR_DOCUMENTAL_TRANSPORTABLE_20251113_204059',
                        ruta_db
                    )
                else:
                    # Si no, agregar RUTA_DOCUMENTOS
                    ruta_completa = os.path.join(RUTA_DOCUMENTOS, ruta_db)
                
                ruta_completa = os.path.normpath(ruta_completa)
                
                if os.path.exists(ruta_completa):
                    # Nombre del archivo en el ZIP: TIPO_NIT.pdf
                    nombre_archivo = f"{doc.tipo_documento}_{tercero.nit}.pdf"
                    zip_file.write(ruta_completa, nombre_archivo)
                    archivos_agregados += 1
        
        if archivos_agregados == 0:
            return jsonify({
                'success': False,
                'error': 'No se encontraron archivos físicos'
            }), 404
        
        zip_buffer.seek(0)
        
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f'{radicado}_Documentos.zip',
            mimetype='application/zip'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
