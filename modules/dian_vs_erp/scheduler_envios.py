"""
============================================
SCHEDULER DE ENVÍOS PROGRAMADOS AUTOMÁTICOS
MÓDULO DIAN VS ERP - PostgreSQL
============================================
Fecha: 25/Diciembre/2025
Descripción: Sistema que ejecuta envíos de emails automáticos
             basados en configuraciones programadas
"""

import json
import logging
from datetime import datetime, time, timedelta
from typing import List, Dict, Any
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import os

from extensions import db
from modules.dian_vs_erp.models import (
    MaestroDianVsErp, 
    EnvioProgramadoDianVsErp, 
    UsuarioCausacionDianVsErp,
    HistorialEnvioDianVsErp,
    # ✅ NUEVOS: Tablas optimizadas para LEFT JOINs (igual que visor_v2)
    Dian,
    ErpFinanciero,
    ErpComercial,
    Acuses
)

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/scheduler_dian_vs_erp.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class EnviosProgramadosSchedulerDianVsErp:
    """Scheduler para envíos automáticos programados del módulo DIAN VS ERP"""
    
    def __init__(self, app=None):
        self.scheduler = BackgroundScheduler()
        self.app = app
        self.smtp_config = {
            'server': os.getenv('MAIL_SERVER', 'smtp.gmail.com'),
            'port': int(os.getenv('MAIL_PORT', 465)),
            'username': os.getenv('MAIL_USERNAME'),
            'password': os.getenv('MAIL_PASSWORD'),
            'use_ssl': os.getenv('MAIL_USE_SSL', 'True') == 'True',
            'use_tls': os.getenv('MAIL_USE_TLS', 'False') == 'True'
        }
        
    def cargar_configuraciones(self) -> List[Dict]:
        """Carga todas las configuraciones activas"""
        try:
            with self.app.app_context():
                configs = EnvioProgramadoDianVsErp.query.filter_by(activo=True).all()
                logger.info(f"📋 Cargadas {len(configs)} configuraciones activas")
                return [c.to_dict() for c in configs]
            
        except Exception as e:
            logger.error(f"❌ Error cargando configuraciones: {e}")
            return []
    
    def iniciar_scheduler(self):
        """Inicia el scheduler con todas las configuraciones"""
        try:
            # Cargar configuraciones activas
            configs = self.cargar_configuraciones()
            
            for config in configs:
                # Crear trigger cron basado en frecuencia
                hora, minuto = config['hora_envio'].split(':')
                
                if config['frecuencia'] == 'DIARIO':
                    trigger = CronTrigger(hour=int(hora), minute=int(minuto))
                elif config['frecuencia'] == 'SEMANAL':
                    dias = json.loads(config['dias_semana'] or '[1]')  # Default: Lunes
                    trigger = CronTrigger(
                        day_of_week=','.join(map(str, dias)),
                        hour=int(hora),
                        minute=int(minuto)
                    )
                else:
                    continue
                
                # Agregar job al scheduler
                job_id = f"dian_vs_erp_envio_{config['id']}"
                self.scheduler.add_job(
                    func=self.ejecutar_envio_programado,
                    trigger=trigger,
                    args=[config['id']],
                    id=job_id,
                    name=config['nombre'],
                    replace_existing=True
                )
                
                logger.info(f"✅ Job agregado: {config['nombre']} ({config['hora_envio']})")
            
            # Iniciar el scheduler
            self.scheduler.start()
            logger.info("🚀 Scheduler DIAN VS ERP iniciado")
            
        except Exception as e:
            logger.error(f"❌ Error iniciando scheduler: {e}")
    
    def detener_scheduler(self):
        """Detiene el scheduler"""
        if self.scheduler.running:
            self.scheduler.shutdown()
            logger.info("🛑 Scheduler DIAN VS ERP detenido")
    
    def reiniciar_scheduler(self):
        """Reinicia el scheduler (al modificar configuraciones)"""
        logger.info("🔄 Reiniciando scheduler...")
        self.detener_scheduler()
        self.iniciar_scheduler()
    
    def ejecutar_envio_programado(self, config_id: int):
        """
        Ejecuta un envío programado específico
        
        Args:
            config_id: ID de la configuración a ejecutar
        """
        with self.app.app_context():  # USAR CONTEXTO DE FLASK
            inicio = datetime.now()
            logger.info(f"🚀 Iniciando envío programado ID={config_id}")
            
            try:
                # Cargar configuración
                config = EnvioProgramadoDianVsErp.query.get(config_id)
                if not config:
                    logger.error(f"❌ Configuración ID={config_id} no encontrada")
                    return
                
                # 🆕 SUPERVISIÓN: Envío a supervisor (consolidado, sin agrupar por NIT)
                if config.es_supervision:
                    logger.info(f"🔍 Ejecutando envío de SUPERVISIÓN: {config.nombre}")
                    resultado = self._enviar_supervision_general(config)
                # Envíos normales (agrupados por NIT)
                elif config.tipo == 'PENDIENTES_DIAS':
                    resultado = self._procesar_pendientes_dias(config)
                elif config.tipo == 'CREDITO_SIN_ACUSES':
                    resultado = self._procesar_credito_sin_acuses(config)
                else:
                    resultado = {'estado': 'FALLIDO', 'mensaje': 'Tipo de configuración desconocido'}
                
                # Actualizar último envío
                config.ultimo_envio = datetime.now()
                config.proximo_envio = self._calcular_proximo_envio(config)
                config.total_ejecuciones = (config.total_ejecuciones or 0) + 1
                config.total_documentos_procesados = (config.total_documentos_procesados or 0) + resultado.get('docs_procesados', 0)
                config.total_emails_enviados = (config.total_emails_enviados or 0) + resultado.get('emails_enviados', 0)
                
                # Registrar en historial
                duracion = (datetime.now() - inicio).total_seconds()
                historial = HistorialEnvioDianVsErp(
                    configuracion_id=config_id,
                    fecha_ejecucion=datetime.now(),
                    documentos_procesados=resultado.get('docs_procesados', 0),
                    documentos_enviados=resultado.get('docs_enviados', 0),
                    emails_enviados=resultado.get('emails_enviados', 0),
                    emails_fallidos=resultado.get('emails_fallidos', 0),
                    destinatarios_enviados=json.dumps(resultado.get('destinatarios_ok', [])),
                    destinatarios_fallidos=json.dumps(resultado.get('destinatarios_error', [])),
                    estado=resultado['estado'],
                    mensaje=resultado['mensaje'],
                    duracion_segundos=duracion
                )
                
                db.session.add(historial)
                db.session.commit()
                
                logger.info(f"✅ Envío programado ID={config_id} completado en {duracion:.2f}s")
                logger.info(f"   📧 Emails enviados: {resultado.get('emails_enviados', 0)}")
                logger.info(f"   📄 Documentos incluidos: {resultado.get('docs_enviados', 0)}")
                
            except Exception as e:
                logger.error(f"❌ Error ejecutando envío programado ID={config_id}: {e}")
                db.session.rollback()
                # Registrar error en historial
                try:
                    historial = HistorialEnvioDianVsErp(
                        configuracion_id=config_id,
                        fecha_ejecucion=datetime.now(),
                        estado='FALLIDO',
                        mensaje=str(e),
                        errores=json.dumps({'error': str(e)})
                    )
                    db.session.add(historial)
                    db.session.commit()
                except:
                    pass
    
    def _procesar_pendientes_dias(self, config: EnvioProgramadoDianVsErp) -> Dict:
        """
        Procesa documentos pendientes por X días
        
        Lógica:
        - Busca docs con dias_desde_emision >= dias_minimos
        - Excluye estados en estados_excluidos
        - Agrupa por usuario asignado (de tabla usuarios_asignados)
        - Envía 1 email consolidado por usuario
        """
        try:
            # ✅ YA NO USAMOS UsuarioAsignadoDianVsErp - usamos usuarios_asignados directamente
            
            dias_min = config.dias_minimos
            
            # Parsear estados_excluidos y asegurar que sea una lista
            estados_excl = []
            if config.estados_excluidos:
                try:
                    parsed = json.loads(config.estados_excluidos)
                    # Si es dict, convertir a lista vacía (era '{}')
                    estados_excl = parsed if isinstance(parsed, list) else []
                except:
                    estados_excl = []
            
            # 🆕 Parsear tipos_tercero (filtro de tipo de tercero)
            tipos_tercero = []
            if config.tipos_tercero:
                try:
                    parsed = json.loads(config.tipos_tercero)
                    tipos_tercero = parsed if isinstance(parsed, list) else []
                except:
                    tipos_tercero = []
            
            # ✅ NUEVA QUERY: Usar tablas optimizadas con LEFT JOINs (igual que visor_v2)
            # Esto garantiza datos en tiempo real y NO envía documentos causados
            query = db.session.query(Dian).outerjoin(
                ErpFinanciero,
                Dian.clave == ErpFinanciero.clave_erp_financiero
            ).outerjoin(
                ErpComercial,
                Dian.clave == ErpComercial.clave_erp_comercial
            ).outerjoin(
                Acuses,
                Dian.cufe_cude == Acuses.cufe  # ✅ CORREGIDO: usar cufe_cude
            ).filter(
                Dian.n_dias >= dias_min,
                # ✅ CRÍTICO: Excluir documentos causados (igual que visor_v2)
                ErpFinanciero.id == None,
                ErpComercial.id == None
            )
            
            if estados_excl:
                query = query.filter(~Acuses.estado_docto.in_(estados_excl))
            
            # 🆕 FILTRO POR RANGO DE FECHAS (si está configurado)
            if config.fecha_inicio:
                query = query.filter(Dian.fecha_emision >= config.fecha_inicio)
            if config.fecha_fin:
                query = query.filter(Dian.fecha_emision <= config.fecha_fin)
            
            # 🆕 Aplicar filtro de tipos_tercero si está configurado
            if tipos_tercero:
                query = query.filter(Dian.tipo_tercero.in_(tipos_tercero))
            
            # ✅ ORDENAR: Más días arriba (más antiguo), menos días abajo (más reciente)
            query = query.order_by(Dian.n_dias.desc())
            
            documentos = query.all()
            
            logger.info(f"   ✅ USANDO TABLAS OPTIMIZADAS (Dian + JOINs) - Datos en tiempo real")
            logger.info(f"   🚫 Documentos causados: EXCLUIDOS automáticamente")
            
            if not documentos:
                return {
                    'estado': 'EXITOSO',
                    'mensaje': 'No hay documentos pendientes que cumplan los criterios',
                    'docs_procesados': 0,
                    'docs_enviados': 0,
                    'emails_enviados': 0
                }
            
            # ✅ Normalizar atributos de objetos Dian para compatibilidad con templates de email
            for doc in documentos:
                doc.cufe = doc.cufe_cude  # cufe_cude → cufe
                doc.valor = doc.total  # total → valor
                doc.razon_social = doc.nombre_emisor  # nombre_emisor → razon_social
                doc.dias_desde_emision = doc.n_dias  # n_dias → dias_desde_emision
            
            # Agrupar por NIT y buscar usuarios asignados
            docs_por_nit = {}
            for doc in documentos:
                nit = doc.nit_emisor
                if nit not in docs_por_nit:
                    docs_por_nit[nit] = []
                docs_por_nit[nit].append(doc)
            
            # Buscar usuarios asignados para cada NIT en la tabla usuarios_asignados
            docs_por_usuario = {}
            for nit, docs in docs_por_nit.items():
                # ✅ USAR CONSULTA DIRECTA SQL A LA TABLA CORRECTA
                query = """
                    SELECT correo, nombres, apellidos 
                    FROM usuarios_asignados 
                    WHERE nit = :nit AND activo = true
                """
                result = db.session.execute(db.text(query), {'nit': nit})
                usuarios = result.fetchall()
                
                for usuario in usuarios:
                    email = usuario[0]  # correo
                    if email not in docs_por_usuario:
                        docs_por_usuario[email] = []
                    docs_por_usuario[email].extend(docs)
            
            # Enviar emails consolidados
            emails_ok = 0
            emails_error = 0
            destinatarios_ok = []
            destinatarios_error = []
            
            for email, docs in docs_por_usuario.items():
                try:
                    self._enviar_email_consolidado(
                        destinatario=email,
                        asunto=f"⏰ Documentos pendientes - {len(docs)} facturas con más de {dias_min} días",
                        documentos=docs,
                        tipo='PENDIENTES_DIAS',
                        dias_min=dias_min
                    )
                    emails_ok += 1
                    destinatarios_ok.append({'email': email, 'documentos': len(docs)})
                    logger.info(f"   ✅ Email enviado a {email} ({len(docs)} docs)")
                except Exception as e:
                    emails_error += 1
                    destinatarios_error.append({'email': email, 'error': str(e)})
                    logger.error(f"   ❌ Error enviando a {email}: {e}")
            
            return {
                'estado': 'EXITOSO' if emails_error == 0 else 'PARCIAL',
                'mensaje': f'Emails enviados: {emails_ok}, Fallidos: {emails_error}',
                'docs_procesados': len(documentos),
                'docs_enviados': len(documentos),
                'emails_enviados': emails_ok,
                'emails_fallidos': emails_error,
                'destinatarios_ok': destinatarios_ok,
                'destinatarios_error': destinatarios_error
            }
            
        except Exception as e:
            logger.error(f"❌ Error en _procesar_pendientes_dias: {e}")
            return {'estado': 'FALLIDO', 'mensaje': str(e)}
    
    def _procesar_credito_sin_acuses(self, config: EnvioProgramadoDianVsErp) -> Dict:
        """
        Procesa documentos de crédito sin acuses completos
        
        Lógica:
        - Busca docs con forma_pago LIKE "%Crédito%" o "%Credit%"
        - Filtra causada=True
        - Verifica acuses_recibidos < requiere_acuses_minimo
        - Agrupa por NIT emisor (igual que pendientes_dias)
        - Busca usuarios asignados por NIT en tabla usuarios_asignados
        - Envía 1 email consolidado por usuario
        """
        try:
            requiere_acuses = config.requiere_acuses_minimo or 2
            logger.info(f"   📊 Buscando documentos de crédito con < {requiere_acuses} acuses...")
            
            # ✅ CORREGIDO: Forma de pago usa códigos numéricos
            # '2' = Crédito (519K docs), '1' = Contado (170K docs), '02' = Crédito alternativo
            forma_pago_credito = ['2', '02']
            
            # 🆕 Parsear tipos_tercero (filtro de tipo de tercero)
            tipos_tercero = []
            if config.tipos_tercero:
                try:
                    parsed = json.loads(config.tipos_tercero)
                    tipos_tercero = parsed if isinstance(parsed, list) else []
                except:
                    tipos_tercero = []
            
            # ✅ NUEVA QUERY: Usar tablas optimizadas con LEFT JOINs (igual que visor_v2)
            # Esto garantiza datos en tiempo real y NO envía documentos causados
            
            # Subconsulta para contar acuses por CUFE (corregido)
            acuses_count = db.session.query(
                Acuses.cufe,  # Tabla Acuses tiene campo cufe
                db.func.count(Acuses.id).label('total_acuses')
            ).group_by(Acuses.cufe).subquery()
            
            # ✅ Query que INCLUYE el conteo de acuses en los resultados
            query_docs = db.session.query(
                Dian,
                db.func.coalesce(acuses_count.c.total_acuses, 0).label('total_acuses')
            ).outerjoin(
                ErpFinanciero,
                Dian.clave == ErpFinanciero.clave_erp_financiero
            ).outerjoin(
                ErpComercial,
                Dian.clave == ErpComercial.clave_erp_comercial
            ).outerjoin(
                acuses_count,
                Dian.cufe_cude == acuses_count.c.cufe
            ).filter(
                Dian.forma_pago.in_(forma_pago_credito),
                # ✅ CRÍTICO: Excluir documentos causados
                ErpFinanciero.id == None,
                ErpComercial.id == None,
                # ✅ Filtro de acuses insuficientes
                db.func.coalesce(acuses_count.c.total_acuses, 0) < requiere_acuses
            )
            
            # 🆕 FILTRO POR RANGO DE FECHAS (si está configurado)
            if config.fecha_inicio:
                query_docs = query_docs.filter(Dian.fecha_emision >= config.fecha_inicio)
            if config.fecha_fin:
                query_docs = query_docs.filter(Dian.fecha_emision <= config.fecha_fin)
            
            # 🆕 Aplicar filtro de tipos_tercero si está configurado
            if tipos_tercero:
                query_docs = query_docs.filter(Dian.tipo_tercero.in_(tipos_tercero))
            
            results = query_docs.order_by(Dian.n_dias.desc()).all()
            
            # ✅ Extraer objetos Dian y agregar conteo de acuses
            documentos = []
            for dian_obj, total_acuses in results:
                # Normalizar atributos
                dian_obj.cufe = dian_obj.cufe_cude
                dian_obj.valor = dian_obj.total
                dian_obj.razon_social = dian_obj.nombre_emisor
                dian_obj.dias_desde_emision = dian_obj.n_dias
                dian_obj.acuses_recibidos = total_acuses
                dian_obj.acuses_requeridos = requiere_acuses
                documentos.append(dian_obj)
            
            logger.info(f"   ✅ USANDO TABLAS OPTIMIZADAS (Dian + JOINs) - Datos en tiempo real")
            logger.info(f"   🚫 Documentos causados: EXCLUIDOS automáticamente")
            
            logger.info(f"   📄 Documentos encontrados para enviar: {len(documentos)}")
            
            if not documentos:
                return {
                    'estado': 'EXITOSO',
                    'mensaje': 'No hay documentos de crédito sin acuses suficientes',
                    'docs_procesados': 0,
                    'docs_enviados': 0,
                    'emails_enviados': 0
                }
            
            # Agrupar por NIT emisor (igual que pendientes_dias)
            docs_por_nit = {}
            for doc in documentos:
                nit = doc.nit_emisor
                if nit not in docs_por_nit:
                    docs_por_nit[nit] = []
                docs_por_nit[nit].append(doc)
            
            logger.info(f"   🏢 NITs con documentos: {len(docs_por_nit)}")
            
            # ✅ BUSCAR USUARIOS POR NIT EN LA TABLA CORRECTA: usuarios_asignados
            docs_por_usuario = {}
            for nit, docs in docs_por_nit.items():
                # Consulta SQL directa a la tabla usuarios_asignados
                query = """
                    SELECT correo, nombres, apellidos 
                    FROM usuarios_asignados 
                    WHERE nit = :nit AND activo = true
                """
                result = db.session.execute(db.text(query), {'nit': nit})
                usuarios = result.fetchall()
                
                logger.info(f"   👤 NIT {nit}: {len(usuarios)} usuarios encontrados")
                
                for usuario in usuarios:
                    email = usuario[0]  # correo
                    if email not in docs_por_usuario:
                        docs_por_usuario[email] = []
                    docs_por_usuario[email].extend(docs)
            
            logger.info(f"   📧 Emails destino: {len(docs_por_usuario)}")
            
            if not docs_por_usuario:
                logger.warning("   ⚠️ No se encontraron usuarios asignados para los NITs con documentos de crédito")
                return {
                    'estado': 'EXITOSO',
                    'mensaje': 'No hay usuarios asignados configurados para los NITs con documentos',
                    'docs_procesados': len(documentos),
                    'docs_enviados': 0,
                    'emails_enviados': 0
                }
            
            # Enviar emails consolidados
            emails_ok = 0
            emails_error = 0
            destinatarios_ok = []
            destinatarios_error = []
            
            for email, docs in docs_por_usuario.items():
                try:
                    self._enviar_email_consolidado(
                        destinatario=email,
                        asunto=f"⚠️ Documentos de crédito sin acuses - {len(docs)} facturas",
                        documentos=docs,
                        tipo='CREDITO_SIN_ACUSES',
                        requiere_acuses=requiere_acuses
                    )
                    emails_ok += 1
                    destinatarios_ok.append({'email': email, 'documentos': len(docs)})
                    logger.info(f"   ✅ Email enviado a {email} ({len(docs)} docs)")
                except Exception as e:
                    emails_error += 1
                    destinatarios_error.append({'email': email, 'error': str(e)})
                    logger.error(f"   ❌ Error enviando a {email}: {e}")
            
            return {
                'estado': 'EXITOSO' if emails_error == 0 else 'PARCIAL',
                'mensaje': f'Emails enviados: {emails_ok}, Fallidos: {emails_error}',
                'docs_procesados': len(documentos),
                'docs_enviados': len(documentos),
                'emails_enviados': emails_ok,
                'emails_fallidos': emails_error,
                'destinatarios_ok': destinatarios_ok,
                'destinatarios_error': destinatarios_error
            }
            
        except Exception as e:
            logger.error(f"❌ Error en _procesar_credito_sin_acuses: {e}")
            return {'estado': 'FALLIDO', 'mensaje': str(e)}
    
    def _enviar_email_consolidado(self, destinatario: str, asunto: str, 
                                   documentos: List, tipo: str, **kwargs):
        """
        Envía email consolidado con lista de documentos
        
        Args:
            destinatario: Email del destinatario
            asunto: Asunto del email
            documentos: Lista de objetos MaestroDianVsErp
            tipo: Tipo de envío (PENDIENTES_DIAS, CREDITO_SIN_ACUSES)
            **kwargs: Parámetros adicionales (dias_min, causador, etc.)
        """
        # ✅ LIMITAR HTML A 20 DOCUMENTOS (pero generar Excel con todos si hay más)
        total_docs = len(documentos)
        docs_para_html = documentos[:20] if total_docs > 20 else documentos
        
        # Construir HTML del email (solo primeros 20)
        html = self._generar_html_email(docs_para_html, tipo, total_documentos=total_docs, **kwargs)
        
        # Crear mensaje
        msg = MIMEMultipart('mixed')  # Cambiar a 'mixed' para soportar adjuntos
        msg['From'] = self.smtp_config['username']
        msg['To'] = destinatario
        msg['Subject'] = asunto
        
        # Agregar HTML
        msg_html = MIMEMultipart('alternative')
        msg_html.attach(MIMEText(html, 'html'))
        msg.attach(msg_html)
        
        # ✅ SI HAY MÁS DE 20 DOCUMENTOS, GENERAR EXCEL ADJUNTO
        if total_docs > 20:
            try:
                excel_content = self._generar_excel_documentos(documentos, tipo, **kwargs)
                from email.mime.application import MIMEApplication
                excel_adjunto = MIMEApplication(excel_content, _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                fecha_archivo = datetime.now().strftime('%Y%m%d_%H%M%S')
                excel_adjunto.add_header('Content-Disposition', 'attachment', filename=f'Documentos_Pendientes_{fecha_archivo}.xlsx')
                msg.attach(excel_adjunto)
                logger.info(f"   📎 Excel adjunto generado con {total_docs} documentos")
            except Exception as e:
                logger.error(f"   ❌ Error generando Excel adjunto: {e}")
        
        # Enviar según configuración (SSL o TLS)
        if self.smtp_config['use_ssl']:
            # Puerto 465 con SSL
            import smtplib
            with smtplib.SMTP_SSL(self.smtp_config['server'], self.smtp_config['port']) as server:
                server.login(self.smtp_config['username'], self.smtp_config['password'])
                server.send_message(msg)
        else:
            # Puerto 587 con TLS
            with smtplib.SMTP(self.smtp_config['server'], self.smtp_config['port']) as server:
                if self.smtp_config['use_tls']:
                    server.starttls()
                server.login(self.smtp_config['username'], self.smtp_config['password'])
                server.send_message(msg)
        
        logger.info(f"📧 Email enviado a {destinatario}")
    
    def _generar_html_email(self, documentos: List, tipo: str, **kwargs) -> str:
        """Genera HTML del email consolidado"""
        
        total_documentos = kwargs.get('total_documentos', len(documentos))
        total_valor = sum(float(doc.valor or 0) for doc in documentos)
        
        # ⚠️ ADVERTENCIA si se muestran solo primeros 20 de más documentos
        advertencia_html = ""
        if total_documentos > 20:
            advertencia_html = f'''
            <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0; border-radius: 4px;">
                <strong>⚠️ Se muestran los primeros 20 documentos de {total_documentos} totales.</strong><br>
                📎 Consulte el archivo Excel adjunto para ver el listado completo.
            </div>
            '''
        
        if tipo == 'PENDIENTES_DIAS':
            dias_min = kwargs.get('dias_min', 3)
            titulo = f"⏰ Documentos pendientes con más de {dias_min} días"
            intro = f"Se le informa que tiene <strong>{total_documentos}</strong> documentos pendientes de causación/recepción con más de {dias_min} días desde su emisión."
            columnas_extra = '<th>Días</th>'
        else:  # CREDITO_SIN_ACUSES
            causador = kwargs.get('causador', 'Usuario')
            requiere_acuses = kwargs.get('requiere_acuses', 2)
            titulo = f"⚠️ Documentos de crédito sin acuses completos"
            intro = f"Estimado/a {causador},<br><br>Se le informa que tiene <strong>{len(documentos)}</strong> documentos de crédito causados que aún no tienen los acuses requeridos (mínimo {requiere_acuses} de 3)."
            columnas_extra = '<th>Acuses</th>'
        
        # HTML
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; background: #f5f5f5; padding: 20px; }}
                .container {{ max-width: 900px; margin: 0 auto; background: white; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .header {{ background: #00c875; color: white; padding: 20px; border-radius: 8px 8px 0 0; }}
                .content {{ padding: 30px; }}
                .summary {{ background: #f0f0f0; padding: 15px; border-radius: 6px; margin: 20px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 13px; }}
                th {{ background: #00c875; color: white; padding: 10px; text-align: left; font-size: 12px; }}
                td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
                tr:hover {{ background: #f9f9f9; }}
                .footer {{ background: #f0f0f0; padding: 15px; text-align: center; border-radius: 0 0 8px 8px; color: #666; font-size: 12px; }}
                .link {{ color: #00c875; text-decoration: none; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2 style="margin:0;">{titulo}</h2>
                    <p style="margin:5px 0 0 0;">Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
                </div>
                
                <div class="content">
                    <p>{intro}</p>
                    
                    {advertencia_html}
                    
                    <div style="margin: 20px 0; padding: 15px; background-color: #fef2f2; border-left: 4px solid #dc2626; border-radius: 4px;">
                        <p style="margin: 0; color: #dc2626; font-weight: bold;">⚠️ IMPORTANTE - Acción Requerida:</p>
                        <p style="margin: 10px 0 5px 0; color: #333; font-size: 15px;">
                            Por favor, <strong>gestione y tramite estos documentos pendientes lo más pronto posible</strong>.
                        </p>
                        <p style="margin: 5px 0 0 0; color: #666; font-size: 14px;">
                            Los documentos con mayor antigüedad requieren atención prioritaria para evitar retrasos en el proceso contable.
                        </p>
                    </div>
                    
                    <div style="margin: 0 0 20px 0; padding: 15px; background-color: #f0fdf4; border-left: 4px solid #166534; border-radius: 4px;">
                        <p style="margin: 0; color: #166534; font-weight: bold;">💡 Información del listado:</p>
                        <ul style="margin: 10px 0; color: #666;">
                            <li><strong>Días:</strong> Días transcurridos desde la fecha de emisión</li>
                            <li><strong>Estado:</strong> Estado de aprobación en el sistema</li>
                            <li><strong>Documento:</strong> Click en "Ver en DIAN" para consultar el documento</li>
                        </ul>
                    </div>
                    
                    <div class="summary">
                        <strong>📊 Resumen:</strong><br>
                        • Total de documentos: {total_documentos}<br>
                        • Valor total: ${total_valor:,.2f}
                    </div>
                    
                    <h3>📄 Detalle de documentos:</h3>
                    <table>
                        <thead>
                            <tr>
                                <th>#</th>
                                <th>NIT</th>
                                <th>Proveedor</th>
                                <th>Factura</th>
                                <th>Fecha Emisión</th>
                                <th>Valor</th>
                                {columnas_extra}
                                <th>Documento DIAN</th>
                            </tr>
                        </thead>
                        <tbody>
        """
        
        for i, doc in enumerate(documentos, 1):
            fecha_emision = doc.fecha_emision.strftime('%Y-%m-%d') if doc.fecha_emision else ''
            cufe = doc.cufe or ''
            link_dian = f'https://catalogo-vpfe.dian.gov.co/User/SearchDocument?DocumentKey={cufe}' if cufe else '#'
            
            if tipo == 'PENDIENTES_DIAS':
                valor_extra = f'<td>{doc.dias_desde_emision or 0}</td>'
            else:
                valor_extra = f'<td>{doc.acuses_recibidos}/{doc.acuses_requeridos}</td>'
            
            html += f"""
                            <tr>
                                <td>{i}</td>
                                <td>{doc.nit_emisor or ''}</td>
                                <td style="max-width:150px; overflow:hidden; text-overflow:ellipsis;">{(doc.razon_social or '')[:35]}...</td>
                                <td>{doc.prefijo or ''}-{doc.folio or ''}</td>
                                <td>{fecha_emision}</td>
                                <td>${float(doc.valor or 0):,.0f}</td>
                                {valor_extra}
                                <td><a href="{link_dian}" class="link" target="_blank">{'Ver en DIAN' if cufe else 'N/A'}</a></td>
                            </tr>
            """
        
        html += """
                        </tbody>
                    </table>
                    
                    <div style="margin-top: 20px; padding: 15px; background: #f0f0f0; border-radius: 6px; text-align: center;">
                        <strong>Sistema de Facturación Electrónica - SUPERTIENDAS CAÑAVERAL</strong><br>
                        <span style="color: #666; font-size: 12px;">Este es un mensaje automático, por favor no responder.</span>
                    </div>
                </div>
            </div>
        </body>
        </html>"""
        
        return html
    
    def _generar_excel_documentos(self, documentos: List, tipo: str, **kwargs) -> bytes:
        """
        Genera un archivo Excel con todos los documentos
        Retorna el contenido del Excel como bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Documentos Pendientes"
            
            # Estilos
            header_fill = PatternFill(start_color="00C875", end_color="00C875", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border_thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezados
            if tipo == 'PENDIENTES_DIAS':
                encabezados = ['#', 'NIT', 'Proveedor', 'Factura', 'Fecha Emisión', 'Valor', 'Días Pendientes', 'Estado', 'CUFE']
            else:
                encabezados = ['#', 'NIT', 'Proveedor', 'Factura', 'Fecha Emisión', 'Valor', 'Acuses', 'Estado', 'CUFE']
            
            for col, header in enumerate(encabezados, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_thin
            
            # Datos
            for i, doc in enumerate(documentos, 1):
                row_num = i + 1
                fecha_emision = doc.fecha_emision.strftime('%Y-%m-%d') if doc.fecha_emision else ''
                cufe = doc.cufe or ''
                
                if tipo == 'PENDIENTES_DIAS':
                    datos = [
                        i,
                        doc.nit_emisor or '',
                        doc.razon_social or '',
                        f"{doc.prefijo or ''}-{doc.folio or ''}",
                        fecha_emision,
                        float(doc.valor or 0),
                        doc.dias_desde_emision or 0,
                        doc.estado_aprobacion or '',
                        cufe  # CUFE (se convertirá en hipervínculo después)
                    ]
                else:
                    datos = [
                        i,
                        doc.nit_emisor or '',
                        doc.razon_social or '',
                        f"{doc.prefijo or ''}-{doc.folio or ''}",
                        fecha_emision,
                        float(doc.valor or 0),
                        f"{doc.acuses_recibidos}/{doc.acuses_requeridos}",
                        doc.estado_aprobacion or '',
                        cufe  # CUFE (se convertirá en hipervínculo después)
                    ]
                
                for col, valor in enumerate(datos, 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.value = valor
                    cell.border = border_thin
                    if col == 6:  # Columna de valor
                        cell.number_format = '$#,##0'
                    elif col == 9 and cufe:  # ✅ COLUMNA CUFE - AGREGAR HIPERVÍNCULO
                        url_dian = f"https://catalogo-vpfe.dian.gov.co/User/SearchDocument?DocumentKey={cufe}"
                        cell.hyperlink = url_dian
                        cell.style = "Hyperlink"  # Estilo azul con subrayado
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 60
            
            # Fila de totales
            total_row = len(documentos) + 2
            ws.cell(row=total_row, column=5).value = "TOTAL:"
            ws.cell(row=total_row, column=5).font = Font(bold=True)
            ws.cell(row=total_row, column=6).value = sum(float(doc.valor or 0) for doc in documentos)
            ws.cell(row=total_row, column=6).number_format = '$#,##0'
            ws.cell(row=total_row, column=6).font = Font(bold=True)
            
            # Auto filtro
            ws.auto_filter.ref = f"A1:I{len(documentos) + 1}"
            
            # Guardar en BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except ImportError:
            logger.error("❌ openpyxl no está instalado. Instalar con: pip install openpyxl")
            raise Exception("openpyxl requerido para generar Excel")
    
    def _enviar_supervision_general(self, config: EnvioProgramadoDianVsErp) -> Dict:
        """
        🆕 Envía correo de supervisión consolidado (TODOS los docs, sin agrupar por NIT)
        
        Args:
            config: Configuración de envío (debe tener es_supervision=True y email_supervisor)
        
        Returns:
            Dict con estado, mensaje, stats
        """
        logger.info(f"📊 === ENVÍO DE SUPERVISIÓN: {config.nombre} ===")
        
        try:
            # Validar configuración
            if not config.email_supervisor:
                return {'estado': 'FALLIDO', 'mensaje': 'Email de supervisor no configurado'}
            
            # 🆕 Parsear tipos_tercero (filtro de tipo de tercero)
            tipos_tercero = []
            if config.tipos_tercero:
                try:
                    parsed = json.loads(config.tipos_tercero)
                    tipos_tercero = parsed if isinstance(parsed, list) else []
                except:
                    tipos_tercero = []
            
            # Construir query según tipo
            if config.tipo == 'PENDIENTES_DIAS':
                # Documentos pendientes +X días (NO REGISTRADAS)
                query = MaestroDianVsErp.query.filter(
                    MaestroDianVsErp.dias_desde_emision >= (config.dias_minimos or 3),
                    MaestroDianVsErp.causada == False  # NO registradas
                )
                if config.estados_incluidos:
                    estados_inc = json.loads(config.estados_incluidos)
                    if estados_inc:
                        query = query.filter(MaestroDianVsErp.estado_aprobacion.in_(estados_inc))
                
                # 🆕 Aplicar filtro de tipos_tercero si está configurado
                if tipos_tercero:
                    query = query.filter(MaestroDianVsErp.tipo_tercero.in_(tipos_tercero))
                
                titulo_tipo = f"Documentos NO REGISTRADOS (+{config.dias_minimos} días)"
                
            elif config.tipo == 'CREDITO_SIN_ACUSES':
                # Créditos causados con <2 acuses
                query = MaestroDianVsErp.query.filter(
                    MaestroDianVsErp.forma_pago == '2',  # Crédito
                    MaestroDianVsErp.causada == True,     # YA causados
                    MaestroDianVsErp.acuses_recibidos < (config.requiere_acuses_minimo or 2)
                )
                
                # 🆕 Aplicar filtro de tipos_tercero si está configurado
                if tipos_tercero:
                    query = query.filter(MaestroDianVsErp.tipo_tercero.in_(tipos_tercero))
                
                titulo_tipo = f"Créditos Causados con <{config.requiere_acuses_minimo or 2} Acuses"
            
            else:
                return {'estado': 'FALLIDO', 'mensaje': f'Tipo {config.tipo} no soportado para supervisión'}
            
            # Aplicar filtros de estados excluidos
            if config.estados_excluidos:
                estados_exc = json.loads(config.estados_excluidos)
                if estados_exc:
                    query = query.filter(~MaestroDianVsErp.estado_aprobacion.in_(estados_exc))
            
            # Ejecutar query y ordenar por antigüedad (más viejos primero)
            documentos = query.order_by(MaestroDianVsErp.dias_desde_emision.desc()).all()
            
            if not documentos:
                logger.info(f"ℹ️  No hay documentos para enviar en supervisión: {config.nombre}")
                return {
                    'estado': 'EXITO',
                    'mensaje': 'No hay documentos pendientes para supervisión',
                    'docs_procesados': 0,
                    'docs_enviados': 0,
                    'emails_enviados': 0
                }
            
            total_documentos = len(documentos)
            logger.info(f"📄 {total_documentos} documentos encontrados para supervisión")
            
            # ⚠️ LIMITAR A 50 DOCUMENTOS para evitar correos demasiado grandes
            MAX_DOCS_EMAIL = 50
            documentos_mostrar = documentos[:MAX_DOCS_EMAIL]
            hay_mas_documentos = total_documentos > MAX_DOCS_EMAIL
            
            logger.info(f"📧 Mostrando {len(documentos_mostrar)} de {total_documentos} documentos en el correo")
            
            # Preparar datos para el email (solo los primeros 50)
            docs_data = []
            for doc in documentos_mostrar:
                docs_data.append({
                    'nit_emisor': doc.nit_emisor,
                    'razon_social': doc.razon_social or 'N/A',
                    'prefijo': doc.prefijo,
                    'folio': doc.folio,
                    'fecha_emision': doc.fecha_emision.strftime('%d/%m/%Y') if doc.fecha_emision else 'N/A',
                    'valor_total': f"${doc.valor:,.2f}" if doc.valor else '$0',
                    'dias_desde_emision': doc.dias_desde_emision or 0,
                    'estado_aprobacion': doc.estado_aprobacion or 'N/A',
                    'estado_contable': 'Causada' if doc.causada else 'No Registrada',
                    'acuses_recibidos': 0  # Campo no disponible en este modelo
                })
            
            # Generar HTML del email
            html_body = self._generar_email_supervision(
                titulo=titulo_tipo,
                documentos=docs_data,
                config_nombre=config.nombre,
                total_documentos=total_documentos,
                hay_mas=hay_mas_documentos
            )
            
            # Enviar correo
            from flask_mail import Message
            from app import mail  # ✅ CORRECTO: mail está en app.py, no en extensions
            
            asunto = f"🔍 {titulo_tipo} - {total_documentos} documento(s)"
            msg = Message(
                subject=asunto,
                recipients=[config.email_supervisor],
                html=html_body
            )
            
            try:
                mail.send(msg)
                logger.info(f"✅ Email de supervisión enviado a: {config.email_supervisor}")
                
                return {
                    'estado': 'EXITO',
                    'mensaje': f'Supervisión enviada exitosamente a {config.email_supervisor}',
                    'docs_procesados': total_documentos,
                    'docs_enviados': len(documentos_mostrar),
                    'emails_enviados': 1,
                    'emails_fallidos': 0,
                    'destinatarios_ok': [config.email_supervisor],
                    'destinatarios_error': []
                }
            
            except Exception as e:
                logger.error(f"❌ Error enviando email de supervisión: {e}")
                return {
                    'estado': 'FALLIDO',
                    'mensaje': f'Error al enviar: {str(e)}',
                    'docs_procesados': len(documentos),
                    'docs_enviados': 0,
                    'emails_enviados': 0,
                    'emails_fallidos': 1,
                    'destinatarios_error': [config.email_supervisor]
                }
        
        except Exception as e:
            logger.error(f"❌ Error en envío de supervisión: {e}")
            return {
                'estado': 'FALLIDO',
                'mensaje': str(e),
                'docs_procesados': 0
            }
    
    def _generar_email_supervision(self, titulo: str, documentos: List[Dict], config_nombre: str, 
                                   total_documentos: int = None, hay_mas: bool = False) -> str:
        """Genera HTML para email de supervisión"""
        
        if total_documentos is None:
            total_documentos = len(documentos)
        
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h2 {{ color: #2c3e50; border-bottom: 3px solid #e74c3c; padding-bottom: 10px; }}
                .info {{ background: #ecf0f1; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                .warning {{ background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin-bottom: 20px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th {{ background: #34495e; color: white; padding: 12px; text-align: left; }}
                td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
                tr:hover {{ background: #f5f5f5; }}
                .valor {{ text-align: right; font-weight: bold; }}
                .dias {{ text-align: center; font-weight: bold; color: #e74c3c; }}
                .footer {{ margin-top: 30px; padding-top: 20px; border-top: 2px solid #ddd; color: #7f8c8d; font-size: 12px; }}
            </style>
        </head>
        <body>
            <h2>🔍 {titulo}</h2>
            
            <div class="info">
                <strong>📋 Configuración:</strong> {config_nombre}<br>
                <strong>📄 Total documentos encontrados:</strong> {total_documentos}<br>
                <strong>📧 Documentos mostrados en este correo:</strong> {len(documentos)}<br>
                <strong>📅 Fecha reporte:</strong> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
            </div>
        """
        
        if hay_mas:
            html += f"""
            <div class="warning">
                <strong>⚠️ AVISO:</strong> Se encontraron {total_documentos} documentos en total, 
                pero solo se muestran los primeros {len(documentos)} para evitar correos demasiado grandes.<br>
                Consulta el sistema para ver el listado completo.
            </div>
            """
        
        html += """
            <table>
                <thead>
                    <tr>
                        <th>NIT</th>
                        <th>Razón Social</th>
                        <th>Prefijo-Folio</th>
                        <th>Fecha Emisión</th>
                        <th>Valor</th>
                        <th>Días</th>
                        <th>Estado Aprob.</th>
                        <th>Estado Contable</th>
                        <th>Acuses</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for doc in documentos:
            html += f"""
                    <tr>
                        <td>{doc['nit_emisor']}</td>
                        <td>{doc['razon_social']}</td>
                        <td>{doc['prefijo']}-{doc['folio']}</td>
                        <td>{doc['fecha_emision']}</td>
                        <td class="valor">{doc['valor_total']}</td>
                        <td class="dias">{doc['dias_desde_emision']}</td>
                        <td>{doc['estado_aprobacion']}</td>
                        <td>{doc['estado_contable']}</td>
                        <td style="text-align:center">{doc['acuses_recibidos']}</td>
                    </tr>
            """
        
        html += """
                </tbody>
            </table>
            
            <div class="footer">
                <p>Este es un correo automático de supervisión del Sistema de Gestión Documental DIAN vs ERP.</p>
                <p>⚠️ No responder a este correo. Para consultas, contactar al administrador del sistema.</p>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def _calcular_proximo_envio(self, config: EnvioProgramadoDianVsErp) -> datetime:
        """Calcula la próxima fecha de envío"""
        hora, minuto = map(int, config.hora_envio.split(':'))
        ahora = datetime.now()
        proximo = ahora.replace(hour=hora, minute=minuto, second=0, microsecond=0)
        
        if config.frecuencia == 'DIARIO':
            if proximo <= ahora:
                proximo += timedelta(days=1)
        elif config.frecuencia == 'SEMANAL':
            dias = json.loads(config.dias_semana or '[1]')
            while proximo.weekday() + 1 not in dias or proximo <= ahora:
                proximo += timedelta(days=1)
        
        return proximo


# Instancia global del scheduler
scheduler_dian_vs_erp_global = None

def iniciar_scheduler_dian_vs_erp(app=None):
    """Inicia el scheduler global al arrancar la aplicación"""
    global scheduler_dian_vs_erp_global
    if scheduler_dian_vs_erp_global is None:
        scheduler_dian_vs_erp_global = EnviosProgramadosSchedulerDianVsErp(app)
        scheduler_dian_vs_erp_global.iniciar_scheduler()
    return scheduler_dian_vs_erp_global

def detener_scheduler_dian_vs_erp():
    """Detiene el scheduler global"""
    global scheduler_dian_vs_erp_global
    if scheduler_dian_vs_erp_global:
        scheduler_dian_vs_erp_global.detener_scheduler()
        scheduler_dian_vs_erp_global = None

def reiniciar_scheduler_dian_vs_erp():
    """Reinicia el scheduler (usar al modificar configuraciones)"""
    global scheduler_dian_vs_erp_global
    if scheduler_dian_vs_erp_global:
        scheduler_dian_vs_erp_global.reiniciar_scheduler()
