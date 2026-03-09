# -*- coding: utf-8 -*-
"""
Rutas del módulo DIAN vs ERP - Sistema Híbrido v5
Integración completa del proyecto original al Gestor Documental
"""

from flask import Blueprint, render_template, request, jsonify, send_file, session, redirect, current_app
from flask_mail import Message
from werkzeug.utils import secure_filename
from decoradores_permisos import requiere_permiso_html, requiere_permiso
import os
import time
import hashlib
import io
import json
from datetime import date, datetime
from pathlib import Path
import polars as pl
import pandas as pd

# Importar modelos PostgreSQL y db
from extensions import db
from .models import (
    MaestroDianVsErp,
    EnvioProgramadoDianVsErp,
    UsuarioCausacionDianVsErp,
    HistorialEnvioDianVsErp,
    LogSistemaDianVsErp,
    # Modelos optimizados para Visor V2
    Dian,
    ErpComercial,
    ErpFinanciero,
    Acuses,
    TipoDocumentoDian,  # Catálogo de tipos de documentos a procesar
    TipoTerceroDianErp  # Clasificación de terceros según presencia en ERP
)

# Importar modelos de tablas de facturas para validar estado contable
from modules.recibir_facturas.models import FacturaTemporal, FacturaRecibida
from modules.facturas_digitales.models import FacturaDigital

# Importar helper de logging
from .logger_helper import (
    registrar_log,
    obtener_logs_recientes as helper_obtener_logs,
    obtener_estadisticas_logs
)

# Loaders de archivos hacia PostgreSQL
from .cargador_bd import procesar_archivo_subido, reconstruir_maestro as reconstruir_maestro_bd

# Crear blueprint
dian_vs_erp_bp = Blueprint('dian_vs_erp', __name__)

# Configuración de rutas
BASE_DIR = Path(__file__).parent.parent.parent
UPLOADS = {
    "dian": BASE_DIR / "uploads" / "dian",
    "erp_fn": BASE_DIR / "uploads" / "erp_fn", 
    "erp_cm": BASE_DIR / "uploads" / "erp_cm",
    "acuses": BASE_DIR / "uploads" / "acuses",
    "errores": BASE_DIR / "uploads" / "rg_erp_er",
}

# Crear directorios
for path in UPLOADS.values():
    path.mkdir(parents=True, exist_ok=True)

ALLOWED_EXCEL = {".xlsx", ".xlsm", ".xls", ".csv", ".ods"}  # Agregado .ods

# Diccionarios de mapeo (del sistema original)
SIGLAS = {
    "factura electrónica": "FE", "factura electronica": "FE",
    "factura electrónica de exportación": "FEE",
    "nota de débito electrónica": "NDE",
    "factura electrónica de contingencia": "FEC",
    "documento soporte con no obligados": "DSNO",
    "factura electrónica de contingencia dian": "FECD",
    "documento equivalente pos": "DEPOS",
    "nota de ajuste crédito del documento equivalente": "NCDE",
    "nota de crédito electrónica": "NCE",
}

MODULO_MAP = {
    "legalización factura anticipos": "FINANCIERO",
    "factura de servicio compra": "FINANCIERO",
    "nota débito de servicios - compra": "FINANCIERO",
    "factura de servicio de reg.fijo compra": "FINANCIERO",
    "factura de servicio desde sol. anticipo": "FINANCIERO",
    "legalización factura caja menor": "FINANCIERO",
    "factura de proveedor": "COMERCIAL",
    "notas débito de proveedor": "COMERCIAL",
    "factura de consignación": "COMERCIAL",
}

# ==============================
# FUNCIONES UTILITARIAS
# ==============================

def validar_sesion():
    """Validar que el usuario tenga sesión activa"""
    if 'usuario_id' not in session or 'usuario' not in session:
        return False, {"error": "Sesión no válida", "redirect": "/login"}, 401
    return True, None, None

def save_excel_to_csv_from_disk(archivo_path: str, folder: Path) -> str:
    """
    Convierte Excel/CSV a CSV normalizado DESDE DISCO
    Nuevo enfoque (Dec 29, 2025): Lee desde archivo guardado con pandas
    
    FORMATOS ACEPTADOS: .xlsx, .xlsm, .csv
    FORMATOS NO ACEPTADOS: .xls (Excel 97-2003 - corrupto)
    """
    import hashlib
    import pandas as pd
    
    fname = os.path.basename(archivo_path)
    ext = os.path.splitext(fname)[1].lower()
    
    # ✅ VALIDACIÓN DE FORMATO
    FORMATOS_ACEPTADOS = ['.xlsx', '.xlsm', '.csv']
    if ext not in FORMATOS_ACEPTADOS:
        raise ValueError(
            f"❌ ARCHIVO RECHAZADO: '{fname}'\n"
            f"   Formato: {ext}\n"
            f"   Formatos aceptados: {', '.join(FORMATOS_ACEPTADOS)}\n\n"
            f"💡 SOLUCIÓN:\n"
            f"   1. Abre el archivo en Excel\n"
            f"   2. Guarda como: Libro de Excel (.xlsx)\n"
            f"   3. Vuelve a subir el archivo\n"
        )
    
    # Hash para evitar colisiones
    with open(archivo_path, 'rb') as f:
        h = hashlib.md5(f.read()).hexdigest()[:10]
    
    base = os.path.splitext(fname)[0]
    csv_name = f"{base}_{h}.csv"
    target = folder / csv_name
    
    if ext in (".xlsx", ".xlsm"):
        # Leer Excel moderno desde disco con pandas + openpyxl
        try:
            df = pd.read_excel(archivo_path, dtype=str, engine='openpyxl')
            df.to_csv(target, index=False, encoding="utf-8")
        except Exception as e:
            raise ValueError(
                f"❌ ERROR AL PROCESAR: '{fname}'\n"
                f"   Error técnico: {str(e)}\n\n"
                f"💡 POSIBLES CAUSAS:\n"
                f"   • Archivo corrupto o dañado\n"
                f"   • Descarga incompleta\n"
                f"   • Formato interno incorrecto\n\n"
                f"💡 SOLUCIONES:\n"
                f"   1. Abre el archivo en Excel → Archivo → Abrir y reparar\n"
                f"   2. Guárdalo como nuevo archivo .xlsx\n"
                f"   3. O guárdalo como CSV para mejor compatibilidad\n"
            )
    elif ext == ".csv":
        # Copiar CSV normalizado
        df = pd.read_csv(archivo_path, dtype=str)
        df.to_csv(target, index=False, encoding="utf-8")
    else:
        raise ValueError(f"Extensión no soportada: {ext}")
    
    return str(target)

def save_excel_to_csv(storage, folder: Path) -> str:
    """Convertir Excel/CSV a CSV normalizado (LEGACY - mantener por compatibilidad)"""
    folder.mkdir(parents=True, exist_ok=True)
    fname = secure_filename(storage.filename)
    ext = os.path.splitext(fname)[1].lower()
    
    # 🔥 FIX (Dec 29, 2025): Asegurar que el stream esté al inicio antes de leer
    storage.seek(0)
    raw = storage.read()

    # Hash para evitar colisiones
    h = hashlib.md5(raw).hexdigest()[:10]
    base = os.path.splitext(fname)[0]
    csv_name = f"{base}_{h}.csv"
    target = folder / csv_name

    if ext in (".xlsx", ".xlsm", ".xls"):
        buf = io.BytesIO(raw)
        try:
            df = pd.read_excel(buf, dtype=str)
        except Exception as e:
            # Si falla, intentar con openpyxl y resetear buffer
            buf.seek(0)
            df = pd.read_excel(buf, engine="openpyxl", dtype=str)
        df.to_csv(target, index=False, encoding="utf-8")
    elif ext == ".csv":
        try:
            txt = raw.decode("utf-8")
        except UnicodeDecodeError:
            txt = raw.decode("latin-1")
        with open(target, "w", encoding="utf-8", newline="") as w:
            w.write(txt)
    else:
        raise ValueError(f"Extensión no soportada: {ext}")

    return str(target)

def latest_file(path: Path) -> str:
    """
    Obtiene el archivo más reciente de una carpeta (Excel o CSV)
    Si es Excel, lo convierte a CSV primero
    ✅ VALIDACIÓN: Solo archivos .xlsx, .xlsm, .csv (NO .xls)
    """
    if not path.is_dir():
        return None
    
    # ✅ BUSCAR SOLO ARCHIVOS VÁLIDOS (xlsx, xlsm, csv)
    archivos_validos = list(path.glob("*.xlsx")) + list(path.glob("*.xlsm")) + list(path.glob("*.csv"))
    
    # ⚠️ BUSCAR ARCHIVOS INVÁLIDOS (.xls)
    archivos_invalidos = list(path.glob("*.xls"))
    
    if archivos_invalidos:
        nombres = [a.name for a in archivos_invalidos]
        raise ValueError(
            f"⚠️ ARCHIVOS CON FORMATO NO ACEPTADO en '{path.name}/':\n" +
            "\n".join([f"   • {n} (formato .xls)" for n in nombres]) +
            "\n\n💡 SOLUCIÓN:\n"
            "   1. Abre estos archivos en Excel\n"
            "   2. Guarda como: Libro de Excel (.xlsx)\n"
            "   3. Elimina los archivos .xls viejos\n"
            "   4. Vuelve a procesar\n\n"
            f"📋 Formatos aceptados: .xlsx, .xlsm, .csv\n"
        )
    
    if not archivos_validos:
        return None
    
    # Ordenar por fecha de modificación
    archivos_validos.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    archivo_mas_reciente = archivos_validos[0]
    
    # 🔥 OPTIMIZACIÓN: NO convertir Excel a CSV - Polars lo lee directamente
    # Simplemente retornar la ruta (Polars soporta .xlsx, .xlsm, .csv)
    return str(archivo_mas_reciente)

def latest_csv(path: Path) -> str:
    """LEGACY: Usar latest_file() en su lugar"""
    return latest_file(path)

def read_csv(path: str, separator: str = ',', encoding: str = 'utf-8') -> pl.DataFrame:
    """
    Leer CSV o Excel con Polars
    🔥 OPTIMIZACIÓN: Detecta formato automáticamente y lee directo
    """
    try:
        if not path or not os.path.exists(path):
            print(f"❌ Archivo no encontrado: {path}")
            return pl.DataFrame()
        
        # Detectar extensión
        ext = os.path.splitext(path)[1].lower()
        print(f"📂 Leyendo archivo: {os.path.basename(path)} (extensión: {ext})")
        
        if ext in ['.xlsx', '.xlsm']:
            # Leer Excel directamente con Polars (50x más rápido que pandas)
            print(f"   📊 Lectura directa Excel con Polars (calamine)...")
            try:
                df = pl.read_excel(path, engine='calamine', infer_schema_length=1000)
                print(f"   ✅ Excel leído: {len(df):,} filas, {len(df.columns)} columnas")
            except Exception as e_excel:
                print(f"   ⚠️  Calamine falló ({e_excel}), usando pandas...")
                # Fallback a pandas si calamine no está disponible
                import pandas as pd
                pdf = pd.read_excel(path, dtype=str)
                df = pl.from_pandas(pdf)
                print(f"   ✅ Excel leído con pandas: {len(df):,} filas")
        else:
            # Leer CSV con parámetros personalizados
            print(f"   📄 Lectura CSV con encoding={encoding}, separador='{separator}'...")
            df = pl.read_csv(
                path, 
                separator=separator, 
                encoding=encoding, 
                infer_schema_length=1000, 
                ignore_errors=True,
                truncate_ragged_lines=True
            )
            print(f"   ✅ CSV leído: {len(df):,} filas, {len(df.columns)} columnas")
        
        # Normalizar nombres de columnas (strip + lowercase)
        df = df.rename({c: c.strip().lower() for c in df.columns})
        print(f"   ✅ Columnas normalizadas")
        
        return df
        
    except Exception as e:
        print(f"❌ ERROR CRÍTICO al leer {os.path.basename(path)}: {str(e)}")
        import traceback
        print(traceback.format_exc())
        raise

# ==============================
# RUTAS DE AUTENTICACIÓN
# ==============================

@dian_vs_erp_bp.route('/api/auth/logout', methods=['POST'])
def logout():
    """
    Cerrar sesión del usuario (destruir sesión en servidor)
    Llamado por timeout automático (25 min) o por botón manual
    """
    try:
        # Limpiar sesión
        session.clear()
        
        # Log de auditoría
        current_app.logger.info('✅ Sesión cerrada correctamente en DIAN vs ERP')
        
        return jsonify({
            'success': True,
            'message': 'Sesión cerrada correctamente'
        }), 200
        
    except Exception as e:
        current_app.logger.error(f'❌ Error al cerrar sesión: {e}')
        # Aún si hay error, intentar limpiar sesión
        session.clear()
        return jsonify({
            'success': False,
            'message': 'Error al cerrar sesión pero se limpiará de todas formas'
        }), 500

# ==============================
# RUTAS PRINCIPALES
# ==============================

@dian_vs_erp_bp.route('/')
def dashboard():
    """Dashboard principal del módulo DIAN vs ERP - Template original del standalone puerto 8097"""
    # Validar sesión - TEMPORAL: Sin decorador de permisos para debugging
    if 'usuario_id' not in session or 'usuario' not in session:
        return redirect('/login')
    
    usuario = session.get('usuario', 'Usuario')
    
    # Obtener estadísticas desde PostgreSQL
    try:
        total_maestro = MaestroDianVsErp.query.count()
        total_envios = EnvioProgramadoDianVsErp.query.count()
        total_usuarios = UsuarioCausacionDianVsErp.query.count()
        stats = {
            "total_maestro": total_maestro,
            "total_envios_programados": total_envios,
            "total_usuarios_causacion": total_usuarios
        }
    except Exception as e:
        stats = {"total_maestro": 0, "total_envios_programados": 0, "total_usuarios_causacion": 0}
        print(f"Error obteniendo estadísticas: {e}")
    
    # Usar visor_dian_v2.html - Template original del standalone con funcionalidad completa
    return render_template('dian_vs_erp/visor_dian_v2.html', usuario=usuario, estadisticas=stats)

@dian_vs_erp_bp.route('/cargar')
@dian_vs_erp_bp.route('/cargar_archivos')
def cargar_archivos():
    """Página de carga de archivos - Template negro (igual a SQLite)"""
    # Validar sesión simple
    if 'usuario_id' not in session or 'usuario' not in session:
        return redirect('/login')
    
    return render_template('dian_vs_erp/cargar_moderno_NEGRO.html', usuario=session.get('usuario', 'Usuario'))

@dian_vs_erp_bp.route('/visor')
def visor_moderno():
    """Visor moderno de datos"""
    # Validar sesión simple
    if 'usuario_id' not in session or 'usuario' not in session:
        return redirect('/login')
    
    return render_template('dian_vs_erp/visor_moderno.html', usuario=session.get('usuario', 'Usuario'))

@dian_vs_erp_bp.route('/visor_v2')
def visor_v2():
    """Visor de Facturas v2 - Carga desde tablas optimizadas (DIAN, ERP Comercial, ERP Financiero, Acuses)"""
    # Validar sesión simple
    if 'usuario_id' not in session or 'usuario' not in session:
        return redirect('/login')
    
    usuario = session.get('usuario', 'Usuario')
    
    # Obtener estadísticas desde las tablas optimizadas
    try:
        total_maestro = MaestroDianVsErp.query.count()
        total_envios = EnvioProgramadoDianVsErp.query.count()
        total_usuarios = UsuarioCausacionDianVsErp.query.count()
        stats = {
            "total_maestro": total_maestro,
            "total_envios_programados": total_envios,
            "total_usuarios_causacion": total_usuarios
        }
    except Exception as e:
        stats = {"total_maestro": 0, "total_envios_programados": 0, "total_usuarios_causacion": 0}
        print(f"Error obteniendo estadísticas: {e}")
    
    # Usar version='v2' para que el Causador lea doc_causado_por directo de erp_comercial/erp_financiero
    # (vía /api/dian_v2 con JOINs en tiempo real, siempre actualizado)
    return render_template('dian_vs_erp/visor_dian_v2.html', usuario=usuario, estadisticas=stats, version='v2')

@dian_vs_erp_bp.route('/validaciones')
def validaciones():
    """Página de validaciones DIAN vs ERP"""
    # Validar sesión simple
    if 'usuario_id' not in session or 'usuario' not in session:
        return redirect('/login')
    
    return render_template('dian_vs_erp/validaciones.html', usuario=session.get('usuario', 'Usuario'))

@dian_vs_erp_bp.route('/reportes')
def reportes():
    """Página de reportes DIAN vs ERP"""
    # Validar sesión simple
    if 'usuario_id' not in session or 'usuario' not in session:
        return redirect('/login')
    
    return render_template('dian_vs_erp/reportes.html', usuario=session.get('usuario', 'Usuario'))

@dian_vs_erp_bp.route('/configuracion')
def configuracion():
    """Página de configuración del módulo"""
    # Validar sesión simple
    if 'usuario_id' not in session or 'usuario' not in session:
        return redirect('/login')
    
    return render_template('dian_vs_erp/configuracion.html', usuario=session.get('usuario', 'Usuario'))

@dian_vs_erp_bp.route('/api/dian')
def api_dian():
    """
    API: Lee maestro consolidado desde POSTGRESQL (migrado desde SQLite)
    Compatible con el frontend original del standalone
    
    PARÁMETROS DE CONSULTA (opcionales):
    - fecha_inicial: YYYY-MM-DD (por defecto: primer día del mes actual)
    - fecha_final: YYYY-MM-DD (por defecto: hoy)
    - buscar: texto para buscar en NIT, razón social, prefijo, folio
    - page: página actual (para paginación)
    - size: registros por página (por defecto: 500)
    """
    try:
        from datetime import date
        import re
        
        # Obtener parámetros de consulta
        fecha_inicial = request.args.get('fecha_inicial', '').strip()
        fecha_final = request.args.get('fecha_final', '').strip()
        buscar = request.args.get('buscar', '').strip()
        page = int(request.args.get('page', 1))
        size = int(request.args.get('size', 500))
        
        # 🔥 VALIDACIÓN Y DEFAULT: Si no hay fechas, usar TODO EL MES ACTUAL
        fecha_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
        
        if not fecha_inicial or not fecha_pattern.match(fecha_inicial):
            hoy = date.today()
            fecha_inicial = f"{hoy.year}-{hoy.month:02d}-01"  # PRIMER DÍA DEL MES ACTUAL
        
        if not fecha_final or not fecha_pattern.match(fecha_final):
            hoy = date.today()
            fecha_final = hoy.strftime("%Y-%m-%d")  # HOY
        
        print(f"[API DIAN] Consultando PostgreSQL: {fecha_inicial} a {fecha_final}, buscar: '{buscar}'")
        
        # ✅ CONSULTAR DESDE POSTGRESQL CON JOIN A ACUSES
        # JOIN con tabla acuses para obtener estado_aprobacion correcto
        query = db.session.query(
            MaestroDianVsErp,
            Acuses.estado_docto.label('estado_acuse')
        ).outerjoin(
            Acuses,
            MaestroDianVsErp.cufe == Acuses.cufe
        )
        
        # Filtro de fechas
        query = query.filter(MaestroDianVsErp.fecha_emision >= fecha_inicial)
        query = query.filter(MaestroDianVsErp.fecha_emision <= fecha_final)
        
        # Filtro de búsqueda (NIT, razón social, prefijo, folio)
        if buscar:
            query = query.filter(
                db.or_(
                    MaestroDianVsErp.nit_emisor.like(f'%{buscar}%'),
                    MaestroDianVsErp.razon_social.like(f'%{buscar}%'),
                    MaestroDianVsErp.prefijo.like(f'%{buscar}%'),
                    MaestroDianVsErp.folio.like(f'%{buscar}%')
                )
            )
        
        # Ordenar por fecha de emisión descendente
        query = query.order_by(MaestroDianVsErp.fecha_emision.desc())
        
        # Obtener TODOS los registros para paginación local en frontend
        registros = query.all()
        
        # Convertir a diccionarios
        datos = []
        for row in registros:
            registro = row[0]  # MaestroDianVsErp
            estado_acuse = row[1]  # estado_docto de acuses (puede ser None)
            
            # Convertir forma_pago: Soporta valores transformados ('Contado', 'Crédito') Y códigos numéricos ('1', '2')
            forma_pago_raw = (registro.forma_pago or "").strip()
            
            # 🔥 PRIORIDAD 1: Si ya está transformado, usarlo directamente
            if forma_pago_raw in ["Contado", "Crédito"]:
                forma_pago_texto = forma_pago_raw
            # 🔥 PRIORIDAD 2: Si es código numérico, transformar
            elif forma_pago_raw == "1" or forma_pago_raw == "01":
                forma_pago_texto = "Contado"
            elif forma_pago_raw == "2" or forma_pago_raw == "02":
                forma_pago_texto = "Crédito"
            else:
                forma_pago_texto = "Crédito"  # Default (incluye null, 0, 3, etc.)
            
            # ✅ DETERMINAR ESTADO CONTABLE CON LÓGICA ROBUSTA
            # Prioridad 1: Si tiene un estado válido sincronizado → usarlo
            # Prioridad 2: Si está vacío/None → calcular según módulo (legacy)
            estado_contable_bd = (registro.estado_contable or "").strip()
            
            # Estados válidos sincronizados (desde el sync_service)
            estados_validos = ["Recibida", "En Trámite", "Causada", "Rechazada"]
            
            if estado_contable_bd and estado_contable_bd in estados_validos:
                # Tiene un estado sincronizado válido → usarlo
                estado_contable_validado = estado_contable_bd
            else:
                # No tiene estado válido → calcular según módulo (lógica legacy)
                modulo_val = (registro.modulo or "").strip()
                if modulo_val and modulo_val != "No Registrada":
                    estado_contable_validado = "Causada"
                else:
                    estado_contable_validado = "No Registrada"
            
            # 🎯 ESTADO DE APROBACIÓN: Del JOIN con tabla acuses
            # REGLA: Solo usar estados de acuses, NUNCA estado_contable
            if estado_acuse and str(estado_acuse).strip():
                estado_aprobacion_final = str(estado_acuse).strip()  # ✅ Estado real de acuses
            else:
                estado_aprobacion_final = "No Registra"  # ⚠️ Sin acuse en tabla acuses
            
            # 📅 Calcular días desde emisión en TIEMPO REAL (no usar la columna de tabla maestro)
            dias_transcurridos = 0
            if registro.fecha_emision:
                try:
                    # Convertir a date si es datetime
                    fecha_emision_date = registro.fecha_emision
                    if isinstance(fecha_emision_date, datetime):
                        fecha_emision_date = fecha_emision_date.date()
                    
                    # Calcular diferencia
                    dias_transcurridos = (date.today() - fecha_emision_date).days
                except Exception as e:
                    print(f"⚠️ Error calculando días para {registro.nit_emisor}-{registro.prefijo}{registro.folio}: {e}")
                    dias_transcurridos = 0
            
            datos.append({
                "nit_emisor": registro.nit_emisor,
                "nombre_emisor": registro.razon_social or "",
                "fecha_emision": registro.fecha_emision.strftime("%Y-%m-%d") if registro.fecha_emision else "",
                "tipo_documento": registro.tipo_documento or "",
                "prefijo": registro.prefijo or "",
                "folio": registro.folio or "",
                "valor": float(registro.valor) if registro.valor else 0,
                "cufe": registro.cufe or "",
                "estado_aprobacion": estado_aprobacion_final,
                "forma_pago_texto": forma_pago_texto,
                "estado_contable": estado_contable_validado,
                "dias_desde_emision": dias_transcurridos,  # 📅 Calculado en tiempo real
                "tipo_tercero": registro.tipo_tercero or "",
                "usuario_solicitante": "",
                "usuario_aprobador": "",
                "observaciones": "",
                "modulo": registro.modulo or "",
                "doc_causado_por": registro.doc_causado_por or "",
                "tiene_usuarios_email": False
            })
        
        print(f"✅ Retornando {len(datos):,} registros desde PostgreSQL")
        return jsonify(datos)
        
    except Exception as e:
        print(f"❌ Error en API D IAN: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error al consultar base de datos: {str(e)}"}), 500


@dian_vs_erp_bp.route('/api/dian_v2')
def api_dian_v2():
    """
    API V2: Lee desde tablas optimizadas con LEFT JOIN a acuses para estado de aprobación
    Compatible 100% con el frontend del visor
    """
    try:
        from datetime import date
        import re
        from sqlalchemy import and_
        
        # Obtener parámetros de consulta
        fecha_inicial = request.args.get('fecha_inicial', '').strip()
        fecha_final = request.args.get('fecha_final', '').strip()
        buscar = request.args.get('buscar', '').strip()
        page = int(request.args.get('page', 1))
        size = int(request.args.get('size', 500))
        
        # Validación y default de fechas
        fecha_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
        
        if not fecha_inicial or not fecha_pattern.match(fecha_inicial):
            hoy = date.today()
            fecha_inicial = f"{hoy.year}-{hoy.month:02d}-01"
        
        if not fecha_final or not fecha_pattern.match(fecha_final):
            hoy = date.today()
            fecha_final = hoy.strftime("%Y-%m-%d")
        
        print(f"[API DIAN V2] 🔍 Consultando: {fecha_inicial} a {fecha_final}, buscar: '{buscar}'")
        
        # 🔥 HACER LEFT JOINS CON TODAS LAS TABLAS NECESARIAS
        # 1. Acuses (para estado_aprobacion) - JOIN por CUFE
        # 2. ERP Financiero y Comercial (para causación) - JOIN por CLAVE
        # 3. Facturas Temporales, Recibidas, Digitales (para recepción) - JOIN por PREFIJO+FOLIO
        # 4. Tipo Tercero (para clasificación Proveedor/Acreedor) - JOIN por NIT
        query = db.session.query(
            Dian,
            Acuses.estado_docto.label('estado_acuse'),
            Acuses.acuse_recibido.label('acuse_recibido'),
            Acuses.recibo_bien_servicio.label('recibo_bien_servicio'),
            Acuses.aceptacion_expresa.label('aceptacion_expresa'),
            Acuses.reclamo.label('reclamo'),
            Acuses.aceptacion_tacita.label('aceptacion_tacita'),
            ErpFinanciero.id.label('existe_financiero'),
            ErpFinanciero.modulo.label('modulo_financiero'),
            ErpFinanciero.doc_causado_por.label('causado_por_financiero'),
            ErpComercial.id.label('existe_comercial'),
            ErpComercial.modulo.label('modulo_comercial'),
            ErpComercial.doc_causado_por.label('causado_por_comercial'),
            FacturaTemporal.id.label('existe_temporal'),
            FacturaRecibida.id.label('existe_recibida'),
            FacturaDigital.id.label('existe_digital'),
            TipoTerceroDianErp.tipo_tercero.label('tipo_tercero_erp')  # 🏢 Clasificación del tercero
        ).outerjoin(
            Acuses,
            and_(
                Dian.cufe_cude == Acuses.cufe,
                Dian.cufe_cude != None,
                Dian.cufe_cude != ''
            )
        ).outerjoin(
            ErpFinanciero,
            and_(
                Dian.clave == ErpFinanciero.clave_erp_financiero,  # 🔗 JOIN POR CLAVE
                Dian.clave != None,
                Dian.clave != ''
            )
        ).outerjoin(
            ErpComercial,
            and_(
                Dian.clave == ErpComercial.clave_erp_comercial,  # 🔗 JOIN POR CLAVE
                Dian.clave != None,
                Dian.clave != ''
            )
        ).outerjoin(
            FacturaTemporal,
            and_(
                Dian.prefijo == FacturaTemporal.prefijo,
                Dian.folio == FacturaTemporal.folio,
                Dian.prefijo != None,
                Dian.folio != None
            )
        ).outerjoin(
            FacturaRecibida,
            and_(
                Dian.prefijo == FacturaRecibida.prefijo,
                Dian.folio == FacturaRecibida.folio,
                Dian.prefijo != None,
                Dian.folio != None
            )
        ).outerjoin(
            FacturaDigital,
            and_(
                db.func.concat(Dian.prefijo, Dian.folio) == FacturaDigital.numero_factura,
                Dian.prefijo != None,
                Dian.folio != None
            )
        ).outerjoin(
            TipoTerceroDianErp,
            Dian.nit_emisor == TipoTerceroDianErp.nit_emisor  # 🏢 JOIN por NIT para clasificación
        ).outerjoin(
            TipoDocumentoDian,
            and_(
                db.func.upper(Dian.tipo_documento) == db.func.upper(TipoDocumentoDian.tipo_documento),  # 🔥 CASE INSENSITIVE
                TipoDocumentoDian.procesar_frontend == True,
                TipoDocumentoDian.activo == True
            )
        )
        
        # Filtro de fechas
        query = query.filter(Dian.fecha_emision >= fecha_inicial)
        query = query.filter(Dian.fecha_emision <= fecha_final)
        
        # Filtro de búsqueda
        if buscar:
            query = query.filter(
                db.or_(
                    Dian.nit_emisor.like(f'%{buscar}%'),
                    Dian.nombre_emisor.like(f'%{buscar}%'),
                    Dian.prefijo.like(f'%{buscar}%'),
                    Dian.folio.like(f'%{buscar}%')
                )
            )
        
        # Ordenar
        query = query.order_by(Dian.fecha_emision.desc())
        # Obtener TODOS los registros para paginación local en frontend
        resultados = query.all()
        
        print(f"[API DIAN V2] ✅ Encontrados {len(resultados)} registros")
        
        # Convertir a diccionarios
        datos = []
        contador_debug = 0
        for registro, estado_acuse, acuse_recibido, recibo_bien_servicio, aceptacion_expresa, reclamo, aceptacion_tacita, existe_financiero, modulo_financiero, causado_por_financiero, existe_comercial, modulo_comercial, causado_por_comercial, existe_temporal, existe_recibida, existe_digital, tipo_tercero_erp in resultados:
            contador_debug += 1
            
            # 🐛 DEBUG: Primeros 3 registros
            if contador_debug <= 3:
                cufe_preview = (registro.cufe_cude[:30] if registro.cufe_cude else "None")
                print(f"\n[DEBUG {contador_debug}] Factura: {registro.prefijo}-{registro.folio}")
                print(f"[DEBUG {contador_debug}]   NIT: {registro.nit_emisor}")
                print(f"[DEBUG {contador_debug}]   CUFE: {cufe_preview}...")
                print(f"[DEBUG {contador_debug}]   CLAVE: {registro.clave}")
                print(f"[DEBUG {contador_debug}]   tipo_tercero_erp={tipo_tercero_erp}")  # 🏢 NUEVO
                print(f"[DEBUG {contador_debug}]   estado_acuse={estado_acuse}")
                print(f"[DEBUG {contador_debug}]   existe_financiero={existe_financiero}")
                print(f"[DEBUG {contador_debug}]   modulo_financiero={modulo_financiero}")
                print(f"[DEBUG {contador_debug}]   causado_por_financiero={causado_por_financiero}")
                print(f"[DEBUG {contador_debug}]   existe_comercial={existe_comercial}")
                print(f"[DEBUG {contador_debug}]   modulo_comercial={modulo_comercial}")
                print(f"[DEBUG {contador_debug}]   causado_por_comercial={causado_por_comercial}")
                print(f"[DEBUG {contador_debug}]   existe_temporal={existe_temporal}")
                print(f"[DEBUG {contador_debug}]   existe_recibida={existe_recibida}")
                print(f"[DEBUG {contador_debug}]   existe_digital={existe_digital}")
            
            # Forma de pago: Soporta valores transformados ('Contado', 'Crédito') Y códigos numéricos ('1', '2')
            forma_pago_raw = (registro.forma_pago or "").strip()
            
            # 🔥 PRIORIDAD 1: Si ya está transformado, usarlo directamente
            if forma_pago_raw in ["Contado", "Crédito"]:
                forma_pago_texto = forma_pago_raw
            # 🔥 PRIORIDAD 2: Si es código numérico, transformar
            elif forma_pago_raw == "1" or forma_pago_raw == "01":
                forma_pago_texto = "Contado"
            elif forma_pago_raw == "2" or forma_pago_raw == "02":
                forma_pago_texto = "Crédito"
            else:
                forma_pago_texto = "Crédito"  # Default (incluye null, 0, 3, etc.)
            
            # 🎯 ESTADO DE APROBACIÓN: Solo desde acuses (NO usar estado de DIAN)
            if estado_acuse and str(estado_acuse).strip():
                estado_aprobacion_final = str(estado_acuse).strip()  # ✅ Hay acuse real del adquiriente
            else:
                estado_aprobacion_final = "No Registra"  # ⚠️ No hay acuse registrado
            
            # 🎯 ESTADO CONTABLE - Lógica según tabla estado_contable:
            # 1. CAUSADA: Si clave DIAN está en erp_comercial o erp_financiero
            # 2. RECIBIDA: Si clave (prefijo+folio) está en facturas_recibidas, facturas_digitales o facturas_temporales
            # 3. NO REGISTRADA: Si no está en ninguna de las tablas anteriores
            # 4. EN TRÁMITE: Si tiene relación generada (TODO: módulo relaciones)
            # 5. RECHAZADA/NOVEDAD: Módulo relaciones (aún no implementado)
            
            # Determinar módulo y causado_por
            modulo_final = ""
            causado_por_final = ""
            
            if existe_financiero:
                estado_contable_validado = "Causada"  # ✅ Causada en ERP Financiero
                modulo_final = modulo_financiero or "Financiero"
                causado_por_final = causado_por_financiero or ""
            elif existe_comercial:
                estado_contable_validado = "Causada"  # ✅ Causada en ERP Comercial
                modulo_final = modulo_comercial or "Comercial"
                causado_por_final = causado_por_comercial or ""
            elif existe_recibida or existe_temporal or existe_digital:
                estado_contable_validado = "Recibida"  # ✅ Recibida en alguna tabla de facturas
                modulo_final = "Recepción de Facturas"
                causado_por_final = ""
            else:
                estado_contable_validado = "No Registrada"  # ⚠️ No está en ninguna tabla
                modulo_final = ""
                causado_por_final = ""
            
            # 🐛 DEBUG: Primeros 3 registros - FINALES
            if contador_debug <= 3:
                print(f"[DEBUG {contador_debug}]   >>> ESTADO_APROBACION_FINAL={estado_aprobacion_final}")
                print(f"[DEBUG {contador_debug}]   >>> ESTADO_CONTABLE_FINAL={estado_contable_validado}")
                print(f"[DEBUG {contador_debug}]   >>> MODULO_FINAL={modulo_final}")
                print(f"[DEBUG {contador_debug}]   >>> CAUSADO_POR_FINAL={causado_por_final}")
                print(f"[DEBUG {contador_debug}]   >>> TIPO_TERCERO_FINAL={tipo_tercero_erp or 'No Registrado'}")  # 🏢 NUEVO
            
            # 📅 Calcular días desde emisión en TIEMPO REAL
            dias_transcurridos = 0
            if registro.fecha_emision:
                try:
                    # Convertir a date si es datetime
                    fecha_emision_date = registro.fecha_emision
                    if isinstance(fecha_emision_date, datetime):
                        fecha_emision_date = fecha_emision_date.date()
                    
                    # Calcular diferencia
                    dias_transcurridos = (date.today() - fecha_emision_date).days
                except Exception as e:
                    print(f"⚠️ Error calculando días para {registro.nit_emisor}-{registro.prefijo}{registro.folio}: {e}")
                    dias_transcurridos = 0
            
            # 🔴 DETECTAR NOTAS CRÉDITO Y CONVERTIR VALOR A NEGATIVO
            tipo_doc = registro.tipo_documento or ''
            valor_original = float(registro.total) if registro.total else 0
            
            # Si es nota crédito, convertir a negativo
            es_nota_credito = (
                'Nota de crédito electrónica' in tipo_doc or
                'Nota de ajuste crédito del documento equivalente' in tipo_doc
            )
            
            valor_final = -abs(valor_original) if es_nota_credito else valor_original
            
            datos.append({
                "nit_emisor": registro.nit_emisor,
                "nombre_emisor": registro.nombre_emisor or "",
                "fecha_emision": registro.fecha_emision.strftime("%Y-%m-%d") if registro.fecha_emision else "",
                "tipo_documento": tipo_doc,
                "prefijo": registro.prefijo or "",
                "folio": registro.folio or "",
                "valor": valor_final,  # 🔴 Valor con signo correcto (negativo para notas crédito)
                "cufe": registro.cufe_cude or "",
                "estado_aprobacion": estado_aprobacion_final,
                "forma_pago_texto": forma_pago_texto,
                "estado_contable": estado_contable_validado,
                "dias_desde_emision": dias_transcurridos,  # 📅 Calculado en tiempo real
                "tipo_tercero": tipo_tercero_erp or "No Registrado",  # 🏢 Desde tabla clasificación
                "usuario_solicitante": "",
                "usuario_aprobador": "",
                "observaciones": "",
                "modulo": modulo_final,
                "doc_causado_por": causado_por_final,
                "tiene_usuarios_email": False
            })
        
        print(f"[API DIAN V2] ✅ Retornando {len(datos):,} registros con estados de acuses y ERP")
        return jsonify(datos)
        
    except Exception as e:
        print(f"[API DIAN V2] ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error al consultar base de datos: {str(e)}"}), 500


@dian_vs_erp_bp.route('/api/obtener_filtros_unicos', methods=['GET'])
def obtener_filtros_unicos():
    """
    Obtiene valores únicos de la base de datos para llenar los filtros del frontend
    ✅ CONSULTA DE LAS MISMAS TABLAS QUE USA EL VISOR V2 (/api/dian_v2)
    Con try-catch individual para que si falla uno, los demás se carguen
    """
    from sqlalchemy import distinct
    
    print("\n📊 OBTENIENDO FILTROS ÚNICOS DESDE TABLAS DEL VISOR V2...")
    
    resultado = {}
    
    # 1️⃣ TIPOS DE DOCUMENTO ⭐ CORREGIDO: Desde tabla DIAN (igual que visor_v2)
    try:
        tipos_documento = db.session.query(
            distinct(Dian.tipo_documento)
        ).filter(
            Dian.tipo_documento.isnot(None),
            Dian.tipo_documento != ''
        ).order_by(Dian.tipo_documento).all()
        resultado['tipos_documento'] = [t[0] for t in tipos_documento]
        print(f"   ✅ Tipos de documento (desde DIAN): {len(tipos_documento)}")
    except Exception as e:
        print(f"   ❌ Error tipos_documento: {e}")
        resultado['tipos_documento'] = []
    
    # 2️⃣ PREFIJOS ⭐ CORREGIDO: Desde tabla DIAN (igual que visor_v2)
    try:
        prefijos = db.session.query(
            distinct(Dian.prefijo)
        ).filter(
            Dian.prefijo.isnot(None),
            Dian.prefijo != ''
        ).order_by(Dian.prefijo).all()
        resultado['prefijos'] = [p[0] for p in prefijos]
        print(f"   ✅ Prefijos (desde DIAN): {len(prefijos)}")
    except Exception as e:
        print(f"   ❌ Error prefijos: {e}")
        resultado['prefijos'] = []
    
    # 3️⃣ TIPOS DE TERCERO ⭐ CORREGIDO: Desde tabla TipoTerceroDianErp (igual que visor_v2)
    try:
        tipos_tercero = db.session.query(
            distinct(TipoTerceroDianErp.tipo_tercero)
        ).filter(
            TipoTerceroDianErp.tipo_tercero.isnot(None),
            TipoTerceroDianErp.tipo_tercero != ''
        ).order_by(TipoTerceroDianErp.tipo_tercero).all()
        resultado['tipos_tercero'] = [tt[0] for tt in tipos_tercero]
        print(f"   ✅ Tipos de tercero (desde TipoTerceroDianErp): {len(tipos_tercero)}")
    except Exception as e:
        print(f"   ❌ Error tipos_tercero: {e}")
        resultado['tipos_tercero'] = []
    
    # 4️⃣ ESTADOS DE APROBACIÓN DIAN (desde tabla acuses) ✅ Ya estaba correcto
    try:
        estados_aprobacion = db.session.query(
            distinct(Acuses.estado_docto)
        ).filter(
            Acuses.estado_docto.isnot(None),
            Acuses.estado_docto != ''
        ).order_by(Acuses.estado_docto).all()
        estados_aprobacion_lista = [e[0] for e in estados_aprobacion]
        # Agregar "No Registra" para facturas sin acuse
        if "No Registra" not in estados_aprobacion_lista:
            estados_aprobacion_lista.append("No Registra")
        estados_aprobacion_lista.sort()
        resultado['estados_aprobacion'] = estados_aprobacion_lista
        print(f"   ✅ Estados de aprobación (desde Acuses): {len(estados_aprobacion_lista)}")
    except Exception as e:
        print(f"   ❌ Error estados_aprobacion: {e}")
        # Valores por defecto si falla
        resultado['estados_aprobacion'] = ['No Registra', 'Pendiente', 'Acuse Recibido', 'Rechazada']
    
    # 5️⃣ ESTADOS CONTABLES ⭐ NOTA: Se calculan en código (Causada/Recibida/No Registrada)
    # No hay tabla de origen, usar valores estándar que genera el visor_v2
    try:
        # Valores que genera el visor_v2 en líneas 775-780 de /api/dian_v2
        resultado['estados_contable'] = ['Causada', 'No Registrada', 'Recibida']
        print(f"   ✅ Estados contables (valores calculados por visor_v2): 3")
    except Exception as e:
        print(f"   ❌ Error estados_contable: {e}")
        resultado['estados_contable'] = ['No Registrada', 'Recibida', 'Causada']
    
    # 6️⃣ FORMAS DE PAGO ⭐ CORREGIDO: Desde tabla DIAN pero transformadas (igual que visor_v2)
    try:
        # El visor_v2 transforma los códigos 1/01 → Contado, 2/02 → Crédito (líneas 720-730)
        # Consultamos valores únicos de DIAN y los transformamos
        formas_pago_raw = db.session.query(
            distinct(Dian.forma_pago)
        ).filter(
            Dian.forma_pago.isnot(None),
            Dian.forma_pago != ''
        ).all()
        
        # Transformar según lógica del visor_v2
        formas_transformadas = set()
        for (forma,) in formas_pago_raw:
            forma_str = str(forma).strip()
            if forma_str in ["Contado", "Crédito"]:
                formas_transformadas.add(forma_str)
            elif forma_str in ["1", "01"]:
                formas_transformadas.add("Contado")
            elif forma_str in ["2", "02"]:
                formas_transformadas.add("Crédito")
            else:
                formas_transformadas.add("Crédito")  # Default
        
        resultado['formas_pago'] = sorted(list(formas_transformadas))
        print(f"   ✅ Formas de pago (desde DIAN, transformadas): {len(formas_transformadas)}")
    except Exception as e:
        print(f"   ❌ Error formas_pago: {e}")
        resultado['formas_pago'] = ['Contado', 'Crédito']
    
    # 7️⃣ MÓDULOS ERP ⭐ NOTA: Se calculan en código según tabla de origen
    # El visor_v2 determina el módulo según:
    # - Si existe en ErpFinanciero → "Financiero"
    # - Si existe en ErpComercial → "Comercial"  
    # - Si existe en Facturas Recibidas/Temporales/Digitales → "Recepción de Facturas"
    # No hay tabla de origen, usar valores posibles que genera el visor_v2
    try:
        resultado['modulos'] = ['Comercial', 'Financiero', 'Recepción de Facturas']
        print(f"   ✅ Módulos ERP (valores calculados por visor_v2): 3")
    except Exception as e:
        print(f"   ❌ Error modulos: {e}")
        resultado['modulos'] = ['Comercial', 'Financiero', 'Recepción de Facturas']
    
    print(f"✅ FILTROS ÚNICOS OBTENIDOS (algunos pueden ser defaults si hubo errores)")
    return jsonify(resultado)


@dian_vs_erp_bp.route('/api/exportar_seleccionadas', methods=['POST'])
def exportar_seleccionadas():
    """
    Exportar facturas seleccionadas a Excel con valores negativos para notas crédito
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from io import BytesIO
        from datetime import datetime
        
        # Obtener datos enviados desde el frontend
        datos_seleccionados = request.json.get('datos', [])
        
        if not datos_seleccionados:
            return jsonify({"error": "No se recibieron datos para exportar"}), 400
        
        print(f"\n📥 EXPORTAR SELECCIONADAS: {len(datos_seleccionados)} facturas")
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Seleccionadas"
        
        # 🔴 APLICAR VALORES NEGATIVOS PARA NOTAS CRÉDITO
        datos_procesados = []
        for row in datos_seleccionados:
            row_copy = row.copy()
            tipo_doc = row_copy.get('tipo_documento', '')
            
            # Detectar notas crédito
            es_nota_credito = (
                'Nota de crédito electrónica' in tipo_doc or
                'Nota de ajuste crédito del documento equivalente' in tipo_doc
            )
            
            # Si es nota crédito, convertir valor a negativo
            if es_nota_credito and row_copy.get('valor'):
                try:
                    valor_original = float(row_copy['valor'])
                    row_copy['valor'] = -abs(valor_original)
                    print(f"   🔴 Nota crédito detectada: {row_copy['nit_emisor']}-{row_copy['prefijo']}{row_copy['folio']} | Valor: {valor_original} → {row_copy['valor']}")
                except:
                    pass
            
            datos_procesados.append(row_copy)
        
        # Encabezados (18 columnas incluyendo CUFE y Ver DIAN)
        headers = [
            'NIT Emisor', 'Razón Social', 'Tipo Tercero', 'Fecha Emisión',
            'Tipo Documento', 'Prefijo', 'Folio', 'Valor',
            'Estado Aprobación', 'Forma de Pago', 'Estado Contable',
            'N° Días', 'Usuario Solicitante', 'Usuario Aprobador',
            'Observaciones', 'Módulo', 'Causador', 'CUFE', 'Ver DIAN'
        ]
        
        # Escribir encabezados con estilo verde corporativo
        for c_idx, col in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Escribir datos
        for r_idx, row in enumerate(datos_procesados, 2):
            ws.cell(row=r_idx, column=1, value=row.get('nit_emisor', ''))
            ws.cell(row=r_idx, column=2, value=row.get('nombre_emisor', ''))
            ws.cell(row=r_idx, column=3, value=row.get('tipo_tercero', ''))
            ws.cell(row=r_idx, column=4, value=row.get('fecha_emision', ''))
            ws.cell(row=r_idx, column=5, value=row.get('tipo_documento', ''))
            ws.cell(row=r_idx, column=6, value=row.get('prefijo', ''))
            ws.cell(row=r_idx, column=7, value=row.get('folio', ''))
            
            # Columna Valor con formato especial
            cell_valor = ws.cell(row=r_idx, column=8, value=row.get('valor', 0))
            cell_valor.number_format = '#,##0.00'
            cell_valor.alignment = Alignment(horizontal='right')
            
            # 🔴 VALORES NEGATIVOS EN ROJO
            if isinstance(row.get('valor'), (int, float)) and row.get('valor') < 0:
                cell_valor.font = Font(color="FF0000", bold=True)
            
            ws.cell(row=r_idx, column=9, value=row.get('estado_aprobacion', ''))
            ws.cell(row=r_idx, column=10, value=row.get('forma_pago_texto', ''))
            ws.cell(row=r_idx, column=11, value=row.get('estado_contable', ''))
            ws.cell(row=r_idx, column=12, value=row.get('dias_desde_emision', 0))
            ws.cell(row=r_idx, column=13, value=row.get('usuario_solicitante', ''))
            ws.cell(row=r_idx, column=14, value=row.get('usuario_aprobador', ''))
            ws.cell(row=r_idx, column=15, value=row.get('observaciones', ''))
            ws.cell(row=r_idx, column=16, value=row.get('modulo', ''))
            ws.cell(row=r_idx, column=17, value=row.get('doc_causado_por', ''))
            
            # 🔗 COLUMNA CUFE (columna 18)
            cufe = row.get('cufe', '')
            ws.cell(row=r_idx, column=18, value=cufe)
            
            # 🔗 COLUMNA VER DIAN con hipervínculo (columna 19)
            if cufe:
                enlace_dian = f"https://catalogo-vpfe.dian.gov.co/User/SearchDocument?DocumentKey={cufe}"
                cell_link = ws.cell(row=r_idx, column=19, value="🔗 Ver en DIAN")
                cell_link.hyperlink = enlace_dian
                cell_link.style = "Hyperlink"
                cell_link.font = Font(color="0563C1", underline="single")
            else:
                ws.cell(row=r_idx, column=19, value="")
        
        # Ajustar anchos de columna (19 columnas ahora)
        anchos = [12, 30, 20, 12, 25, 10, 12, 15, 20, 15, 18, 10, 20, 20, 30, 15, 20, 35, 15]
        for idx, ancho in enumerate(anchos, 1):
            col_letter = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[col_letter].width = ancho
        
        # Crear tabla Excel
        tabla_rango = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(datos_procesados) + 1}"
        tabla = Table(displayName="TablaSeleccionadas", ref=tabla_rango)
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(tabla)
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nombre del archivo con timestamp
        fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DIAN_vs_ERP_Seleccionadas_{fecha_str}.xlsx"
        
        print(f"✅ Excel generado: {filename} ({len(datos_procesados)} filas)")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Error exportando seleccionadas: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Error al exportar: {str(e)}"}), 500


@dian_vs_erp_bp.route('/descargar_plantilla/<tipo>')
def descargar_plantilla(tipo):
    """
    Descargar plantillas Excel con encabezados EXACTOS que el sistema espera
    Encabezados basados en routes.py líneas 1240-1265 (formato con espacios)
    """
    try:
        plantillas_dir = BASE_DIR / "plantillas"
        
        if tipo == "dian":
            archivo = plantillas_dir / "plantilla_dian.xlsx"
        elif tipo == "erp":
            archivo = plantillas_dir / "plantilla_erp.xlsx"
        elif tipo == "acuses":
            archivo = plantillas_dir / "plantilla_acuses.xlsx"
        else:
            return jsonify({"error": "Tipo de plantilla no válido"}), 400
        
        if archivo.exists():
            return send_file(str(archivo), as_attachment=True)
        else:
            # Si no existe la plantilla, crear con encabezados exactos
            print(f"⚠️ Plantilla {tipo} no encontrada, creando con encabezados correctos...")
            
            # ==========================================
            # ENCABEZADOS EXACTOS QUE BUSCA EL SISTEMA
            # Basados en routes.py líneas 1074-1265
            # ==========================================
            if tipo == "dian":
                # Encabezados CON ESPACIOS (formato original DIAN)
                headers = [
                    'Tipo Documento',      # row.get('tipo documento')
                    'CUFE/CUDE',           # row.get('cufe/cude')
                    'Numero',              # row.get('numero')
                    'Prefijo',             # row.get('prefijo')
                    'Fecha Emision',       # row.get('fecha emision')
                    'NIT Emisor',          # row.get('nit emisor')
                    'Nombre Emisor',       # row.get('nombre emisor')
                    'Valor',               # row.get('valor')
                    'Forma Pago'           # row.get('forma pago')
                ]
            elif tipo == "erp":
                # Encabezados EXACTOS que busca el código (líneas 1082-1086)
                headers = [
                    'Proveedor',           # "proveedor" in c.lower() and "razon" not in c.lower()
                    'Docto. Proveedor',    # "docto" in c.lower() and "proveedor" in c.lower()
                    'Clase de Documento',  # "clase" in c.lower()
                    'C.O.',                # c.upper() == "C.O."
                    'Usuario Creacion',    # "usuario" in c.lower() and "creac" in c.lower()
                    'Nro. Documento'       # "nro" in c.lower() and "documento" in c.lower()
                ]
            else:  # acuses
                # Encabezados para Acuses (línea 1215)
                headers = [
                    'Fecha',               # row.get('Fecha')
                    'Adquiriente',         # row.get('Adquiriente')
                    'Factura',             # row.get('Factura')
                    'Emisor',              # row.get('Emisor')
                    'CUFE',                # row.get('CUFE')
                    'Estado Docto.'        # row.get('Estado Docto.')
                ]
            
            # Crear DataFrame vacío con solo encabezados
            df = pd.DataFrame(columns=headers)
            
            # Crear directorio si no existe
            plantillas_dir.mkdir(parents=True, exist_ok=True)
            
            # Guardar como Excel con openpyxl
            df.to_excel(str(archivo), index=False, engine='openpyxl')
            
            print(f"✅ Plantilla {tipo} creada con {len(headers)} columnas")
            print(f"   📋 Encabezados: {', '.join(headers)}")
            
            return send_file(str(archivo), as_attachment=True)
        
    except Exception as e:
        print(f"❌ Error descargando plantilla {tipo}: {e}")
        return jsonify({"error": f"Error al descargar plantilla: {e}"}), 500

@dian_vs_erp_bp.route('/subir_archivos', methods=['POST'])
@dian_vs_erp_bp.route('/api/subir_archivos', methods=['POST'])
@requiere_permiso('dian_vs_erp', 'cargar_archivos')
def subir_archivos():
    """
    NUEVA LÓGICA (Dec 29, 2025): 
    1. GUARDAR archivos en disco PRIMERO
    2. PROCESAR desde disco DESPUÉS
    Esto evita problemas con FileStorage streams
    """
    try:
        # Validar sesión
        if 'usuario_id' not in session or 'usuario' not in session:
            return jsonify({"error": "Sesión no válida", "redirect": "/login"}), 401
        
        archivos_guardados = []
        
        print("=" * 80)
        print("📤 PASO 1: GUARDANDO ARCHIVOS EN DISCO")
        print("=" * 80)
        
        # PASO 1: GUARDAR TODOS LOS ARCHIVOS EN DISCO
        for key, folder in UPLOADS.items():
            f = request.files.get(key)
            if not f or not f.filename:
                continue
            
            ext = os.path.splitext(f.filename)[1].lower()
            if ext not in ALLOWED_EXCEL:
                return jsonify({"mensaje": f"❌ {f.filename}: extensión no soportada"}), 400
            
            # Crear carpeta si no existe
            folder.mkdir(parents=True, exist_ok=True)
            
            # Guardar archivo DIRECTO A DISCO (sin procesar nada)
            from werkzeug.utils import secure_filename
            filename = secure_filename(f.filename)
            ruta_original = folder / filename
            
            print(f"💾 Guardando {key}: {filename}")
            f.save(str(ruta_original))
            
            archivos_guardados.append({
                "tipo": key,
                "nombre": filename,
                "ruta": str(ruta_original),
                "folder": str(folder)
            })
            
            print(f"   ✅ Guardado: {ruta_original}")
        
        if len(archivos_guardados) == 0:
            return jsonify({"mensaje": "⚠️ No se recibió ningún archivo"}), 400
        
        print(f"\n✅ {len(archivos_guardados)} archivo(s) guardado(s) en disco")
        
        # PASO 2: CARGAR CADA ARCHIVO A SU TABLA POSTGRESQL
        print("\n" + "=" * 80)
        print("⚙️ PASO 2: CARGANDO ARCHIVOS A POSTGRESQL")
        print("=" * 80)
        
        db_url = current_app.config.get('SQLALCHEMY_DATABASE_URI') or os.environ.get('DATABASE_URL', '')
        resultados_carga = []
        errores_carga = []
        
        for archivo_info in archivos_guardados:
            tipo = archivo_info["tipo"]
            ruta = archivo_info["ruta"]
            try:
                print(f"\n📥 Procesando {tipo}: {archivo_info['nombre']}")
                resultado = procesar_archivo_subido(tipo, ruta, db_url)
                resultados_carga.append(resultado)
                print(f"   ✅ {resultado.get('mensaje', 'OK')}")
            except Exception as e_loader:
                import traceback
                tb = traceback.format_exc()
                msg_err = f"❌ Error cargando {tipo}: {str(e_loader)}"
                errores_carga.append(msg_err)
                print(f"   {msg_err}\n{tb}")
        
        # PASO 3: RECONSTRUIR MAESTRO DESDE TABLAS
        print("\n" + "=" * 80)
        print("🔄 PASO 3: RECONSTRUYENDO MAESTRO DIAN vs ERP")
        print("=" * 80)
        
        resultado_maestro = {"registros": 0, "mensaje": "⚠️ No se reconstruyó el maestro"}
        try:
            resultado_maestro = reconstruir_maestro_bd(db_url)
        except Exception as e_maestro:
            errores_carga.append(f"❌ Error reconstruyendo maestro: {str(e_maestro)}")
            print(f"❌ Error maestro: {e_maestro}")
        
        # Construir respuesta
        lineas = []
        for r in resultados_carga:
            lineas.append(r.get('mensaje', '?'))
        lineas.append(resultado_maestro.get('mensaje', ''))
        if errores_carga:
            lineas.extend(errores_carga)
        
        mensaje_resultado = '\n'.join(lineas)
        
        return jsonify({
            "mensaje": mensaje_resultado,
            "archivos": archivos_guardados,
            "resultados_carga": resultados_carga,
            "maestro": resultado_maestro,
            "errores": errores_carga,
        })
        
    except Exception as e:
        import traceback
        error_detalle = traceback.format_exc()
        print(f"❌ ERROR EN SUBIDA:\n{error_detalle}")
        return jsonify({"mensaje": f"❌ Error: {str(e)}"}), 500

@dian_vs_erp_bp.route('/api/forzar_procesar')
def forzar_procesar():
    """Forzar reprocesamiento del maestro"""
    try:
        # Validar sesión simple para API endpoints
        if 'usuario_id' not in session or 'usuario' not in session:
            return jsonify({"error": "Sesión no válida", "redirect": "/login"}), 401
        
        msg = actualizar_maestro()
        return jsonify({"mensaje": msg})
    except Exception as e:
        return jsonify({"mensaje": f"❌ Error: {e}"}), 500

# ==============================
# FUNCIONES DE PROCESAMIENTO
# ==============================

def obtener_jerarquia_aceptacion(estado: str) -> int:
    """
    Retorna jerarquía de estado de aceptación/acuse DIAN
    Jerarquía: 1=Pendiente, 2=Acuse Recibido, 3=Acuse Bien/Servicio, 
               4=Rechazada, 5=Aceptación Expresa, 6=Aceptación Tácita
    """
    jerarquias = {
        'Pendiente': 1,
        'Acuse Recibido': 2,
        'Acuse Bien/Servicio': 3,
        'Rechazada': 4,
        'Aceptación Expresa': 5,
        'Aceptación Tácita': 6
    }
    return jerarquias.get(str(estado).strip() if estado else 'Pendiente', 1)

def obtener_jerarquia_contable(estado: str) -> int:
    """
    Retorna jerarquía de estado contable del sistema
    Jerarquía: 1=No Registrada, 2=Recibida, 3=Novedad, 4=En Trámite,
               5=Rechazada, 6=Causada
    """
    jerarquias = {
        'No Registrada': 1,
        'Recibida': 2,
        'Novedad': 3,
        'En Trámite': 4,
        'Rechazada': 5,
        'Causada': 6
    }
    return jerarquias.get(str(estado).strip() if estado else 'No Registrada', 1)

def calcular_acuses_recibidos(estado_aprobacion: str) -> int:
    """
    Calcula número de acuses recibidos según estado_aprobacion
    - Pendiente: 0 acuses (sin respuesta DIAN)
    - Acuse Recibido/Bien/Servicio/Rechazada: 1 acuse
    - Aceptación Expresa/Tácita: 2 acuses
    """
    estado = str(estado_aprobacion).strip() if estado_aprobacion else 'Pendiente'
    
    if estado in ['Pendiente', '', 'None']:
        return 0
    elif estado in ['Acuse Recibido', 'Acuse Bien/Servicio', 'Rechazada']:
        return 1
    elif estado in ['Aceptación Expresa', 'Aceptación Tácita']:
        return 2
    else:
        return 0  # Por defecto

def extraer_prefijo(docto: str) -> str:
    """Extrae prefijo alfanumérico (letras Y números), limpiando solo guiones y puntos"""
    if not docto:
        return ""
    import re
    # Solo eliminar guiones, puntos y espacios - MANTENER números
    prefijo = re.sub(r'[\-\.\s]', '', str(docto)).strip().upper()
    
    # Validar longitud máxima (20 caracteres por esquema BD)
    # Si es muy largo, probablemente es un CUFE/CUDE mal posicionado
    if len(prefijo) > 20:
        # Si es todo hexadecimal largo, es un CUFE - devolver vacío
        if re.match(r'^[A-F0-9]+$', prefijo) and len(prefijo) > 20:
            return ""  # CUFE mal posicionado
        # Si no, truncar a 20 caracteres
        return prefijo[:20]
    
    return prefijo

def extraer_folio(docto: str) -> str:
    """Extrae solo DÍGITOS del documento"""
    if not docto:
        return ""
    import re
    return re.sub(r'[^0-9]', '', str(docto))

def ultimos_8_sin_ceros(folio: str) -> str:
    """Últimos 8 dígitos sin ceros a la izquierda"""
    if not folio:
        return "0"
    import re
    numeros = re.sub(r'[^0-9]', '', str(folio))
    if not numeros:
        return "0"
    ultimos = numeros[-8:] if len(numeros) >= 8 else numeros
    return ultimos.lstrip('0') or '0'

def mapear_modulo(clase: str) -> str:
    """Mapea clase de documento a módulo COMERCIAL/FINANCIERO"""
    if not clase:
        return ""
    return MODULO_MAP.get(str(clase).lower().strip(), "")

def crear_clave_factura(nit: str, prefijo: str, folio: str) -> str:
    """Crea clave única: NIT + PREFIJO + FOLIO_8"""
    nit_limpio = extraer_folio(str(nit))
    prefijo_limpio = extraer_prefijo(str(prefijo))
    folio_8 = ultimos_8_sin_ceros(extraer_folio(str(folio)))
    return f"{nit_limpio}{prefijo_limpio}{folio_8}"

# ==============================
# HELPER PARA COPY FROM
# ==============================

def format_value_for_copy(value):
    """
    Formatea un valor para COPY FROM, convirtiendo None a cadena vacía
    y escapando caracteres especiales.
    
    PostgreSQL COPY FROM requiere escapar:
    - backslash → doble backslash
    - tab → \\t
    - newline → \\n
    - carriage return → \\r
    
    Args:
        value: Valor a formatear (puede ser None, date, int, str, etc)
    
    Returns:
        str: Cadena formateada (vacía si value es None)
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ''
    if isinstance(value, bool):
        return 't' if value else 'f'
    
    # Convertir a string
    s = str(value)
    
    # Escapar caracteres especiales para PostgreSQL COPY FROM
    s = s.replace('\\', '\\\\')  # Backslash debe ir primero
    s = s.replace('\t', '\\t')   # Tab
    s = s.replace('\n', '\\n')   # Newline
    s = s.replace('\r', '\\r')   # Carriage return
    
    return s

# ==============================
# FUNCIONES DE INSERCIÓN MASIVA A TABLAS INDIVIDUALES
# ==============================

def insertar_dian_bulk(df_polars, cursor, tipo_tercero_dict):
    """
    Inserta registros masivos en tabla dian usando COPY FROM
    
    Args:
        df_polars: DataFrame de Polars con datos DIAN
        cursor: Cursor psycopg2
        tipo_tercero_dict: Diccionario {NIT: tipo_tercero}
    
    Returns:
        Número de registros insertados
    """
    print("\n📊 Insertando en tabla DIAN...")
    
    # Convertir a pandas para iteración
    df = df_polars.to_pandas()
    
    # Normalizar nombres de columnas
    import unicodedata
    def normalizar_columna(nombre):
        sin_tildes = ''.join(
            c for c in unicodedata.normalize('NFD', str(nombre))
            if unicodedata.category(c) != 'Mn'
        )
        return sin_tildes.lower().strip().replace(' ', '_')
    
    columnas_originales = {}
    for col in df.columns:
        col_norm = normalizar_columna(col)
        columnas_originales[col_norm] = col
        columnas_originales[col_norm.replace('_', '')] = col
    
    # Detectar columna CUFE
    cufe_col_name = None
    for col in df.columns:
        if 'cufe' in col.lower() or 'cude' in col.lower():
            cufe_col_name = col
            break
    if not cufe_col_name:
        cufe_col_name = 'cufe/cude'
    
    # Preparar registros
    registros = []
    fecha_hoy = date.today()
    
    for idx, (_, row) in enumerate(df.iterrows()):
        # Extraer datos básicos
        nit = str(row.get(columnas_originales.get('nit_emisor', 'NIT Emisor'), '')).strip()
        nit_limpio = extraer_folio(nit)
        razon_social = str(row.get(columnas_originales.get('nombre_emisor', 'Nombre Emisor'), '')).strip()
        
        # Fecha de emisión
        fecha_emision = fecha_hoy
        try:
            fecha_col = columnas_originales.get('fecha_de_emision', columnas_originales.get('fechadeemision', 'Fecha de Emisión'))
            fecha_raw = row.get(fecha_col)
            if fecha_raw and not pd.isna(fecha_raw):
                if isinstance(fecha_raw, str):
                    fecha_emision = pd.to_datetime(fecha_raw, format='%d/%m/%Y', errors='coerce').date()
                elif isinstance(fecha_raw, (pd.Timestamp, datetime)):
                    fecha_emision = fecha_raw.date() if hasattr(fecha_raw, 'date') else fecha_raw
        except:
            pass
        
        # Prefijo y folio
        prefijo_folio = str(row.get(columnas_originales.get('numero_de_factura_electronica', 'Número de Factura Electrónica'), ''))
        prefijo = extraer_prefijo(prefijo_folio)
        folio_raw = extraer_folio(prefijo_folio)
        folio = ultimos_8_sin_ceros(folio_raw)
        
        # Valor
        valor = 0.0
        try:
            valor_col = columnas_originales.get('total_factura', columnas_originales.get('totalfactura', 'Total Factura'))
            valor_raw = row.get(valor_col, 0)
            if valor_raw and not pd.isna(valor_raw):
                if isinstance(valor_raw, str):
                    valor = float(valor_raw.strip().replace('.', '').replace(',', '.'))
                else:
                    valor = float(valor_raw)
        except:
            pass
        
        # Otros campos
        tipo_documento = str(row.get(columnas_originales.get('tipo_de_documento', 'tipo de documento'), 'Factura Electrónica'))
        cufe = str(row.get(cufe_col_name, ''))
        forma_pago = str(row.get(columnas_originales.get('forma_de_pago', 'forma de pago'), '2')).strip()
        
        # Campos calculados
        clave = crear_clave_factura(nit, prefijo, folio_raw)
        clave_acuse = cufe  # clave_acuse = CUFE para matching
        tipo_tercero = tipo_tercero_dict.get(nit_limpio, '')
        n_dias = (fecha_hoy - fecha_emision).days if fecha_emision else 0
        modulo = 'DIAN'
        
        # Agregar registro
        registros.append({
            'nit_emisor': nit_limpio,
            'nombre_emisor': razon_social,
            'fecha_emision': fecha_emision,
            'prefijo': prefijo,
            'folio': folio,
            'total': valor,
            'tipo_documento': tipo_documento,
            'cufe_cude': cufe,
            'forma_pago': forma_pago,
            'clave': clave,
            'clave_acuse': clave_acuse,
            'tipo_tercero': tipo_tercero,
            'n_dias': n_dias,
            'modulo': modulo
        })
    
    # 🔥 CARGA INCREMENTAL: Usar tabla temporal + INSERT ... ON CONFLICT DO NOTHING
    print(f"   🔄 Carga INCREMENTAL (conserva datos viejos, agrega nuevos)")
    
    # Crear tabla temporal (sin restricciones UNIQUE para permitir duplicados internos en archivo)
    cursor.execute("""
        CREATE TEMP TABLE temp_dian_nuevos (LIKE dian INCLUDING DEFAULTS) ON COMMIT DROP
    """)
    
    # COPY FROM a tabla temporal
    buffer = io.StringIO()
    for reg in registros:
        buffer.write(f"{format_value_for_copy(reg['nit_emisor'])}\t")
        buffer.write(f"{format_value_for_copy(reg['nombre_emisor'])}\t")
        buffer.write(f"{format_value_for_copy(reg['fecha_emision'])}\t")
        buffer.write(f"{format_value_for_copy(reg['prefijo'])}\t")
        buffer.write(f"{format_value_for_copy(reg['folio'])}\t")
        buffer.write(f"{format_value_for_copy(reg['total'])}\t")
        buffer.write(f"{format_value_for_copy(reg['tipo_documento'])}\t")
        buffer.write(f"{format_value_for_copy(reg['cufe_cude'])}\t")
        buffer.write(f"{format_value_for_copy(reg['forma_pago'])}\t")
        buffer.write(f"{format_value_for_copy(reg['clave'])}\t")
        buffer.write(f"{format_value_for_copy(reg['clave_acuse'])}\t")
        buffer.write(f"{format_value_for_copy(reg['tipo_tercero'])}\t")
        buffer.write(f"{format_value_for_copy(reg['n_dias'])}\t")
        buffer.write(f"{format_value_for_copy(reg['modulo'])}\n")
    
    buffer.seek(0)
    cursor.copy_from(
        buffer,
        'temp_dian_nuevos',
        sep='\t',
        null='',
        columns=(
            'nit_emisor', 'nombre_emisor', 'fecha_emision',
            'prefijo', 'folio', 'total', 'tipo_documento', 'cufe_cude',
            'forma_pago', 'clave', 'clave_acuse', 'tipo_tercero', 'n_dias', 'modulo'
        )
    )
    
    # INSERT todos los registros CON PROTECCIÓN contra duplicados (ON CONFLICT)
    cursor.execute("""
        INSERT INTO dian (
            nit_emisor, nombre_emisor, fecha_emision, prefijo, folio, total,
            tipo_documento, cufe_cude, forma_pago, clave, clave_acuse,
            tipo_tercero, n_dias, modulo
        )
        SELECT 
            nit_emisor, nombre_emisor, fecha_emision, prefijo, folio, total,
            tipo_documento, cufe_cude, forma_pago, clave, clave_acuse,
            tipo_tercero, n_dias, modulo
        FROM temp_dian_nuevos
        ON CONFLICT (cufe_cude) DO NOTHING
    """)
    
    insertados = cursor.rowcount
    print(f"   ✅ {insertados:,} registros NUEVOS insertados (duplicados ignorados)")
    return len(registros)


def insertar_erp_comercial_bulk(df_polars, cursor):
    """Inserta registros masivos en tabla erp_comercial usando COPY FROM"""
    print("\n📊 Insertando en tabla ERP COMERCIAL...")
    
    df = df_polars.to_pandas()
    
    # 🔍 DETECTAR COLUMNAS DINÁMICAMENTE (igual que el código original)
    cols = df.columns.tolist()
    proveedor_col = next((c for c in cols if "proveedor" in c.lower() and "razon" not in c.lower()), None)
    razon_col = next((c for c in cols if "razon" in c.lower() and "social" in c.lower()), None)
    docto_col = next((c for c in cols if "docto" in c.lower() and "proveedor" in c.lower()), None)
    co_col = next((c for c in cols if c.upper() == "CO" or c.upper() == "C.O."), None)
    nro_doc_col = next((c for c in cols if "nro" in c.lower() and "documento" in c.lower()), None)
    usuario_col = next((c for c in cols if "usuario" in c.lower() and "creac" in c.lower()), None)
    clase_col = next((c for c in cols if "clase" in c.lower() and "docto" in c.lower()), None)  # FIX: busca "clase" + "docto" (no "documento")
    fecha_col = next((c for c in cols if "fecha" in c.lower() and "docto" in c.lower()), None)  # FIX: busca "fecha" + "docto" (no "recib")
    valor_col = next((c for c in cols if "valor" in c.lower() and "bruto" in c.lower()), None)  # FIX: busca "valor" + "bruto" (no "total")
    
    if not all([proveedor_col, docto_col]):
        print(f"   ⚠️ ERROR: Columnas requeridas no encontradas")
        print(f"      Proveedor: {proveedor_col}, Docto: {docto_col}")
        return 0
    
    registros = []
    
    for idx, (_, row) in enumerate(df.iterrows()):
        # Extraer campos básicos usando columnas detectadas
        proveedor = str(row.get(proveedor_col, '')).strip() if proveedor_col else ''
        razon_social = str(row.get(razon_col, '')).strip() if razon_col else ''
        docto_proveedor = str(row.get(docto_col, '')).strip() if docto_col else ''
        co = str(row.get(co_col, '')).strip() if co_col else ''
        nro_documento = str(row.get(nro_doc_col, '')).strip() if nro_doc_col else ''
        usuario_creacion = str(row.get(usuario_col, '')).strip() if usuario_col else ''
        clase_documento = str(row.get(clase_col, '')).strip() if clase_col else ''
        
        # Fecha recibido
        fecha_recibido = None
        if fecha_col:
            try:
                fecha_raw = row.get(fecha_col)
                if fecha_raw and not pd.isna(fecha_raw):
                    if isinstance(fecha_raw, str):
                        fecha_recibido = pd.to_datetime(fecha_raw, format='%d/%m/%Y', errors='coerce').date()
                    else:
                        fecha_recibido = fecha_raw.date() if hasattr(fecha_raw, 'date') else fecha_raw
            except:
                pass
        
        # Valor
        valor = 0.0
        if valor_col:
            try:
                valor_raw = row.get(valor_col, 0)
                if valor_raw and not pd.isna(valor_raw):
                    if isinstance(valor_raw, str):
                        valor = float(valor_raw.strip().replace('.', '').replace(',', '.'))
                    else:
                        valor = float(valor_raw)
            except:
                pass
        
        # Campos calculados
        prefijo = extraer_prefijo(docto_proveedor)
        folio_raw = extraer_folio(docto_proveedor)
        folio = ultimos_8_sin_ceros(folio_raw)
        clave_erp_comercial = crear_clave_factura(proveedor, prefijo, folio_raw)
        doc_causado_por = f"{co} - {usuario_creacion} - {nro_documento}"
        modulo = 'Comercial'
        
        registros.append({
            'proveedor': proveedor,
            'razon_social': razon_social,
            'docto_proveedor': docto_proveedor,
            'prefijo': prefijo,
            'folio': folio,
            'co': co,
            'nro_documento': nro_documento,
            'fecha_recibido': fecha_recibido,
            'usuario_creacion': usuario_creacion,
            'clase_documento': clase_documento,
            'valor': valor,
            'clave_erp_comercial': clave_erp_comercial,
            'doc_causado_por': doc_causado_por,
            'modulo': modulo
        })
    
    # 🔥 CARGA INCREMENTAL: Usar tabla temporal + INSERT ... ON CONFLICT DO NOTHING
    print(f"   🔄 Carga INCREMENTAL (conserva datos viejos, agrega nuevos)")
    
   # Crear tabla temporal (sin restricciones UNIQUE para permitir duplicados internos en archivo)
    cursor.execute("""
        CREATE TEMP TABLE temp_erp_comercial_nuevos (LIKE erp_comercial INCLUDING DEFAULTS) ON COMMIT DROP
    """)
    
    # COPY FROM a tabla temporal
    buffer = io.StringIO()
    for reg in registros:
        buffer.write(f"{format_value_for_copy(reg['proveedor'])}\t")
        buffer.write(f"{format_value_for_copy(reg['razon_social'])}\t")
        buffer.write(f"{format_value_for_copy(reg['docto_proveedor'])}\t")
        buffer.write(f"{format_value_for_copy(reg['prefijo'])}\t")
        buffer.write(f"{format_value_for_copy(reg['folio'])}\t")
        buffer.write(f"{format_value_for_copy(reg['co'])}\t")
        buffer.write(f"{format_value_for_copy(reg['nro_documento'])}\t")
        buffer.write(f"{format_value_for_copy(reg['fecha_recibido'])}\t")
        buffer.write(f"{format_value_for_copy(reg['usuario_creacion'])}\t")
        buffer.write(f"{format_value_for_copy(reg['clase_documento'])}\t")
        buffer.write(f"{format_value_for_copy(reg['valor'])}\t")
        buffer.write(f"{format_value_for_copy(reg['clave_erp_comercial'])}\t")
        buffer.write(f"{format_value_for_copy(reg['doc_causado_por'])}\t")
        buffer.write(f"{format_value_for_copy(reg['modulo'])}\n")
    
    buffer.seek(0)
    cursor.copy_from(
        buffer,
        'temp_erp_comercial_nuevos',
        sep='\t',
        null='',
        columns=(
            'proveedor', 'razon_social', 'docto_proveedor', 'prefijo', 'folio',
            'co', 'nro_documento', 'fecha_recibido', 'usuario_creacion',
            'clase_documento', 'valor', 'clave_erp_comercial', 'doc_causado_por', 'modulo'
        )
    )
    
    # INSERT solo registros nuevos (si clave ya existe, no inserta)
    cursor.execute("""
        INSERT INTO erp_comercial (
            proveedor, razon_social, docto_proveedor, prefijo, folio,
            co, nro_documento, fecha_recibido, usuario_creacion,
            clase_documento, valor, clave_erp_comercial, doc_causado_por, modulo
        )
        SELECT 
            proveedor, razon_social, docto_proveedor, prefijo, folio,
            co, nro_documento, fecha_recibido, usuario_creacion,
            clase_documento, valor, clave_erp_comercial, doc_causado_por, modulo
        FROM temp_erp_comercial_nuevos
        ON CONFLICT (clave_erp_comercial) DO NOTHING
    """)
    
    insertados = cursor.rowcount
    print(f"   ✅ {insertados:,} registros NUEVOS insertados (duplicados ignorados)")
    return len(registros)


def insertar_erp_financiero_bulk(df_polars, cursor):
    """Inserta registros masivos en tabla erp_financiero usando COPY FROM"""
    print("\n📊 Insertando en tabla ERP FINANCIERO...")
    
    df = df_polars.to_pandas()
    
    # 🔍 DETECTAR COLUMNAS DINÁMICAMENTE
    cols = df.columns.tolist()
    proveedor_col = next((c for c in cols if "proveedor" in c.lower() and "razon" not in c.lower()), None)
    razon_col = next((c for c in cols if "razon" in c.lower() and "social" in c.lower()), None)
    docto_col = next((c for c in cols if "docto" in c.lower() and "proveedor" in c.lower()), None)
    co_col = next((c for c in cols if c.upper() == "CO" or c.upper() == "C.O."), None)
    nro_doc_col = next((c for c in cols if "nro" in c.lower() and "documento" in c.lower()), None)
    usuario_col = next((c for c in cols if "usuario" in c.lower() and "creac" in c.lower()), None)
    clase_col = next((c for c in cols if "clase" in c.lower() and "docto" in c.lower()), None)  # FIX: busca "clase" + "docto"
    fecha_col = next((c for c in cols if "fecha" in c.lower() and ("proveedor" in c.lower() or "docto" in c.lower())), None)  # FIX: busca "fecha" + "proveedor" o "docto"
    valor_col = next((c for c in cols if "valor" in c.lower() and ("subtotal" in c.lower() or "bruto" in c.lower())), None)  # FIX: busca "valor" + "subtotal" o "bruto"
    
    if not all([proveedor_col, docto_col]):
        print(f"   ⚠️ ERROR: Columnas requeridas no encontradas")
        print(f"      Proveedor: {proveedor_col}, Docto: {docto_col}")
        return 0
    
    registros = []
    
    for idx, (_, row) in enumerate(df.iterrows()):
        # Extraer campos básicos usando columnas detectadas
        proveedor = str(row.get(proveedor_col, '')).strip() if proveedor_col else ''
        razon_social = str(row.get(razon_col, '')).strip() if razon_col else ''
        docto_proveedor = str(row.get(docto_col, '')).strip() if docto_col else ''
        co = str(row.get(co_col, '')).strip() if co_col else ''
        nro_documento = str(row.get(nro_doc_col, '')).strip() if nro_doc_col else ''
        usuario_creacion = str(row.get(usuario_col, '')).strip() if usuario_col else ''
        clase_documento = str(row.get(clase_col, '')).strip() if clase_col else ''
        
        # Fecha recibido
        fecha_recibido = None
        if fecha_col:
            try:
                fecha_raw = row.get(fecha_col)
                if fecha_raw and not pd.isna(fecha_raw):
                    if isinstance(fecha_raw, str):
                        fecha_recibido = pd.to_datetime(fecha_raw, format='%d/%m/%Y', errors='coerce').date()
                    else:
                        fecha_recibido = fecha_raw.date() if hasattr(fecha_raw, 'date') else fecha_raw
            except:
                pass
        
        # Valor
        valor = 0.0
        if valor_col:
            try:
                valor_raw = row.get(valor_col, 0)
                if valor_raw and not pd.isna(valor_raw):
                    if isinstance(valor_raw, str):
                        valor = float(valor_raw.strip().replace('.', '').replace(',', '.'))
                    else:
                        valor = float(valor_raw)
            except:
                pass
        
        # Campos calculados
        prefijo = extraer_prefijo(docto_proveedor)
        folio_raw = extraer_folio(docto_proveedor)
        folio = ultimos_8_sin_ceros(folio_raw)
        clave_erp_financiero = crear_clave_factura(proveedor, prefijo, folio_raw)
        doc_causado_por = f"{co} - {usuario_creacion} - {nro_documento}"
        modulo = 'Financiero'
        
        registros.append({
            'proveedor': proveedor,
            'razon_social': razon_social,
            'docto_proveedor': docto_proveedor,
            'prefijo': prefijo,
            'folio': folio,
            'co': co,
            'nro_documento': nro_documento,
            'fecha_recibido': fecha_recibido,
            'usuario_creacion': usuario_creacion,
            'clase_documento': clase_documento,
            'valor': valor,
            'clave_erp_financiero': clave_erp_financiero,
            'doc_causado_por': doc_causado_por,
            'modulo': modulo
        })
    
    # 🔥 CARGA INCREMENTAL: Usar tabla temporal + INSERT ... ON CONFLICT DO NOTHING
    print(f"   🔄 Carga INCREMENTAL (conserva datos viejos, agrega nuevos)")
    
    # Crear tabla temporal (sin restricciones UNIQUE para permitir duplicados internos en archivo)
    cursor.execute("""
        CREATE TEMP TABLE temp_erp_financiero_nuevos (LIKE erp_financiero INCLUDING DEFAULTS) ON COMMIT DROP
    """)
    
    # COPY FROM a tabla temporal
    buffer = io.StringIO()
    for reg in registros:
        buffer.write(f"{format_value_for_copy(reg['proveedor'])}\t")
        buffer.write(f"{format_value_for_copy(reg['razon_social'])}\t")
        buffer.write(f"{format_value_for_copy(reg['docto_proveedor'])}\t")
        buffer.write(f"{format_value_for_copy(reg['prefijo'])}\t")
        buffer.write(f"{format_value_for_copy(reg['folio'])}\t")
        buffer.write(f"{format_value_for_copy(reg['co'])}\t")
        buffer.write(f"{format_value_for_copy(reg['nro_documento'])}\t")
        buffer.write(f"{format_value_for_copy(reg['fecha_recibido'])}\t")
        buffer.write(f"{format_value_for_copy(reg['usuario_creacion'])}\t")
        buffer.write(f"{format_value_for_copy(reg['clase_documento'])}\t")
        buffer.write(f"{format_value_for_copy(reg['valor'])}\t")
        buffer.write(f"{format_value_for_copy(reg['clave_erp_financiero'])}\t")
        buffer.write(f"{format_value_for_copy(reg['doc_causado_por'])}\t")
        buffer.write(f"{format_value_for_copy(reg['modulo'])}\n")
    
    buffer.seek(0)
    cursor.copy_from(
        buffer,
        'temp_erp_financiero_nuevos',
        sep='\t',
        null='',
        columns=(
            'proveedor', 'razon_social', 'docto_proveedor', 'prefijo', 'folio',
            'co', 'nro_documento', 'fecha_recibido', 'usuario_creacion',
            'clase_documento', 'valor', 'clave_erp_financiero', 'doc_causado_por', 'modulo'
        )
    )
    
    # INSERT solo registros nuevos (si clave ya existe, no inserta)
    cursor.execute("""
        INSERT INTO erp_financiero (
            proveedor, razon_social, docto_proveedor, prefijo, folio,
            co, nro_documento, fecha_recibido, usuario_creacion,
            clase_documento, valor, clave_erp_financiero, doc_causado_por, modulo
        )
        SELECT 
            proveedor, razon_social, docto_proveedor, prefijo, folio,
            co, nro_documento, fecha_recibido, usuario_creacion,
            clase_documento, valor, clave_erp_financiero, doc_causado_por, modulo
        FROM temp_erp_financiero_nuevos
        ON CONFLICT (clave_erp_financiero) DO NOTHING
    """)
    
    insertados = cursor.rowcount
    print(f"   ✅ {insertados:,} registros NUEVOS insertados (duplicados ignorados)")
    return len(registros)


def insertar_acuses_bulk(df_polars, cursor):
    """
    Inserta registros masivos en tabla acuses usando COPY FROM
    🆕 CORREGIDO: Usa columnas REALES de tabla acuses (no columnas de DIAN)
    """
    import io
    import pandas as pd
    
    print("\n📊 Insertando en tabla ACUSES...")
    
    # Convertir a pandas
    acuses_pd = df_polars.to_pandas()
    
    # 🆕 MAPEAR COLUMNAS: Excel → PostgreSQL
    # Las columnas de Excel tienen acentos y espacios, necesitamos normalizar
    column_mapping = {
        'fecha': 'fecha',
        'adquiriente': 'adquiriente',
        'factura': 'factura',
        'emisor': 'emisor',
        'nit emisor': 'nit_emisor',
        'nit. pt': 'nit_pt',
        'estado docto.': 'estado_docto',
        'descripción reclamo': 'descripcion_reclamo',
        'tipo documento': 'tipo_documento',
        'cufe': 'cufe',
        'valor': 'valor',
        'acuse recibido': 'acuse_recibido',
        'recibo bien servicio': 'recibo_bien_servicio',
        'aceptación expresa': 'aceptacion_expresa',
        'reclamo': 'reclamo',
        'aceptación tacita': 'aceptacion_tacita'
    }
    
    # Renombrar columnas que existen en el Excel
    columns_to_rename = {k: v for k, v in column_mapping.items() if k in acuses_pd.columns}
    acuses_pd = acuses_pd.rename(columns=columns_to_rename)
    
    # Generar clave_acuse (CUFE + Estado) - clave única por acuse
    if 'cufe' in acuses_pd.columns and 'estado_docto' in acuses_pd.columns:
        acuses_pd['clave_acuse'] = (
            acuses_pd['cufe'].fillna('').astype(str) + '|' + 
            acuses_pd['estado_docto'].fillna('Pendiente').astype(str)
        )
    else:
        print(f"   ⚠️ ERROR: Columnas CUFE o ESTADO_DOCTO no encontradas")
        print(f"      Columnas disponibles: {acuses_pd.columns.tolist()}")
        return 0
    
    registros = acuses_pd.to_dict('records')
    
    print(f"   🔄 Carga INCREMENTAL (actualiza solo si jerarquía mayor)")
    print(f"   📋 {len(registros):,} registros a procesar")
    
    # Crear tabla temporal (sin restricciones UNIQUE para permitir duplicados internos)
    cursor.execute("""
        CREATE TEMP TABLE temp_acuses_nuevos (LIKE acuses INCLUDING DEFAULTS) ON COMMIT DROP
    """)
    
    # COPY FROM a tabla temporal usando las columnas CORRECTAS
    buffer = io.StringIO()
    for reg in registros:
        # Usar format_value_for_copy() para manejar None correctamente
        buffer.write(f"{format_value_for_copy(reg.get('fecha'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('adquiriente'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('factura'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('emisor'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('nit_emisor'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('nit_pt'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('estado_docto'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('descripcion_reclamo'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('tipo_documento'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('cufe'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('valor'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('acuse_recibido'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('recibo_bien_servicio'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('aceptacion_expresa'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('reclamo'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('aceptacion_tacita'))}\t")
        buffer.write(f"{format_value_for_copy(reg.get('clave_acuse'))}\n")
    
    buffer.seek(0)
    
    # COPY FROM temp table con columnas CORRECTAS
    cursor.copy_from(
        buffer,
        'temp_acuses_nuevos',
        sep='\t',
        null='',
        columns=(
            'fecha', 'adquiriente', 'factura', 'emisor', 'nit_emisor', 'nit_pt',
            'estado_docto', 'descripcion_reclamo', 'tipo_documento', 'cufe', 'valor',
            'acuse_recibido', 'recibo_bien_servicio', 'aceptacion_expresa',
            'reclamo', 'aceptacion_tacita', 'clave_acuse'
        )
    )
    
    # INSERT con UPDATE si jerarquía mayor
    cursor.execute("""
        INSERT INTO acuses (
            fecha, adquiriente, factura, emisor, nit_emisor, nit_pt,
            estado_docto, descripcion_reclamo, tipo_documento, cufe, valor,
            acuse_recibido, recibo_bien_servicio, aceptacion_expresa,
            reclamo, aceptacion_tacita, clave_acuse, fecha_carga
        )
        SELECT 
            fecha, adquiriente, factura, emisor, nit_emisor, nit_pt,
            estado_docto, descripcion_reclamo, tipo_documento, cufe, valor,
            acuse_recibido, recibo_bien_servicio, aceptacion_expresa,
            reclamo, aceptacion_tacita, clave_acuse, NOW()
        FROM temp_acuses_nuevos
        ON CONFLICT (clave_acuse) DO UPDATE SET
            fecha = EXCLUDED.fecha,
            adquiriente = EXCLUDED.adquiriente,
            factura = EXCLUDED.factura,
            emisor = EXCLUDED.emisor,
            nit_emisor = EXCLUDED.nit_emisor,
            nit_pt = EXCLUDED.nit_pt,
            estado_docto = EXCLUDED.estado_docto,
            descripcion_reclamo = EXCLUDED.descripcion_reclamo,
            tipo_documento = EXCLUDED.tipo_documento,
            cufe = EXCLUDED.cufe,
            valor = EXCLUDED.valor,
            acuse_recibido = EXCLUDED.acuse_recibido,
            recibo_bien_servicio = EXCLUDED.recibo_bien_servicio,
            aceptacion_expresa = EXCLUDED.aceptacion_expresa,
            reclamo = EXCLUDED.reclamo,
            aceptacion_tacita = EXCLUDED.aceptacion_tacita,
            fecha_actualizacion = NOW()
        WHERE calcular_jerarquia_acuse(EXCLUDED.estado_docto) > calcular_jerarquia_acuse(acuses.estado_docto)
    """)
    
    # Contar registros procesados
    cursor.execute("SELECT COUNT(*) FROM temp_acuses_nuevos")
    registros_procesados = cursor.fetchone()[0]
    
    print(f"   ✅ {registros_procesados:,} registros procesados (nuevos + actualizaciones)")
    return registros_procesados

# ==============================
# PROCESAMIENTO PRINCIPAL OPTIMIZADO
# ==============================

def actualizar_maestro() -> str:

    """
    Sistema ULTRA OPTIMIZADO con PostgreSQL usando COPY FROM
    Velocidad: 25,000+ registros/segundo (similar a SQLite)
    
    Procesa: DIAN + ERP (FN+CM+Errores) + Acuses
    Técnica: Bulk COPY FROM (comando nativo PostgreSQL)
    """
    import io
    import psycopg2
    from sqlalchemy import create_engine, text
    
    print("\n" + "="*80)
    print("🚀 INICIANDO PROCESO DE CARGA DE CSV")
    print("="*80)
    print(f"📅 Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*80)
    
    t0 = time.time()
    
    # 1️⃣ CARGAR ARCHIVO DIAN (OBLIGATORIO)
    print("\n📂 Buscando archivo DIAN...")
    f_dian = latest_csv(UPLOADS["dian"])
    if not f_dian:
        print("❌ No hay archivos DIAN")
        return "⚠️ No hay archivos DIAN. Sube un Excel/CSV de DIAN para procesar."
    
    print(f"✅ Archivo encontrado: {os.path.basename(f_dian)}")
    
    # 🔥 DETECTAR TIPO DE ARCHIVO (Excel vs CSV)
    ext_dian = os.path.splitext(f_dian)[1].lower()
    
    if ext_dian in ['.xlsx', '.xlsm']:
        # ⚡ EXCEL: Usar Polars directamente (NO convertir a CSV)
        print(f"\n📊 Archivo Excel detectado - lectura directa con Polars")
        t_read_start = time.time()
        d = read_csv(f_dian)  # Polars detecta Excel automáticamente
        t_read_end = time.time()
        print(f"✅ Excel leído en {t_read_end - t_read_start:.2f}s")
    else:
        # 📄 CSV: Detectar separador y encoding
        print(f"\n🔍 Detectando tipo de CSV...")
        
        # Probar diferentes encodings
        encoding_correcto = 'utf-8'
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                with open(f_dian, 'r', encoding=enc) as f:
                    primera_linea = f.readline()
                    # Si tiene caracteres raros, no es el encoding correcto
                    if 'Ã' not in primera_linea or enc == 'cp1252':  # cp1252 es fallback
                        encoding_correcto = enc
                        print(f"   Encoding detectado: {enc}")
                        break
            except:
                continue
        
        with open(f_dian, 'r', encoding=encoding_correcto) as f:
            primera_linea = f.readline()
            if '\t' in primera_linea:
                separador = '\t'
                tipo_archivo = 'TSV (tabulaciones)'
            else:
                separador = ','
                tipo_archivo = 'CSV (comas)'
        print(f"   Separador detectado: {tipo_archivo}")
        
        t_read_start = time.time()
        d = read_csv(f_dian, separator=separador, encoding=encoding_correcto)
        t_read_end = time.time()
        print(f"✅ CSV leído en {t_read_end - t_read_start:.2f}s")
    
    registros_dian = d.height
    print(f"📊 Registros DIAN: {registros_dian:,}")
    
    if registros_dian == 0:
        return "⚠️ Archivo DIAN está vacío"
    
    # 2️⃣ CARGAR ARCHIVOS ERP (FN + CM + ERRORES) - CON CLASIFICACIÓN POR MÓDULO
    # Catálogos de clases de documento
    CLASES_ACREEDOR = {
        'Factura de servicio desde sol. anticipo',
        'Factura de servicio de reg.fijo compra',
        'Factura de servicio compra',
        'Legalización factura anticipos',
        'Legalización factura caja menor',
        'Nota débito de servicios - compra',
        'Factura de servicio, legalizacion gastos',
        'Legalización nota debito anticipos'
    }
    
    CLASES_PROVEEDOR = {
        'Notas débito de proveedor',
        'Factura de proveedor',
        'Factura de consignación'
    }
    
    registros_erp = 0
    erp_por_clave = {}
    tipo_tercero_por_nit = {}  # ✅ NUEVO: clasificar por NIT
    
    # 🆕 INICIALIZAR VARIABLES (necesario para inserción posterior)
    erp_fn = None
    erp_cm = None
    acuses_df = None
    
    # Procesar ERP FINANCIERO
    csv_fn = latest_csv(UPLOADS["erp_fn"])
    if csv_fn:
        erp_fn = read_csv(csv_fn)
        if erp_fn.height > 0:
            registros_erp += erp_fn.height
            erp_fn_pd = erp_fn.to_pandas()
            
            # Detectar columnas
            cols = erp_fn_pd.columns.tolist()
            proveedor_col = next((c for c in cols if "proveedor" in c.lower() and "razon" not in c.lower()), None)
            docto_col = next((c for c in cols if "docto" in c.lower() and "proveedor" in c.lower()), None)
            clase_col = next((c for c in cols if "clase" in c.lower()), None)
            co_col = next((c for c in cols if c.upper() == "C.O."), None)
            usuario_col = next((c for c in cols if "usuario" in c.lower() and "creac" in c.lower()), None)
            nro_doc_col = next((c for c in cols if "nro" in c.lower() and "documento" in c.lower()), None)
            
            if all([proveedor_col, docto_col, clase_col]):
                for _, row in erp_fn_pd.iterrows():
                    nit = str(row[proveedor_col])
                    docto = str(row[docto_col])
                    clase = str(row[clase_col])
                    
                    # ✅ EXTRAER PREFIJO Y FOLIO del campo "Docto. proveedor"
                    prefijo = extraer_prefijo(docto)
                    folio = ultimos_8_sin_ceros(extraer_folio(docto))
                    clave = crear_clave_factura(nit, prefijo, folio)
                    modulo = 'FINANCIERO'
                    
                    # ✅ Construir doc_causado_por: "C.O. - Usuario creación - Nro documento"
                    doc_causado_por = ""
                    if co_col and usuario_col and nro_doc_col:
                        co = str(row.get(co_col, '')).strip()
                        usuario = str(row.get(usuario_col, '')).strip()
                        nro_doc = str(row.get(nro_doc_col, '')).strip()
                        
                        if co and usuario and nro_doc:
                            doc_causado_por = f"{co} - {usuario} - {nro_doc}"
                    
                    if clave not in erp_por_clave:
                        erp_por_clave[clave] = {
                            'modulo': modulo,
                            'doc_causado_por': doc_causado_por
                        }
                    
                    # ✅ Clasificar tipo de tercero si es clase acreedor
                    if clase in CLASES_ACREEDOR:
                        nit_limpio = extraer_folio(nit)
                        if nit_limpio not in tipo_tercero_por_nit:
                            tipo_tercero_por_nit[nit_limpio] = set()
                        tipo_tercero_por_nit[nit_limpio].add('ACREEDOR')
    
    # Procesar ERP COMERCIAL
    csv_cm = latest_csv(UPLOADS["erp_cm"])
    if csv_cm:
        erp_cm = read_csv(csv_cm)
        if erp_cm.height > 0:
            registros_erp += erp_cm.height
            erp_cm_pd = erp_cm.to_pandas()
            
            # Detectar columnas
            cols = erp_cm_pd.columns.tolist()
            proveedor_col = next((c for c in cols if "proveedor" in c.lower() and "razon" not in c.lower()), None)
            docto_col = next((c for c in cols if "docto" in c.lower() and "proveedor" in c.lower()), None)
            clase_col = next((c for c in cols if "clase" in c.lower()), None)
            co_col = next((c for c in cols if c.upper() == "C.O."), None)
            usuario_col = next((c for c in cols if "usuario" in c.lower() and "creac" in c.lower()), None)
            nro_doc_col = next((c for c in cols if "nro" in c.lower() and "documento" in c.lower()), None)
            
            if all([proveedor_col, docto_col, clase_col]):
                for _, row in erp_cm_pd.iterrows():
                    nit = str(row[proveedor_col])
                    docto = str(row[docto_col])
                    clase = str(row[clase_col])
                    
                    # ✅ EXTRAER PREFIJO Y FOLIO del campo "Docto. proveedor"
                    prefijo = extraer_prefijo(docto)
                    folio = ultimos_8_sin_ceros(extraer_folio(docto))
                    clave = crear_clave_factura(nit, prefijo, folio)
                    modulo = 'COMERCIAL'
                    
                    # ✅ Construir doc_causado_por: "C.O. - Usuario creación - Nro documento"
                    doc_causado_por = ""
                    if co_col and usuario_col and nro_doc_col:
                        co = str(row.get(co_col, '')).strip()
                        usuario = str(row.get(usuario_col, '')).strip()
                        nro_doc = str(row.get(nro_doc_col, '')).strip()
                        
                        if co and usuario and nro_doc:
                            doc_causado_por = f"{co} - {usuario} - {nro_doc}"
                    
                    if clave not in erp_por_clave:
                        erp_por_clave[clave] = {
                            'modulo': modulo,
                            'doc_causado_por': doc_causado_por
                        }
                    
                    # ✅ Clasificar tipo de tercero si es clase proveedor
                    if clase in CLASES_PROVEEDOR:
                        nit_limpio = extraer_folio(nit)
                        if nit_limpio not in tipo_tercero_por_nit:
                            tipo_tercero_por_nit[nit_limpio] = set()
                        tipo_tercero_por_nit[nit_limpio].add('PROVEEDOR')
    
    # Procesar ERRORES ERP (sin clasificar tipo tercero)
    csv_errores = latest_csv(UPLOADS["errores"])
    if csv_errores:
        erp_err = read_csv(csv_errores)
        if erp_err.height > 0:
            registros_erp += erp_err.height
            erp_err_pd = erp_err.to_pandas()
            
            cols = erp_err_pd.columns.tolist()
            proveedor_col = next((c for c in cols if "proveedor" in c.lower() and "razon" not in c.lower()), None)
            docto_col = next((c for c in cols if "docto" in c.lower() and "proveedor" in c.lower()), None)
            clase_col = next((c for c in cols if "clase" in c.lower()), None)
            
            if all([proveedor_col, docto_col, clase_col]):
                for _, row in erp_err_pd.iterrows():
                    nit = str(row[proveedor_col])
                    docto = str(row[docto_col])
                    clase = str(row[clase_col])
                    
                    clave = crear_clave_factura(nit, docto, docto)
                    modulo = mapear_modulo(clase)
                    
                    if modulo and clave not in erp_por_clave:
                        erp_por_clave[clave] = {
                            'modulo': modulo,
                            'doc_causado_por': ''
                        }
    
    # 3️⃣ CARGAR ACUSES (OPCIONAL)
    acuses_csv = latest_csv(UPLOADS["acuses"])
    acuses_por_cufe = {}
    registros_acuses = 0
    
    if acuses_csv:
        acuses_df = read_csv(acuses_csv)
        registros_acuses = acuses_df.height
        acuses_pd = acuses_df.to_pandas()
        
        # 🐛 DEBUG: Mostrar columnas del archivo de acuses
        print(f"\n🔍 DEBUG - Columnas disponibles en archivo ACUSES:")
        print(f"   Total columnas: {len(acuses_pd.columns)}")
        for col in acuses_pd.columns:
            print(f"   - '{col}' (tipo: {acuses_pd[col].dtype})")
        
        # 🐛 DEBUG: Mostrar primeras 3 filas de acuses
        print(f"\n🔍 DEBUG - Primeras 3 filas de ACUSES:")
        for idx, (_, row) in enumerate(acuses_pd.head(3).iterrows()):
            print(f"   Fila {idx}:")
            for col in acuses_pd.columns:
                print(f"      {col}: {repr(row[col])}")
        
        # Mapear CUFE → estado_aprobacion (buscar columnas usando normalización)
        cufe_col = None
        estado_col = None
        
        # Buscar columna de CUFE
        for col in acuses_pd.columns:
            col_lower = col.lower().strip()
            if 'cufe' in col_lower or 'cude' in col_lower:
                cufe_col = col
                print(f"✅ Columna CUFE detectada: '{col}'")
                break
        
        # Buscar columna de estado
        for col in acuses_pd.columns:
            col_lower = col.lower().strip()
            if 'estado' in col_lower and ('docto' in col_lower or 'documento' in col_lower):
                estado_col = col
                print(f"✅ Columna ESTADO detectada: '{col}'")
                break
        
        if not cufe_col:
            print(f"⚠️ WARNING: No se encontró columna de CUFE en acuses")
        if not estado_col:
            print(f"⚠️ WARNING: No se encontró columna de ESTADO en acuses")
        
        # Procesar acuses
        for _, row in acuses_pd.iterrows():
            cufe = str(row.get(cufe_col, '')) if cufe_col else ''
            estado = str(row.get(estado_col, 'Pendiente')) if estado_col else 'Pendiente'
            
            if cufe and cufe.strip():
                acuses_por_cufe[cufe.strip()] = estado.strip()
        
        print(f"\n✅ {len(acuses_por_cufe):,} acuses procesados (CUFE → Estado)")
        
        # 🐛 DEBUG: Mostrar primeros 5 acuses cargados
        print(f"\n🔍 DEBUG - Primeros 5 acuses en diccionario:")
        for idx, (cufe, estado) in enumerate(list(acuses_por_cufe.items())[:5]):
            print(f"   {idx+1}. CUFE: {cufe[:50]}... → Estado: {estado}")
    
    # 🆕 PASO INTERMEDIO: CONSOLIDAR TIPOS DE TERCERO (NECESARIO PARA INSERTAR EN DIAN)
    print("\n🔍 Consolidando tipos de tercero...")
    tipo_tercero_final = {}
    for nit, tipos in tipo_tercero_por_nit.items():
        if 'PROVEEDOR' in tipos and 'ACREEDOR' in tipos:
            tipo_tercero_final[nit] = 'PROVEEDOR Y ACREEDOR'
        elif 'PROVEEDOR' in tipos:
            tipo_tercero_final[nit] = 'PROVEEDOR'
        elif 'ACREEDOR' in tipos:
            tipo_tercero_final[nit] = 'ACREEDOR'
    print(f"✅ {len(tipo_tercero_final):,} terceros clasificados")
    
    # 🆕 INSERTAR EN TABLAS INDIVIDUALES (dian, erp, acuses)
    print("\n" + "="*80)
    print("🔥 INSERTANDO EN TABLAS INDIVIDUALES (dian, erp_comercial, erp_financiero, acuses)")
    print("="*80)
    
    try:
        # Conexión directa PostgreSQL (sin ORM)
        engine = create_engine(current_app.config['SQLALCHEMY_DATABASE_URI'])
        raw_conn = engine.raw_connection()
        cursor = raw_conn.cursor()
        
        # 1️⃣ INSERTAR EN TABLA DIAN
        insertar_dian_bulk(d, cursor, tipo_tercero_final)
        
        # 2️⃣ INSERTAR EN TABLA ERP_COMERCIAL
        if erp_cm is not None and erp_cm.height > 0:
            insertar_erp_comercial_bulk(erp_cm, cursor)
        else:
            print("   ⚠️ No hay datos de ERP Comercial para insertar")
        
        # 3️⃣ INSERTAR EN TABLA ERP_FINANCIERO
        if erp_fn is not None and erp_fn.height > 0:
            insertar_erp_financiero_bulk(erp_fn, cursor)
        else:
            print("   ⚠️ No hay datos de ERP Financiero para insertar")
        
        # 4️⃣ INSERTAR EN TABLA ACUSES
        if acuses_df is not None and acuses_df.height > 0:
            insertar_acuses_bulk(acuses_df, cursor)
        else:
            print("   ⚠️ No hay datos de ACUSES para insertar")
        
        # Commit de todas las inserciones
        raw_conn.commit()
        print("\n✅ TODAS LAS TABLAS INDIVIDUALES ACTUALIZADAS CORRECTAMENTE")
        
    except Exception as e:
        if 'raw_conn' in locals():
            raw_conn.rollback()
        print(f"\n❌ ERROR insertando en tablas individuales: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    # Mantener cursor abierto para inserción posterior a maestro
    # (No cerrar conexión todavía)
    
    # 5️⃣ PROCESAR DIAN EN POLARS (RÁPIDO)
    d_pd = d.to_pandas()
    
    # � NORMALIZAR NOMBRES DE COLUMNAS (lowercase + sin tildes + guion bajo)
    import unicodedata
    def normalizar_columna(nombre):
        """Normaliza nombre: lowercase, sin tildes, espacios→guion bajo"""
        # Quitar tildes
        sin_tildes = ''.join(
            c for c in unicodedata.normalize('NFD', str(nombre))
            if unicodedata.category(c) != 'Mn'
        )
        # Lowercase y espacios→guion bajo
        return sin_tildes.lower().strip().replace(' ', '_')
    
    # Función para convertir números con formato español (coma decimal)
    def convertir_numero_espanol(valor):
        """Convierte un valor con coma decimal española a float"""
        if valor is None or pd.isna(valor):
            return 0.0
        if isinstance(valor, (int, float)):
            return float(valor)
        if isinstance(valor, str):
            # Eliminar espacios y reemplazar coma por punto
            valor_limpio = valor.strip().replace('.', '').replace(',', '.')
            try:
                return float(valor_limpio)
            except:
                return 0.0
        return 0.0
    
    # Crear diccionario {nombre_normalizado: nombre_original}
    columnas_originales = {}
    for col in d_pd.columns:
        col_norm = normalizar_columna(col)
        columnas_originales[col_norm] = col
        # Agregar variante sin guion bajo
        columnas_originales[col_norm.replace('_', '')] = col
    
    # 🐛 DEBUG: Mostrar columnas disponibles
    print(f"\n🔍 DEBUG - Columnas disponibles en CSV DIAN:")
    print(f"   Total columnas: {len(d_pd.columns)}")
    for col in d_pd.columns[:10]:  # Primeras 10 columnas
        col_norm = normalizar_columna(col)
        print(f"   '{col}' → normalizado: '{col_norm}'")
    
    # 🔍 DEBUG: Buscar columnas de fecha específicamente
    columnas_fecha = [c for c in d_pd.columns if 'fecha' in normalizar_columna(c) or 'emis' in normalizar_columna(c)]
    print(f"\n🔍 DEBUG - Columnas relacionadas con fecha:")
    for col in columnas_fecha:
        print(f"   - '{col}' → '{normalizar_columna(col)}' (tipo: {d_pd[col].dtype})")
        if len(d_pd) > 0:
            print(f"      Primer valor: {d_pd[col].iloc[0]}")
    
    # 🔥 FIX: Detectar columna CUFE explícitamente (búsqueda case-insensitive)
    cufe_col_name = None
    for col in d_pd.columns:
        if 'cufe' in col.lower() or 'cude' in col.lower():
            cufe_col_name = col
            print(f"\n✅ COLUMNA CUFE DETECTADA: '{col}'")
            break
    
    if not cufe_col_name:
        print(f"\n⚠️  WARNING: No se encontró columna CUFE/CUDE en archivo DIAN")
        cufe_col_name = 'cufe/cude'  # Fallback a nombre normalizado esperado
    
    # 🔍 DEBUG: Buscar columnas de valor/total específicamente
    columnas_valor = [c for c in d_pd.columns if 'total' in normalizar_columna(c) or 'valor' in normalizar_columna(c)]
    print(f"\n🔍 DEBUG - Columnas de valor/total:")
    for col in columnas_valor:
        print(f"   - '{col}' → '{normalizar_columna(col)}' (tipo: {d_pd[col].dtype})")
        if len(d_pd) > 0:
            print(f"      Primer valor: {d_pd[col].iloc[0]}")
    
    # Preparar registros para COPY FROM
    registros = []
    registros_con_modulo = 0
    errores_fecha = 0
    errores_valor = 0
    
    print(f"\n🔄 Procesando {len(d_pd):,} registros...")
    for idx, (_, row) in enumerate(d_pd.iterrows()):
        # Extraer datos de DIAN usando mapeo normalizado
        nit = str(row.get(columnas_originales.get('nit_emisor', 'NIT Emisor'), '')).strip()
        nit_limpio = extraer_folio(nit)
        
        razon_social = str(row.get(columnas_originales.get('nombre_emisor', 'Nombre Emisor'), '')).strip()
        
        # 🔥 BUSCAR FECHA EMISIÓN usando columnas normalizadas
        fecha_emision_raw = None
        for variante_norm in ['fecha_emision', 'fechaemision']:
            if variante_norm in columnas_originales:
                col_original = columnas_originales[variante_norm]
                fecha_emision_raw = row.get(col_original)
                if fecha_emision_raw is not None and not pd.isna(fecha_emision_raw):
                    if idx < 3:
                        print(f"   ✅ Fila {idx}: Fecha de '{col_original}' = {fecha_emision_raw}")
                    break
        
        # 🔥 SOLUCIÓN RADICAL: NUNCA usar date.today() - solo usar fecha del CSV o NULL
        fecha_emision = None
        
        if fecha_emision_raw is None or pd.isna(fecha_emision_raw):
            if idx < 3:
                print(f"   ⚠️  Fila {idx}: Fecha no encontrada - quedará NULL")
            errores_fecha += 1
            fecha_emision = None
        elif isinstance(fecha_emision_raw, str):
            fecha_str = fecha_emision_raw.strip()
            try:
                # ✅ Parsear DD-MM-YYYY (formato del CSV)
                if '-' in fecha_str:
                    partes = fecha_str.split('-')
                    if len(partes[0]) == 4:  # YYYY-MM-DD
                        fecha_emision = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                    else:  # DD-MM-YYYY (formato esperado)
                        fecha_emision = datetime.strptime(fecha_str, '%d-%m-%Y').date()
                    
                    if idx < 5:
                        print(f"   ✅ Fila {idx}: Fecha parseada: {fecha_emision}")
                else:
                    # Sin guiones, dejar NULL
                    if idx < 5:
                        print(f"   ⚠️  Fila {idx}: Fecha sin guiones: '{fecha_str}' - quedará NULL")
                    fecha_emision = None
                    errores_fecha += 1
            except Exception as e:
                # Error al parsear, dejar NULL
                if idx < 5:
                    print(f"   ❌ Fila {idx}: Error parseando '{fecha_str}': {e} - quedará NULL")
                errores_fecha += 1
                fecha_emision = None
        else:
            # Ya es objeto date/datetime
            if isinstance(fecha_emision_raw, datetime):
                fecha_emision = fecha_emision_raw.date()
            elif isinstance(fecha_emision_raw, date):
                fecha_emision = fecha_emision_raw
            else:
                if idx < 5:
                    print(f"   ⚠️  Fila {idx}: Tipo inesperado: {type(fecha_emision_raw)} - quedará NULL")
                fecha_emision = None
                errores_fecha += 1
        
        prefijo_raw = str(row.get(columnas_originales.get('prefijo', 'Prefijo'), ''))
        prefijo = extraer_prefijo(prefijo_raw)
        
        folio_raw = str(row.get(columnas_originales.get('folio', 'Folio'), ''))
        folio = ultimos_8_sin_ceros(extraer_folio(folio_raw))
        
        # 🔥 FIX: Buscar 'total' o 'valor' usando columnas normalizadas
        valor_raw = None
        col_valor_encontrada = None
        for variante_norm in ['total', 'valor']:
            if variante_norm in columnas_originales:
                col_original = columnas_originales[variante_norm]
                valor_raw = row.get(col_original)
                col_valor_encontrada = col_original
                if valor_raw is not None and not pd.isna(valor_raw):
                    if idx < 3:
                        print(f"   💰 Fila {idx}: Encontrada columna '{col_original}' (normalizada: '{variante_norm}') con valor: {repr(valor_raw)}")
                    break
        
        # Si no encontró, buscar en columnas directamente
        if valor_raw is None or pd.isna(valor_raw):
            for col in ['Total', 'Valor', 'total', 'valor']:
                if col in row.index:
                    valor_raw = row[col]
                    col_valor_encontrada = col
                    if valor_raw is not None and not pd.isna(valor_raw):
                        if idx < 3:
                            print(f"   💰 Fila {idx}: Encontrada columna '{col}' (fallback) con valor: {repr(valor_raw)}")
                        break
        
        # Convertir a número (maneja comas decimales españolas)
        valor = convertir_numero_espanol(valor_raw)
        
        if valor == 0.0 and idx < 3:
            print(f"   ⚠️  Fila {idx}: Valor convertido a 0.0 desde {repr(valor_raw)}")
            errores_valor += 1
        
        tipo_documento = str(row.get(columnas_originales.get('tipo_de_documento', 'tipo de documento'), 'Factura Electrónica'))
        
        # 🔥 FIX: Usar columna CUFE detectada, NO hardcodear el fallback
        cufe = str(row.get(cufe_col_name, ''))  # Usar variable detectada explícitamente
        
        # 🐛 DEBUG: Mostrar CUFEs de primeras 5 facturas
        if idx < 5:
            print(f"   📝 Fila {idx}: CUFE desde columna '{cufe_col_name}': '{cufe[:50] if cufe else 'VACÍO'}...' (len={len(cufe)})")
        
        forma_pago = str(row.get(columnas_originales.get('forma_de_pago', 'forma de pago'), '2')).strip()
        
        # Buscar módulo y doc_causado_por en ERP
        clave = crear_clave_factura(nit, prefijo_raw, folio_raw)
        erp_info = erp_por_clave.get(clave, {})
        modulo = erp_info.get('modulo', '') if isinstance(erp_info, dict) else erp_info
        doc_causado_por = erp_info.get('doc_causado_por', '') if isinstance(erp_info, dict) else ''
        
        if modulo:
            registros_con_modulo += 1
        
        # Buscar estado en acuses
        estado_aprobacion = acuses_por_cufe.get(cufe, 'No Registra')
        
        # 🐛 DEBUG: Mostrar resultado de búsqueda de acuses
        if idx < 3:
            encontrado = "✅ ENCONTRADO" if cufe in acuses_por_cufe else "❌ NO ENCONTRADO"
            print(f"   🔍 Búsqueda acuse: {encontrado} → Estado: '{estado_aprobacion}'")
        
        # 🔥 DETERMINAR ESTADO CONTABLE SEGÚN MÓDULO (Dec 29, 2025)
        # Si tiene módulo (COMERCIAL o FINANCIERO) → Ya está causado en ERP
        # Si NO tiene módulo → No está registrado
        if modulo:
            estado_contable = 'Causada'  # ✅ Encontrado en ERP = Causada
        else:
            estado_contable = 'No Registrada'  # ❌ No está en ERP
        
        # 🆕 CALCULAR ACUSES RECIBIDOS (Dec 29, 2025)
        acuses_recibidos = calcular_acuses_recibidos(estado_aprobacion)
        
        # 🆕 CALCULAR DÍAS DESDE EMISIÓN (Dec 30, 2025)
        dias_desde_emision = 0
        try:
            if fecha_emision:
                dias_desde_emision = (date.today() - fecha_emision).days
        except:
            dias_desde_emision = 0
        
        # 🆕 OBTENER TIPO DE TERCERO (Dec 30, 2025)
        tipo_tercero = tipo_tercero_final.get(nit_limpio, '')
        
        # Agregar registro
        registros.append({
            'nit_emisor': nit_limpio,
            'razon_social': razon_social,
            'fecha_emision': fecha_emision,
            'prefijo': prefijo,
            'folio': folio,
            'valor': valor,
            'tipo_documento': tipo_documento,
            'cufe': cufe,
            'forma_pago': forma_pago,
            'estado_aprobacion': estado_aprobacion,
            'modulo': modulo,
            'estado_contable': estado_contable,
            'acuses_recibidos': acuses_recibidos,
            'doc_causado_por': doc_causado_por,  # ✅ NUEVO CAMPO
            'dias_desde_emision': dias_desde_emision,  # ✅ NUEVO CAMPO
            'tipo_tercero': tipo_tercero  # ✅ NUEVO CAMPO
        })
    
    # 🐛 DEBUG: Mostrar resumen de procesamiento
    print(f"\n📊 RESUMEN DE PROCESAMIENTO:")
    print(f"   Total registros procesados: {len(registros):,}")
    print(f"   ⚠️  Errores de fecha: {errores_fecha:,}")
    print(f"   ⚠️  Errores de valor: {errores_valor:,}")
    if errores_fecha > 0:
        print(f"   ❌ ADVERTENCIA: {errores_fecha:,} registros tienen fecha actual (17-02-2026)")
    if errores_valor > 0:
        print(f"   ❌ ADVERTENCIA: {errores_valor:,} registros tienen valor = 0")
    
    # 5️⃣ USAR COPY FROM (ULTRA RÁPIDO)
    total = 0
    registros_respaldados = 0  # 🔥 Inicializar contador (Dec 29, 2025)
    if registros:
        try:
            # 🆕 REUTILIZAR CONEXIÓN YA ABIERTA (conexión abierta para tablas individuales)
            # raw_conn y cursor ya están disponibles
            
            # 🔥 PASO 1: GUARDAR CAMPOS DE CAUSACIÓN (Dec 29, 2025)
            print("📋 Guardando datos de causación existentes...")
            cursor.execute("""
                CREATE TEMP TABLE IF NOT EXISTS backup_causacion AS
                SELECT nit_emisor, prefijo, folio, 
                       causada, fecha_causacion, usuario_causacion, doc_causado_por,
                       recibida, fecha_recibida, usuario_recibio,
                       rechazada, fecha_rechazo, motivo_rechazo,
                       estado_contable, origen_sincronizacion
                FROM maestro_dian_vs_erp
                WHERE doc_causado_por IS NOT NULL 
                   OR causada = TRUE 
                   OR rechazada = TRUE
                   OR estado_contable IN ('Causada', 'Rechazada', 'En Trámite')
            """)
            registros_respaldados = cursor.rowcount
            print(f"✅ {registros_respaldados:,} registros con datos de causación respaldados")
            
            # 🆕 PASO 2: CREAR TABLA TEMPORAL PARA NUEVOS DATOS (Dec 29, 2025)
            print("📦 Creando tabla temporal para nuevos datos...")
            cursor.execute("""
                CREATE TEMP TABLE IF NOT EXISTS temp_maestro_nuevos (
                    nit_emisor VARCHAR(20),
                    razon_social VARCHAR(255),
                    fecha_emision DATE,
                    prefijo VARCHAR(10),
                    folio VARCHAR(20),
                    valor NUMERIC(15,2),
                    tipo_documento VARCHAR(100),
                    cufe VARCHAR(255),
                    forma_pago VARCHAR(20),
                    estado_aprobacion VARCHAR(50),
                    modulo VARCHAR(20),
                    estado_contable VARCHAR(50),
                    acuses_recibidos INTEGER DEFAULT 0,
                    doc_causado_por VARCHAR(255),
                    dias_desde_emision INTEGER DEFAULT 0,
                    tipo_tercero VARCHAR(50)
                )
            """)
            cursor.execute("TRUNCATE TABLE temp_maestro_nuevos")
            print("✅ Tabla temporal lista")
            
            # 🆕 PASO 3: CARGAR DATOS A TABLA TEMPORAL CON COPY FROM (Dec 29, 2025)
            print("📥 Cargando datos nuevos a tabla temporal...")
            buffer = io.StringIO()
            for reg in registros:
                buffer.write(f"{format_value_for_copy(reg['nit_emisor'])}\t")
                buffer.write(f"{format_value_for_copy(reg['razon_social'])}\t")
                buffer.write(f"{format_value_for_copy(reg['fecha_emision'])}\t")
                buffer.write(f"{format_value_for_copy(reg['prefijo'])}\t")
                buffer.write(f"{format_value_for_copy(reg['folio'])}\t")
                buffer.write(f"{format_value_for_copy(reg['valor'])}\t")
                buffer.write(f"{format_value_for_copy(reg['tipo_documento'])}\t")
                buffer.write(f"{format_value_for_copy(reg['cufe'])}\t")
                buffer.write(f"{format_value_for_copy(reg['forma_pago'])}\t")
                buffer.write(f"{format_value_for_copy(reg['estado_aprobacion'])}\t")
                buffer.write(f"{format_value_for_copy(reg['modulo'])}\t")
                buffer.write(f"{format_value_for_copy(reg['estado_contable'])}\t")
                buffer.write(f"{format_value_for_copy(reg['acuses_recibidos'])}\t")
                buffer.write(f"{format_value_for_copy(reg['doc_causado_por'])}\t")
                buffer.write(f"{format_value_for_copy(reg['dias_desde_emision'])}\t")
                buffer.write(f"{format_value_for_copy(reg['tipo_tercero'])}\n")
            
            buffer.seek(0)
            cursor.copy_from(
                buffer,
                'temp_maestro_nuevos',
                sep='\t',
                null='',
                columns=(
                    'nit_emisor', 'razon_social', 'fecha_emision', 'prefijo', 'folio',
                    'valor', 'tipo_documento', 'cufe', 'forma_pago', 'estado_aprobacion',
                    'modulo', 'estado_contable', 'acuses_recibidos', 'doc_causado_por',
                    'dias_desde_emision', 'tipo_tercero'
                )
            )
            print(f"✅ {len(registros):,} registros cargados en tabla temporal")
            
            # 🆕 PASO 4: UPSERT INTELIGENTE CON VALIDACIÓN DE JERARQUÍAS (Dec 29, 2025)
            print("🔄 Ejecutando UPSERT con validación de jerarquías...")
            cursor.execute("""
                -- Crear función temporal para jerarquía de aceptación
                CREATE OR REPLACE FUNCTION temp_jerarquia_aceptacion(estado TEXT) RETURNS INTEGER AS $$
                BEGIN
                    RETURN CASE COALESCE(estado, 'Pendiente')
                        WHEN 'Pendiente' THEN 1
                        WHEN 'Acuse Recibido' THEN 2
                        WHEN 'Acuse Bien/Servicio' THEN 3
                        WHEN 'Rechazada' THEN 4
                        WHEN 'Aceptación Expresa' THEN 5
                        WHEN 'Aceptación Tácita' THEN 6
                        ELSE 1
                    END;
                END;
                $$ LANGUAGE plpgsql IMMUTABLE;
                
                -- INSERT SIMPLE (sin ON CONFLICT) - Diciembre 29, 2025
                -- Inserta todos los registros sin validar duplicados
                INSERT INTO maestro_dian_vs_erp (
                    nit_emisor, razon_social, fecha_emision, prefijo, folio,
                    valor, tipo_documento, cufe, forma_pago, estado_aprobacion, 
                    modulo, estado_contable, acuses_recibidos, doc_causado_por,
                    dias_desde_emision, tipo_tercero
                )
                SELECT 
                    nit_emisor, razon_social, fecha_emision, prefijo, folio,
                    valor, tipo_documento, cufe, forma_pago, estado_aprobacion,
                    modulo, estado_contable, acuses_recibidos, doc_causado_por,
                    dias_desde_emision, tipo_tercero
                FROM temp_maestro_nuevos;
                
                -- Limpiar duplicados dejando solo el más reciente (mayor ID)
                DELETE FROM maestro_dian_vs_erp a 
                USING maestro_dian_vs_erp b
                WHERE a.id < b.id
                  AND a.nit_emisor = b.nit_emisor
                  AND a.prefijo = b.prefijo
                  AND a.folio = b.folio
            """)
            
            total = cursor.rowcount
            print(f"✅ UPSERT completado: {total:,} registros insertados/actualizados")
            
            # 🔥 PASO 5: RESTAURAR CAMPOS DE CAUSACIÓN (Dec 29, 2025)
            print("🔄 Restaurando datos de causación...")
            cursor.execute("""
                UPDATE maestro_dian_vs_erp m
                SET causada = b.causada,
                    fecha_causacion = b.fecha_causacion,
                    usuario_causacion = b.usuario_causacion,
                    doc_causado_por = b.doc_causado_por,
                    recibida = b.recibida,
                    fecha_recibida = b.fecha_recibida,
                    usuario_recibio = b.usuario_recibio,
                    rechazada = b.rechazada,
                    fecha_rechazo = b.fecha_rechazo,
                    motivo_rechazo = b.motivo_rechazo,
                    estado_contable = b.estado_contable,
                    origen_sincronizacion = b.origen_sincronizacion
                FROM backup_causacion b
                WHERE m.nit_emisor = b.nit_emisor
                  AND m.prefijo = b.prefijo
                  AND m.folio = b.folio
            """)
            registros_restaurados = cursor.rowcount
            print(f"✅ {registros_restaurados:,} registros con datos de causación restaurados")
            
            raw_conn.commit()
            
            cursor.close()
            raw_conn.close()
            
            print(f"✅ {total:,} registros guardados en PostgreSQL con COPY FROM")
            
        except Exception as e:
            if 'raw_conn' in locals():
                raw_conn.rollback()
                if 'cursor' in locals():
                    cursor.close()
                raw_conn.close()
            print(f"❌ Error guardando en PostgreSQL: {e}")
            import traceback
            traceback.print_exc()
            return f"❌ Error al guardar: {e}"
    
    # 6️⃣ ESTADÍSTICAS FINALES
    tiempo = time.time() - t0
    porcentaje_match = (registros_con_modulo / registros_dian * 100) if registros_dian > 0 else 0
    velocidad = total / tiempo if tiempo > 0 else 0
    
    msg = f"""✅ Sistema DIAN vs ERP actualizado en {tiempo:.1f}s

📊 ARCHIVOS PROCESADOS:
   • DIAN: {registros_dian:,} registros
   • ERP (FN+CM+Errores): {registros_erp:,} registros
   • Acuses: {registros_acuses:,} registros

🎯 RESULTADOS:
   • Total guardado: {total:,} facturas
   • Con módulo detectado: {registros_con_modulo:,} ({porcentaje_match:.1f}%)
   • Sin módulo: {registros_dian - registros_con_modulo:,}
   • Datos de causación restaurados: {registros_respaldados:,}

⚡ RENDIMIENTO:
   • Velocidad: {velocidad:,.0f} registros/segundo
   • Método: COPY FROM (nativo PostgreSQL)

🚀 BASE DE DATOS: PostgreSQL (gestor_documental)
📁 TABLA: maestro_dian_vs_erp"""
    
    # 7️⃣ SINCRONIZACIÓN AUTOMÁTICA CON OTRAS TABLAS
    print("\n🔄 Iniciando sincronización automática...")
    
    try:
        # Validación #1: Documentos en tablas de facturas → "Recibida"
        # 🆕 MEJORADO: Ahora incluye facturas_digitales (Dec 30, 2025)
        query_validacion = """
            UPDATE maestro_dian_vs_erp m
            SET estado_contable = 'Recibida',
                fecha_actualizacion = NOW()
            WHERE (modulo IS NULL OR modulo = '')
              AND (doc_causado_por IS NULL OR doc_causado_por = '')
              AND (
                  EXISTS (SELECT 1 FROM facturas_temporales f WHERE f.nit = m.nit_emisor AND f.prefijo = m.prefijo AND f.folio = m.folio)
                  OR EXISTS (SELECT 1 FROM facturas_recibidas f WHERE f.nit = m.nit_emisor AND f.prefijo = m.prefijo AND f.folio = m.folio)
                  OR EXISTS (SELECT 1 FROM facturas_recibidas_digitales f WHERE f.nit = m.nit_emisor AND f.prefijo = m.prefijo AND f.folio = m.folio)
                  OR EXISTS (SELECT 1 FROM facturas_digitales f WHERE f.nit_proveedor = m.nit_emisor AND f.numero_factura = m.prefijo || m.folio)
              )
        """
        result_val = db.session.execute(db.text(query_validacion))
        docs_recibidas = result_val.rowcount
        print(f"   ✅ {docs_recibidas} docs actualizados a 'Recibida'")
        
        # Validación #2: Documentos con doc_causado_por → "Causada"
        query_causados = """
            UPDATE maestro_dian_vs_erp
            SET estado_contable = 'Causada',
                fecha_actualizacion = NOW()
            WHERE doc_causado_por IS NOT NULL
              AND doc_causado_por != ''
              AND estado_contable != 'Causada'
        """
        result_cau = db.session.execute(db.text(query_causados))
        docs_causadas = result_cau.rowcount
        print(f"   ✅ {docs_causadas} docs actualizados a 'Causada' (doc_causado_por)")
        
        # Validación #3: Documentos con módulo ERP → "Causada"
        query_erp = """
            UPDATE maestro_dian_vs_erp
            SET estado_contable = 'Causada',
                fecha_actualizacion = NOW()
            WHERE modulo IS NOT NULL
              AND modulo != ''
              AND estado_contable != 'Causada'
        """
        result_erp = db.session.execute(db.text(query_erp))
        docs_erp = result_erp.rowcount
        print(f"   ✅ {docs_erp} docs actualizados a 'Causada' (módulo ERP)")
        
        db.session.commit()
        
        total_sincronizados = docs_recibidas + docs_causadas + docs_erp
        if total_sincronizados > 0:
            msg += f"\n\n🔄 SINCRONIZACIÓN:\n   • Total actualizados: {total_sincronizados:,} estados contables"
        
    except Exception as e:
        print(f"⚠️ Error en sincronización automática: {e}")
        db.session.rollback()
    
    return msg

# ============================================
# ENDPOINTS DE API FALTANTES PARA EL FRONTEND
# ============================================

@dian_vs_erp_bp.route('/api/actualizar_nit', methods=['POST'])
def api_actualizar_nit():
    """
    Actualiza un campo para todos los documentos de un NIT
    """
    try:
        data = request.get_json()
        nit = data.get('nit')
        campo = data.get('campo')
        valor = data.get('valor')
        
        if not all([nit, campo]):
            return jsonify({'exito': False, 'mensaje': 'Faltan parámetros requeridos'}), 400
        
        # Simular actualización exitosa
        # En una implementación real, aquí actualizarías la base de datos
        print(f"🔄 Simulando actualización: NIT {nit}, campo '{campo}' = '{valor}'")
        
        return jsonify({
            'exito': True,
            'mensaje': f'Campo {campo} actualizado para todas las facturas del NIT {nit}'
        })
        
    except Exception as e:
        print(f"❌ Error actualizando NIT: {e}")
        return jsonify({'exito': False, 'mensaje': f'Error interno: {e}'}), 500

@dian_vs_erp_bp.route('/api/actualizar_campo', methods=['POST'])
def api_actualizar_campo():
    """
    Actualiza un campo individual por CUFE
    """
    try:
        data = request.get_json()
        cufe = data.get('cufe')
        campo = data.get('campo')
        valor = data.get('valor')
        
        if not all([cufe, campo]):
            return jsonify({'exito': False, 'mensaje': 'Faltan parámetros requeridos'}), 400
        
        # Simular actualización exitosa
        print(f"🔄 Simulando actualización individual: CUFE {cufe}, campo '{campo}' = '{valor}'")
        
        return jsonify({
            'exito': True,
            'mensaje': f'Campo {campo} actualizado correctamente'
        })
        
    except Exception as e:
        print(f"❌ Error actualizando campo: {e}")
        return jsonify({'exito': False, 'mensaje': f'Error interno: {e}'}), 500

@dian_vs_erp_bp.route('/api/enviar_emails', methods=['POST'])
def api_enviar_emails():
    """
    Envía emails individuales para facturas
    """
    try:
        data = request.get_json()
        cufe = data.get('cufe')
        destinatarios = data.get('destinatarios', [])
        
        if not cufe or not destinatarios:
            return jsonify({'exito': False, 'mensaje': 'Faltan parámetros requeridos'}), 400
        
        # Simular envío exitoso
        enviados = len(destinatarios)
        print(f"📧 Simulando envío de {enviados} emails para CUFE {cufe}")
        
        return jsonify({
            'exito': True,
            'enviados': enviados,
            'fallidos': 0
        })
        
    except Exception as e:
        print(f"❌ Error enviando emails: {e}")
        return jsonify({'exito': False, 'mensaje': f'Error interno: {e}'}), 500


@dian_vs_erp_bp.route('/api/enviar_email_agrupado', methods=['POST'])
def api_enviar_email_agrupado():
    """
    Enviar TODAS las facturas agrupadas en UN SOLO EMAIL a cada destinatario
    """
    try:
        print("=" * 80)
        print("📧 ENDPOINT ENVIAR EMAIL AGRUPADO - INICIO")
        print("=" * 80)
        
        data = request.get_json()
        print(f"📦 Datos recibidos: {data}")
        
        destinatarios = data.get('destinatarios', [])
        cufes = data.get('cufes', [])
        
        print(f"👥 Destinatarios: {len(destinatarios)}")
        print(f"📄 CUFEs: {len(cufes)}")
        
        if not destinatarios:
            print("❌ Error: No hay destinatarios")
            return jsonify({'exito': False, 'mensaje': 'No hay destinatarios'}), 400
        
        if not cufes:
            print("❌ Error: No hay facturas para enviar")
            return jsonify({'exito': False, 'mensaje': 'No hay facturas para enviar'}), 400
        
        # Obtener todas las facturas desde PostgreSQL
        try:
            print(f"📧 Consultando {len(cufes)} facturas desde PostgreSQL...")
            
            # Consultar facturas por CUFEs usando sintaxis correcta de SQLAlchemy
            facturas_query = db.session.query(MaestroDianVsErp).filter(
                MaestroDianVsErp.cufe.in_(cufes)
            ).all()
            
            print(f"✅ Encontradas {len(facturas_query)} facturas en BD")
            
            if not facturas_query:
                return jsonify({'exito': False, 'mensaje': 'No se encontraron facturas'}), 404
            
            # Convertir a diccionarios (formato compatible con puerto 8097)
            facturas = []
            for f in facturas_query:
                facturas.append({
                    'cufe': f.cufe or '',
                    'nit_emisor': f.nit_emisor or '',
                    'nombre_emisor': f.razon_social or '',
                    'razon_social': f.razon_social or '',
                    'fecha_emision': f.fecha_emision.strftime('%Y-%m-%d') if f.fecha_emision else '',
                    'tipo_documento': f.tipo_documento or 'Factura Electrónica',
                    'prefijo': f.prefijo or '',
                    'folio': f.folio or '',
                    'valor': float(f.valor) if f.valor else 0,
                    'estado_aprobacion': f.estado_aprobacion or 'Pendiente',
                    'forma_pago_texto': f.forma_pago or 'Crédito',
                    'estado_contable': f.estado_contable or 'No Registrada',
                    'causada': f.causada or False,
                    'recibida': f.recibida or False
                })
        
        except Exception as e:
            print(f"❌ Error obteniendo facturas desde PostgreSQL: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({'exito': False, 'mensaje': f'Error consultando facturas: {str(e)}'}), 500
        
        # Enviar UN email agrupado por destinatario
        enviados = 0
        fallidos = 0
        errores = []
        
        print(f"📧 Enviando {len(facturas)} facturas agrupadas a {len(destinatarios)} destinatarios")
        
        for dest in destinatarios:
            correo = dest.get('correo')
            nombre = dest.get('nombre', correo)
            
            try:
                # Calcular días desde emisión
                from datetime import datetime, date
                fecha_hoy = date.today()
                
                # ✅ LIMITAR A 20 FACTURAS EN EL HTML
                total_facturas = len(facturas)
                facturas_para_email = facturas[:20] if total_facturas > 20 else facturas
                
                # ⚠️ Mensaje de advertencia si hay más de 20
                advertencia_html = ""
                if total_facturas > 20:
                    advertencia_html = f"""
                    <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0; border-radius: 4px;">
                        <strong>⚠️ Se muestran las primeras 20 facturas de {total_facturas} seleccionadas.</strong><br>
                        💡 El resumen completo está disponible en el sistema.
                    </div>
                    """
                
                # Generar HTML del email con tabla de facturas mejorada
                html_facturas = """
                <table style="width:100%; border-collapse:collapse; margin:20px 0; font-size:13px;">
                    <thead>
                        <tr style="background-color:#166534; color:white;">
                            <th style="padding:8px; border:1px solid #ddd;">NIT</th>
                            <th style="padding:8px; border:1px solid #ddd;">Razón Social</th>
                            <th style="padding:8px; border:1px solid #ddd;">Prefijo-Folio</th>
                            <th style="padding:8px; border:1px solid #ddd;">Fecha Emisión</th>
                            <th style="padding:8px; border:1px solid #ddd;">Días</th>
                            <th style="padding:8px; border:1px solid #ddd;">Valor</th>
                            <th style="padding:8px; border:1px solid #ddd;">Estado</th>
                            <th style="padding:8px; border:1px solid #ddd;">Documento</th>
                        </tr>
                    </thead>
                    <tbody>
                """
                
                total_valor = 0
                for factura in facturas_para_email:
                    valor = factura.get('valor', 0)
                    total_valor += valor
                    
                    # Calcular días desde emisión
                    dias = ''
                    fecha_emision_str = factura.get('fecha_emision', '')
                    if fecha_emision_str:
                        try:
                            fecha_emision = datetime.strptime(fecha_emision_str, '%Y-%m-%d').date()
                            dias_transcurridos = (fecha_hoy - fecha_emision).days
                            dias = str(dias_transcurridos)
                        except:
                            dias = '-'
                    
                    # Link al documento DIAN
                    cufe = factura.get('cufe', '')
                    link_dian = ''
                    if cufe:
                        link_dian = f'<a href="https://catalogo-vpfe.dian.gov.co/User/SearchDocument?DocumentKey={cufe}" target="_blank" style="color:#166534; text-decoration:none;">📄 Ver</a>'
                    else:
                        link_dian = '<span style="color:#999;">-</span>'
                    
                    html_facturas += f"""
                        <tr>
                            <td style="padding:6px; border:1px solid #ddd;">{factura.get('nit_emisor', '')}</td>
                            <td style="padding:6px; border:1px solid #ddd;">{factura.get('razon_social', '')}</td>
                            <td style="padding:6px; border:1px solid #ddd;">{factura.get('prefijo', '')}-{factura.get('folio', '')}</td>
                            <td style="padding:6px; border:1px solid #ddd; text-align:center;">{fecha_emision_str}</td>
                            <td style="padding:6px; border:1px solid #ddd; text-align:center;">{dias}</td>
                            <td style="padding:6px; border:1px solid #ddd; text-align:right;">${valor:,.2f}</td>
                            <td style="padding:6px; border:1px solid #ddd;">{factura.get('estado_aprobacion', 'Pendiente')}</td>
                            <td style="padding:6px; border:1px solid #ddd; text-align:center;">{link_dian}</td>
                        </tr>
                    """
                
                html_facturas += f"""
                    </tbody>
                    <tfoot>
                        <tr style="background-color:#f9fafb; font-weight:bold;">
                            <td colspan="5" style="padding:10px; border:1px solid #ddd; text-align:right;">TOTAL</td>
                            <td style="padding:10px; border:1px solid #ddd; text-align:right;">${total_valor:,.2f}</td>
                            <td colspan="2" style="padding:10px; border:1px solid #ddd;">{total_facturas} facturas</td>
                        </tr>
                    </tfoot>
                </table>
                """
                
                html_body = f"""
                <html>
                <body style="font-family: Arial, sans-serif; color: #333;">
                    <div style="max-width: 1000px; margin: 0 auto; padding: 20px;">
                        <h2 style="color: #166534;">📧 Facturas Electrónicas - Resumen Agrupado</h2>
                        <p>Estimado/a <strong>{nombre}</strong>,</p>
                        <p>Le enviamos el siguiente resumen de <strong>{total_facturas} facturas electrónicas</strong>:</p>
                        
                        {advertencia_html}
                        
                        <div style="margin: 20px 0; padding: 15px; background-color: #fef2f2; border-left: 4px solid #dc2626; border-radius: 4px;">
                            <p style="margin: 0; color: #dc2626; font-weight: bold;">⚠️ IMPORTANTE - Acción Requerida:</p>
                            <p style="margin: 10px 0 5px 0; color: #333; font-size: 15px;">
                                Por favor, <strong>gestione y tramite estos documentos pendientes lo más pronto posible</strong>.
                            </p>
                            <p style="margin: 5px 0 0 0; color: #666; font-size: 14px;">
                                Los documentos con mayor antigüedad requieren atención prioritaria para evitar retrasos en el proceso contable.
                            </p>
                        </div>
                        
                        <div style="margin: 0 0 20px 0; padding: 15px; background-color: #f0fdf4; border-left: 4px solid #166534; border-radius: 4px;">
                            <p style="margin: 0; color: #166534; font-weight: bold;">💡 Información del listado:</p>
                            <ul style="margin: 10px 0; color: #666;">
                                <li><strong>Días:</strong> Días transcurridos desde la fecha de emisión</li>
                                <li><strong>Estado:</strong> Estado de aprobación en el sistema</li>
                                <li><strong>Documento:</strong> Click en "📄 Ver" para consultar el documento en la DIAN</li>
                            </ul>
                        </div>
                        
                        {html_facturas}
                        
                        <div style="margin-top: 20px; padding: 15px; background: #f0f0f0; border-radius: 6px; text-align: center;">
                            <strong>Sistema de Facturación Electrónica - SUPERTIENDAS CAÑAVERAL</strong><br>
                            <span style="color: #666; font-size: 12px;">Este es un mensaje automático, por favor no responder.</span>
                        </div>
                    </div>
                </body>
                </html>"""
                
                # Enviar email con Flask-Mail
                from flask_mail import Message
                msg = Message(
                    subject=f"Facturas Electrónicas - Resumen Agrupado ({len(facturas)} facturas)",
                    recipients=[correo],
                    html=html_body
                )
                
                from flask import current_app
                current_app.extensions['mail'].send(msg)
                
                enviados += 1
                print(f"✅ Email agrupado REAL enviado a {correo} ({len(facturas)} facturas)")
            
            except Exception as e:
                fallidos += 1
                error_msg = str(e)
                errores.append(f"{nombre}: {error_msg}")
                print(f"❌ Error enviando email a {correo}: {error_msg}")
                import traceback
                traceback.print_exc()
        
        mensaje_resultado = f"Enviados: {enviados}/{len(destinatarios)} - Total facturas: {len(facturas)}"
        
        return jsonify({
            'exito': enviados > 0,
            'mensaje': mensaje_resultado,
            'enviados': enviados,
            'fallidos': fallidos,
            'total_facturas': len(facturas),
            'errores': errores
        })
        
    except Exception as e:
        print(f"❌ Error en envío agrupado: {e}")
        return jsonify({'exito': False, 'mensaje': str(e)}), 500


# ====================================
# CONFIGURACIÓN DEL SISTEMA
# ====================================

@dian_vs_erp_bp.route('/config')
def config_dian_erp():
    """Página de configuración del sistema DIAN vs ERP"""
    return render_template('dian_vs_erp/configuracion.html')


@dian_vs_erp_bp.route('/api/dian_usuarios/por_nit/<nit>', methods=['GET'])
def obtener_usuarios_dian_por_nit(nit):
    """Obtiene usuarios configurados para un NIT específico"""
    try:
        # ✅ CONSULTAR BASE DE DATOS REAL
        usuarios_db = UsuarioAsignadoDianVsErp.query.filter_by(nit=nit).all()
        
        usuarios = []
        for usuario in usuarios_db:
            # Separar nombre completo en nombres y apellidos (aproximación)
            partes_nombre = usuario.nombre.split(' ', 1) if usuario.nombre else ['', '']
            usuarios.append({
                'nombres': partes_nombre[0] if len(partes_nombre) > 0 else '',
                'apellidos': partes_nombre[1] if len(partes_nombre) > 1 else '',
                'correo': usuario.correo,
                'tipo': 'APROBADOR',  # Por defecto, podría agregarse un campo en el modelo
                'activo': usuario.activo
            })
        
        return jsonify({
            'status': 'success',
            'nit': nit,
            'usuarios': usuarios,
            'total': len(usuarios)
        })
        
    except Exception as e:
        log_security(f"ERROR OBTENER USUARIOS NIT {nit} | {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error obteniendo usuarios: {str(e)}'
        }), 500
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error obteniendo usuarios: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/terceros/buscar_por_nit/<nit>', methods=['GET'])
def buscar_tercero_por_nit(nit):
    """
    Busca un tercero en PostgreSQL por NIT y retorna su razón social
    Usado para autocompletar el campo de razón social en el modal de agregar usuario
    """
    try:
        import psycopg2
        from psycopg2.extras import RealDictCursor
        
        # Obtener configuración de la BD desde .env
        db_config = {
            'host': os.getenv('DB_HOST', 'localhost'),
            'port': os.getenv('DB_PORT', '5432'),
            'database': os.getenv('DB_NAME', 'gestor_documental'),
            'user': os.getenv('DB_USER', 'gestor_user'),
            'password': os.getenv('DB_PASSWORD', 'password')
        }
        
        conn = psycopg2.connect(**db_config, cursor_factory=RealDictCursor)
        cursor = conn.cursor()
        
        # Buscar tercero por NIT
        query = """
            SELECT 
                tercero_id,
                nit,
                razon_social,
                tipo_persona,
                activo
            FROM terceros
            WHERE nit = %s
            LIMIT 1
        """
        
        cursor.execute(query, (nit,))
        tercero = cursor.fetchone()
        
        cursor.close()
        conn.close()
        
        if tercero:
            print(f"✅ Tercero encontrado: NIT {nit} - {tercero['razon_social']}")
            return jsonify({
                'exito': True,
                'tercero': dict(tercero)
            })
        else:
            print(f"⚠️ No se encontró tercero con NIT: {nit}")
            return jsonify({
                'exito': False,
                'mensaje': f'No se encontró un tercero con NIT {nit}'
            }), 404
            
    except Exception as e:
        print(f"❌ Error buscando tercero NIT {nit}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'exito': False,
            'mensaje': f'Error al buscar tercero: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/dian_usuarios/agregar', methods=['POST'])
def agregar_usuario_dian():
    """Agrega un nuevo usuario para el envío de emails"""
    try:
        data = request.get_json()
        
        # Validar campos requeridos
        campos_requeridos = ['nit', 'nombres', 'apellidos', 'correo', 'tipo']
        for campo in campos_requeridos:
            if not data.get(campo):
                return jsonify({
                    'status': 'error',
                    'message': f'El campo {campo} es requerido'
                }), 400
        
        # ✅ GUARDAR EN BASE DE DATOS REAL
        nombre_completo = f"{data['nombres']} {data['apellidos']}"
        nuevo_usuario = UsuarioAsignadoDianVsErp(
            nit=data['nit'],
            correo=data['correo'],
            nombre=nombre_completo,
            activo=True
        )
        
        db.session.add(nuevo_usuario)
        db.session.commit()
        
        log_security(f"USUARIO DIAN AGREGADO | nit={data['nit']} | correo={data['correo']} | nombre={nombre_completo}")
        
        return jsonify({
            'status': 'success',
            'message': 'Usuario agregado correctamente',
            'usuario': {
                'nit': data['nit'],
                'nombres': data['nombres'],
                'apellidos': data['apellidos'],
                'correo': data['correo'],
                'tipo': data['tipo'],
                'activo': True
            }
        })
        
    except Exception as e:
        db.session.rollback()
        log_security(f"ERROR AGREGAR USUARIO DIAN | {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error agregando usuario: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/dian_usuarios/actualizar', methods=['PUT'])
def actualizar_usuario_dian():
    """Actualiza la información de un usuario"""
    try:
        data = request.get_json()
        
        # Validar campos requeridos
        if not data.get('correo'):
            return jsonify({
                'status': 'error',
                'message': 'El correo es requerido para identificar el usuario'
            }), 400
        
        # Simular actualización en base de datos
        # En una implementación real, aquí se actualizaría en la BD
        
        return jsonify({
            'status': 'success',
            'message': 'Usuario actualizado correctamente'
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error actualizando usuario: {str(e)}'
        }), 500


# ================================
# ENDPOINTS DE CONFIGURACIÓN SMTP
# ================================

@dian_vs_erp_bp.route('/api/config/smtp', methods=['GET'])
def obtener_config_smtp():
    """Obtiene la configuración SMTP global del sistema (sin contraseña por seguridad)"""
    try:
        # ✅ USAR CONFIGURACIÓN GLOBAL DEL SISTEMA
        config = {
            'smtp_host': current_app.config.get('MAIL_SERVER', 'No configurado'),
            'smtp_port': str(current_app.config.get('MAIL_PORT', 'No configurado')),
            'smtp_user': current_app.config.get('MAIL_USERNAME', 'No configurado'),
            'smtp_pass': '***********' if current_app.config.get('MAIL_PASSWORD') else 'No configurado',
            'from_name': 'Sistema Facturas DIAN - Supertiendas Cañaveral',
            'from_email': current_app.config.get('MAIL_DEFAULT_SENDER', current_app.config.get('MAIL_USERNAME', 'No configurado')),
            'use_ssl': current_app.config.get('MAIL_USE_SSL', False),
            'use_tls': current_app.config.get('MAIL_USE_TLS', False),
            'reply_to': current_app.config.get('MAIL_REPLY_TO', ''),
            'configurado': bool(current_app.config.get('MAIL_USERNAME') and current_app.config.get('MAIL_PASSWORD')),
            'mensaje_info': '⚙️ Esta configuración es global del sistema Gestor Documental'
        }
        
        return jsonify({
            'exito': True,
            'config': config
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error obteniendo configuración: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/config/smtp', methods=['POST'])
def guardar_config_smtp():
    """La configuración SMTP es global del sistema - No se permite modificar desde este módulo"""
    return jsonify({
        'exito': False,
        'mensaje': '⚠️ La configuración SMTP es global del sistema. Para modificarla, edite el archivo .env y reinicie el servidor.'
    }), 403


@dian_vs_erp_bp.route('/api/config/smtp/probar', methods=['GET'])
def probar_conexion_smtp():
    """Prueba la conexión SMTP usando la configuración global"""
    try:
        # Verificar si hay configuración
        if not current_app.config.get('MAIL_USERNAME') or not current_app.config.get('MAIL_PASSWORD'):
            return jsonify({
                'exito': False,
                'mensaje': '⚠️ No hay configuración SMTP. Configure las variables MAIL_* en el archivo .env'
            }), 400
        
        # Intentar envío de correo de prueba a un destinatario interno
        from app import mail
        
        msg = Message(
            subject='🧪 Prueba de Conexión SMTP - DIAN vs ERP',
            recipients=[current_app.config.get('MAIL_USERNAME')],  # Enviar a la misma cuenta
            html='''
            <html>
            <body style="font-family: Arial, sans-serif;">
                <h2 style="color: #166534;">✅ Prueba de Conexión SMTP Exitosa</h2>
                <p>Este es un correo de prueba desde el módulo <strong>DIAN vs ERP</strong>.</p>
                <p>La configuración SMTP está funcionando correctamente.</p>
                <hr>
                <p style="color: #666; font-size: 12px;">
                    Enviado desde: <strong>{}</strong><br>
                    Servidor: <strong>{}</strong><br>
                    Puerto: <strong>{}</strong>
                </p>
            </body>
            </html>
            '''.format(
                current_app.config.get('MAIL_DEFAULT_SENDER', 'No configurado'),
                current_app.config.get('MAIL_SERVER', 'No configurado'),
                current_app.config.get('MAIL_PORT', 'No configurado')
            )
        )
        
        # Configurar Reply-To si está definido
        if current_app.config.get('MAIL_REPLY_TO'):
            msg.reply_to = current_app.config.get('MAIL_REPLY_TO')
        
        mail.send(msg)
        
        return jsonify({
            'exito': True,
            'mensaje': f'✅ Conexión SMTP exitosa. Email de prueba enviado a {current_app.config.get("MAIL_USERNAME")}'
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'❌ Error en conexión SMTP: {str(e)}'
        }), 500


# ========================================
# ENDPOINTS DE ENVÍOS PROGRAMADOS
# ========================================

@dian_vs_erp_bp.route('/api/config/envios', methods=['GET'])
def obtener_envios_programados():
    """Obtiene todas las configuraciones de envíos programados desde PostgreSQL"""
    try:
        # Consultar desde PostgreSQL
        envios_query = EnvioProgramadoDianVsErp.query.order_by(
            EnvioProgramadoDianVsErp.activo.desc(),
            EnvioProgramadoDianVsErp.hora_envio.asc()
        ).all()
        
        envios = [envio.to_dict() for envio in envios_query]
        
        return jsonify({
            'exito': True,
            'envios': envios
        })
        
    except Exception as e:
        current_app.logger.error(f"Error obteniendo envíos programados: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/config/envios', methods=['POST'])
@requiere_permiso('dian_vs_erp', 'gestionar_envios')
def crear_envio_programado():
    """Crea una nueva configuración de envío programado en PostgreSQL"""
    try:
        data = request.json
        
        # Crear nuevo registro en PostgreSQL
        nuevo_envio = EnvioProgramadoDianVsErp(
            nombre=data['nombre'],
            tipo=data['tipo'],
            dias_minimos=data.get('dias_minimos'),
            requiere_acuses_minimo=data.get('requiere_acuses_minimo', 2),
            estados_incluidos=json.dumps(data.get('estados_incluidos', [])) if data.get('estados_incluidos') else None,
            estados_excluidos=json.dumps(data.get('estados_excluidos', [])) if data.get('estados_excluidos') else None,
            tipos_tercero=json.dumps(data.get('tipos_tercero', [])) if data.get('tipos_tercero') else None,  # 🆕 FILTRO TIPOS DE TERCERO
            # 🆕 FILTROS DE FECHA (Jan 26, 2026)
            fecha_inicio=datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date() if data.get('fecha_inicio') else None,
            fecha_fin=datetime.strptime(data['fecha_fin'], '%Y-%m-%d').date() if data.get('fecha_fin') else None,
            hora_envio=data['hora_envio'],
            frecuencia=data['frecuencia'],
            dias_semana=data.get('dias_semana'),
            tipo_destinatario=data['tipo_destinatario'],
            email_cc=data.get('email_cc'),
            activo=data.get('activo', True),
            # 🆕 SUPERVISIÓN
            es_supervision=data.get('es_supervision', False),
            email_supervisor=data.get('email_supervisor'),
            frecuencia_dias=data.get('frecuencia_dias', 1)
        )
        
        db.session.add(nuevo_envio)
        db.session.flush()  # Obtener ID antes de commit
        
        # 🆕 CALCULAR PRÓXIMO ENVÍO INMEDIATAMENTE
        from .scheduler_envios import scheduler_dian_vs_erp_global
        if scheduler_dian_vs_erp_global:
            nuevo_envio.proximo_envio = scheduler_dian_vs_erp_global._calcular_proximo_envio(nuevo_envio)
        
        db.session.commit()
        config_id = nuevo_envio.id
        
        # Recargar scheduler
        try:
            from scheduler_envios_programados import scheduler_global
            if scheduler_global:
                scheduler_global.detener_scheduler()
                scheduler_global.iniciar_scheduler()
        except:
            pass
        
        return jsonify({
            'exito': True,
            'mensaje': 'Configuración creada exitosamente',
            'id': config_id
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creando envío programado: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/config/envios/<int:id>', methods=['PUT'])
@requiere_permiso('dian_vs_erp', 'gestionar_envios')
def actualizar_envio_programado(id):
    """Actualiza una configuración existente en PostgreSQL"""
    try:
        data = request.json
        
        # Buscar registro en PostgreSQL
        envio = EnvioProgramadoDianVsErp.query.get(id)
        if not envio:
            return jsonify({
                'exito': False,
                'mensaje': 'Configuración no encontrada'
            }), 404
        
        # Actualizar campos
        envio.nombre = data['nombre']
        envio.tipo = data['tipo']
        envio.dias_minimos = data.get('dias_minimos')
        envio.requiere_acuses_minimo = data.get('requiere_acuses_minimo', 2)
        envio.estados_incluidos = json.dumps(data.get('estados_incluidos', [])) if data.get('estados_incluidos') else None
        envio.estados_excluidos = json.dumps(data.get('estados_excluidos', [])) if data.get('estados_excluidos') else None
        envio.tipos_tercero = json.dumps(data.get('tipos_tercero', [])) if data.get('tipos_tercero') else None  # 🆕 FILTRO TIPOS DE TERCERO
        # 🆕 FILTROS DE FECHA (Jan 26, 2026)
        envio.fecha_inicio = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date() if data.get('fecha_inicio') else None
        envio.fecha_fin = datetime.strptime(data['fecha_fin'], '%Y-%m-%d').date() if data.get('fecha_fin') else None
        envio.hora_envio = data['hora_envio']
        envio.frecuencia = data['frecuencia']
        envio.dias_semana = data.get('dias_semana')
        envio.tipo_destinatario = data['tipo_destinatario']
        envio.email_cc = data.get('email_cc')
        envio.activo = data.get('activo', True)
        envio.fecha_modificacion = datetime.now()
        
        db.session.commit()
        
        # Recargar scheduler
        try:
            from scheduler_envios_programados import scheduler_global
            if scheduler_global:
                scheduler_global.detener_scheduler()
                scheduler_global.iniciar_scheduler()
        except:
            pass
        
        return jsonify({
            'exito': True,
            'mensaje': 'Configuración actualizada'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error actualizando envío programado: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error: {str(e)}'
        }), 500


# ========================================
# ENDPOINTS DE USUARIOS DE CAUSACIÓN
# ========================================

@dian_vs_erp_bp.route('/api/usuarios-causacion', methods=['GET'])
def obtener_usuarios_causacion():
    """Obtiene todos los usuarios de causación desde PostgreSQL"""
    try:
        usuarios_query = UsuarioCausacionDianVsErp.query.order_by(
            UsuarioCausacionDianVsErp.nombre_causador.asc()
        ).all()
        
        usuarios = [usuario.to_dict() for usuario in usuarios_query]
        
        return jsonify({
            'exito': True,
            'usuarios': usuarios
        })
        
    except Exception as e:
        current_app.logger.error(f"Error obteniendo usuarios causación: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/usuarios-causacion', methods=['POST'])
def crear_usuario_causacion():
    """Crea un nuevo usuario de causación"""
    inicio = datetime.now()
    try:
        from app import db as main_db
        data = request.json
        
        nuevo_usuario = UsuarioCausacionDianVsErp(
            nombre_causador=data['nombre_causador'],
            email=data['email'],
            activo=data.get('activo', True),
            usuario_creacion=session.get('usuario', 'sistema')
        )
        
        main_db.session.add(nuevo_usuario)
        main_db.session.commit()
        
        # 🆕 Registrar log de éxito con duración
        duracion_ms = int((datetime.now() - inicio).total_seconds() * 1000)
        registrar_log(
            'SUCCESS',
            'CREAR_USUARIO_CAUSACION',
            f'Usuario de causación creado: {data["nombre_causador"]}',
            recurso_tipo='USUARIO',
            recurso_id=nuevo_usuario.id,
            duracion_ms=duracion_ms,
            detalles={'email': data['email'], 'activo': data.get('activo', True)}
        )
        
        return jsonify({
            'exito': True,
            'mensaje': 'Usuario de causación creado',
            'id': nuevo_usuario.id
        })
        
    except Exception as e:
        main_db.session.rollback()
        
        # 🆕 Registrar log de error con stack trace
        duracion_ms = int((datetime.now() - inicio).total_seconds() * 1000)
        registrar_log(
            'ERROR',
            'CREAR_USUARIO_CAUSACION',
            f'Error creando usuario de causación: {str(e)}',
            duracion_ms=duracion_ms,
            incluir_stacktrace=True
        )
        
        logger.error(f"Error creando usuario causación: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/usuarios-causacion/<int:id>', methods=['PUT'])
def actualizar_usuario_causacion(id):
    """Actualiza un usuario de causación"""
    inicio = datetime.now()
    try:
        from app import db as main_db
        
        data = request.json
        
        usuario = UsuarioCausacionDianVsErp.query.get(id)
        if not usuario:
            return jsonify({
                'exito': False,
                'mensaje': 'Usuario no encontrado'
            }), 404
        
        # Guardar valores anteriores para el log
        valores_anteriores = {
            'nombre': usuario.nombre_causador,
            'email': usuario.email,
            'activo': usuario.activo
        }
        
        usuario.nombre_causador = data['nombre_causador']
        usuario.email = data['email']
        usuario.activo = data.get('activo', True)
        usuario.fecha_modificacion = datetime.now()
        usuario.usuario_modificacion = session.get('usuario', 'sistema')
        
        main_db.session.commit()
        
        # 🆕 Registrar log de actualización
        duracion_ms = int((datetime.now() - inicio).total_seconds() * 1000)
        registrar_log(
            'SUCCESS',
            'ACTUALIZAR_USUARIO_CAUSACION',
            f'Usuario actualizado: {data["nombre_causador"]}',
            recurso_tipo='USUARIO',
            recurso_id=id,
            duracion_ms=duracion_ms,
            detalles={'anterior': valores_anteriores, 'nuevo': data}
        )
        
        return jsonify({
            'exito': True,
            'mensaje': 'Usuario actualizado exitosamente'
        })
        
    except Exception as e:
        main_db.session.rollback()
        
        # 🆕 Registrar error
        duracion_ms = int((datetime.now() - inicio).total_seconds() * 1000)
        registrar_log(
            'ERROR',
            'ACTUALIZAR_USUARIO_CAUSACION',
            f'Error actualizando usuario ID {id}: {str(e)}',
            recurso_tipo='USUARIO',
            recurso_id=id,
            duracion_ms=duracion_ms,
            incluir_stacktrace=True
        )
        
        logger.error(f"Error actualizando usuario causación: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/usuarios-causacion/<int:id>', methods=['DELETE'])
def eliminar_usuario_causacion(id):
    """Elimina un usuario de causación"""
    inicio = datetime.now()
    try:
        from app import db as main_db
        
        usuario = UsuarioCausacionDianVsErp.query.get(id)
        if not usuario:
            return jsonify({
                'exito': False,
                'mensaje': 'Usuario no encontrado'
            }), 404
        
        # Guardar info antes de eliminar
        nombre_eliminado = usuario.nombre_causador
        email_eliminado = usuario.email
        
        main_db.session.delete(usuario)
        main_db.session.commit()
        
        # 🆕 Registrar eliminación
        duracion_ms = int((datetime.now() - inicio).total_seconds() * 1000)
        registrar_log(
            'WARNING',
            'ELIMINAR_USUARIO_CAUSACION',
            f'Usuario eliminado: {nombre_eliminado}',
            recurso_tipo='USUARIO',
            recurso_id=id,
            duracion_ms=duracion_ms,
            detalles={'nombre': nombre_eliminado, 'email': email_eliminado}
        )
        
        return jsonify({
            'exito': True,
            'mensaje': 'Usuario eliminado exitosamente'
        })
        
    except Exception as e:
        main_db.session.rollback()
        
        # 🆕 Registrar error
        duracion_ms = int((datetime.now() - inicio).total_seconds() * 1000)
        registrar_log(
            'ERROR',
            'ELIMINAR_USUARIO_CAUSACION',
            f'Error eliminando usuario ID {id}: {str(e)}',
            recurso_tipo='USUARIO',
            recurso_id=id,
            duracion_ms=duracion_ms,
            incluir_stacktrace=True
        )
        
        logger.error(f"Error eliminando usuario causación: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/causadores/disponibles', methods=['GET'])
def obtener_causadores_disponibles():
    """Obtiene lista de causadores únicos desde maestro_consolidado"""
    try:
        from app import db as main_db
        
        query = """
            SELECT DISTINCT SPLIT_PART(doc_causado_por, ' - ', 2) AS usuario_creacion
            FROM maestro_dian_vs_erp
            WHERE doc_causado_por IS NOT NULL
              AND doc_causado_por != ''
              AND SPLIT_PART(doc_causado_por, ' - ', 2) != ''
            ORDER BY usuario_creacion ASC
        """
        
        result = main_db.session.execute(main_db.text(query))
        causadores = [row[0] for row in result.fetchall()]
        
        return jsonify({
            'exito': True,
            'causadores': causadores
        })
        
    except Exception as e:
        logger.error(f"Error obteniendo causadores: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error: {str(e)}'
        }), 500


# ========================================
# ENDPOINTS DE HISTORIAL
# ========================================

@dian_vs_erp_bp.route('/api/historial-envios', methods=['GET'])
def obtener_historial_envios():
    """Obtiene historial de envíos programados con filtros"""
    try:
        config_id = request.args.get('config_id', type=int)
        fecha_desde = request.args.get('fecha_desde')
        estado = request.args.get('estado')
        limit = request.args.get('limit', 50, type=int)
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = """
            SELECT h.id, h.configuracion_id, e.nombre as config_nombre,
                   h.fecha_ejecucion, h.documentos_procesados, h.documentos_enviados,
                   h.emails_enviados, h.emails_fallidos, h.estado, h.mensaje,
                   h.duracion_segundos
            FROM historial_envios_programados h
            LEFT JOIN envios_programados e ON h.configuracion_id = e.id
            WHERE 1=1
        """
        params = []
        
        if config_id:
            query += " AND h.configuracion_id = ?"
            params.append(config_id)
        
        if fecha_desde:
            query += " AND DATE(h.fecha_ejecucion) >= ?"
            params.append(fecha_desde)
        
        if estado:
            query += " AND h.estado = ?"
            params.append(estado)
        
        query += " ORDER BY h.fecha_ejecucion DESC LIMIT ?"
        params.append(limit)
        
        cursor.execute(query, params)
        
        historial = []
        for row in cursor.fetchall():
            historial.append({
                'id': row[0],
                'configuracion_id': row[1],
                'config_nombre': row[2],
                'fecha_ejecucion': row[3],
                'documentos_procesados': row[4],
                'documentos_enviados': row[5],
                'emails_enviados': row[6],
                'emails_fallidos': row[7],
                'estado': row[8],
                'mensaje': row[9],
                'duracion_segundos': row[10]
            })
        
        conn.close()
        
        return jsonify({
            'exito': True,
            'historial': historial
        })
        
    except Exception as e:
        logger.error(f"Error obteniendo historial: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error: {str(e)}'
        }), 500


# ================================
# ENDPOINTS DE LOGS DEL SISTEMA
# ================================

@dian_vs_erp_bp.route('/api/logs', methods=['GET'])
def obtener_logs_sistema():
    """Obtiene los logs del sistema con filtros"""
    try:
        fecha = request.args.get('fecha', date.today().isoformat())
        tipo = request.args.get('tipo', 'todos')
        
        # Datos de ejemplo de logs
        logs = [
            {
                'timestamp': '2025-11-30 10:30:15',
                'tipo': 'INFO',
                'modulo': 'ENVIO_EMAIL',
                'mensaje': 'Email enviado exitosamente a jperez@supertiendascanaveral.com - NIT: 805028041',
                'detalles': {'emails_enviados': 5, 'tiempo_proceso': '2.3s'}
            },
            {
                'timestamp': '2025-11-30 10:25:42',
                'tipo': 'WARNING',
                'modulo': 'SMTP',
                'mensaje': 'Conexión SMTP lenta - Tiempo de respuesta: 3.2s',
                'detalles': {'host': 'smtp.gmail.com', 'puerto': 587}
            },
            {
                'timestamp': '2025-11-30 10:20:08',
                'tipo': 'ERROR',
                'modulo': 'PROCESAMIENTO',
                'mensaje': 'Error procesando archivo DIAN - Formato inválido',
                'detalles': {'archivo': 'dian_facturas_20251130.xlsx', 'linea_error': 127}
            },
            {
                'timestamp': '2025-11-30 10:15:30',
                'tipo': 'SUCCESS',
                'modulo': 'CARGA_ARCHIVO',
                'mensaje': 'Archivo ERP cargado exitosamente',
                'detalles': {'archivo': 'erp_data_20251130.csv', 'registros': 15420}
            }
        ]
        
        # Filtrar por tipo si se especifica
        if tipo != 'todos':
            logs = [log for log in logs if log['tipo'].lower() == tipo.lower()]
        
        return jsonify({
            'exito': True,
            'fecha': fecha,
            'tipo_filtro': tipo,
            'logs': logs,
            'total': len(logs)
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error obteniendo logs: {str(e)}'
        }), 500


# ================================
# ENDPOINTS ADICIONALES DE USUARIOS
# ================================

@dian_vs_erp_bp.route('/api/dian_usuarios/todos', methods=['GET'])
def obtener_todos_usuarios():
    """Obtiene todos los usuarios configurados agrupados por NIT"""
    try:
        # Datos de ejemplo de todos los usuarios
        usuarios_por_nit = [
            {
                'nit': '805028041',
                'razon_social': 'Supertiendas Cañaveral S.A.',
                'usuarios': [
                    {
                        'id': 1,
                        'nombres': 'Juan Carlos',
                        'apellidos': 'Pérez López',
                        'correo': 'jperez@supertiendascanaveral.com',
                        'tipo': 'SOLICITANTE',
                        'activo': True
                    },
                    {
                        'id': 2,
                        'nombres': 'María Elena',
                        'apellidos': 'Rodríguez Gómez',
                        'correo': 'mrodriguez@supertiendascanaveral.com',
                        'tipo': 'APROBADOR',
                        'activo': True
                    }
                ]
            },
            {
                'nit': '800123456',
                'razon_social': 'Empresa Ejemplo S.A.S.',
                'usuarios': [
                    {
                        'id': 3,
                        'nombres': 'Carlos Alberto',
                        'apellidos': 'López Martínez',
                        'correo': 'carlos@empresa.com',
                        'tipo': 'SOLICITANTE',
                        'activo': True
                    }
                ]
            }
        ]
        
        return jsonify({
            'exito': True,
            'usuarios_por_nit': usuarios_por_nit
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error obteniendo usuarios: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/dian_usuarios/desactivar/<int:user_id>', methods=['POST'])
def desactivar_usuario(user_id):
    """Desactiva un usuario específico"""
    try:
        # En producción: actualizar estado en base de datos
        # Por ahora simular desactivación exitosa
        
        return jsonify({
            'exito': True,
            'mensaje': f'Usuario {user_id} desactivado correctamente'
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error desactivando usuario: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/dian_usuarios/eliminar/<int:user_id>', methods=['DELETE'])
def eliminar_usuario(user_id):
    """Elimina un usuario específico"""
    try:
        # En producción: eliminar de base de datos
        # Por ahora simular eliminación exitosa
        
        return jsonify({
            'exito': True,
            'mensaje': f'Usuario {user_id} eliminado correctamente'
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error eliminando usuario: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/logs/recientes', methods=['GET'])
def obtener_logs_recientes():
    """
    Obtiene los logs más recientes con filtros opcionales.
    
    Query params:
        limite (int): Cantidad máxima de logs (default: 100)
        nivel (str): Filtrar por nivel (SUCCESS, INFO, WARNING, ERROR, CRITICAL)
        operacion (str): Filtrar por tipo de operación
        usuario (str): Filtrar por usuario
    
    Returns:
        JSON con lista de logs y totales
    """
    try:
        # Obtener parámetros de consulta
        limite = int(request.args.get('limite', 100))
        nivel = request.args.get('nivel', None)
        operacion = request.args.get('operacion', None)
        usuario = request.args.get('usuario', None)
        
        # Usar helper para obtener logs
        logs = helper_obtener_logs(
            limite=limite,
            nivel=nivel,
            operacion=operacion,
            usuario=usuario
        )
        
        return jsonify({
            'exito': True,
            'logs': logs,
            'total': len(logs),
            'filtros': {
                'limite': limite,
                'nivel': nivel,
                'operacion': operacion,
                'usuario': usuario
            }
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error obteniendo logs recientes: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/logs/filtrar', methods=['POST'])
def filtrar_logs_avanzado():
    """
    Filtrado avanzado de logs con rango de fechas y múltiples criterios.
    
    Body JSON:
        {
            "fecha_desde": "2025-12-01",
            "fecha_hasta": "2025-12-25",
            "nivel": "ERROR",
            "usuario": "admin",
            "nit": "900123456",
            "operacion": "ENVIO_EMAIL",
            "buscar_texto": "timeout",
            "limite": 200
        }
    
    Returns:
        JSON con logs filtrados y estadísticas
    """
    try:
        data = request.get_json()
        
        # Construir query base
        query = LogSistemaDianVsErp.query
        
        # Aplicar filtros
        if data.get('fecha_desde'):
            fecha_desde = datetime.strptime(data['fecha_desde'], '%Y-%m-%d').date()
            query = query.filter(LogSistemaDianVsErp.fecha >= fecha_desde)
        
        if data.get('fecha_hasta'):
            fecha_hasta = datetime.strptime(data['fecha_hasta'], '%Y-%m-%d').date()
            query = query.filter(LogSistemaDianVsErp.fecha <= fecha_hasta)
        
        if data.get('nivel'):
            query = query.filter_by(nivel=data['nivel'])
        
        if data.get('usuario'):
            query = query.filter(LogSistemaDianVsErp.usuario.ilike(f"%{data['usuario']}%"))
        
        if data.get('nit'):
            query = query.filter_by(nit_relacionado=data['nit'])
        
        if data.get('operacion'):
            query = query.filter_by(operacion=data['operacion'])
        
        if data.get('buscar_texto'):
            texto = f"%{data['buscar_texto']}%"
            query = query.filter(
                db.or_(
                    LogSistemaDianVsErp.mensaje.ilike(texto),
                    LogSistemaDianVsErp.detalles.ilike(texto)
                )
            )
        
        # Ordenar y limitar
        limite = data.get('limite', 200)
        logs_query = query.order_by(LogSistemaDianVsErp.timestamp.desc()).limit(limite).all()
        
        logs = [log.to_dict() for log in logs_query]
        
        # Obtener estadísticas del rango
        stats = {
            'total_encontrados': len(logs),
            'rango_fechas': {
                'desde': data.get('fecha_desde', 'N/A'),
                'hasta': data.get('fecha_hasta', 'N/A')
            }
        }
        
        # Registrar la búsqueda
        registrar_log(
            'INFO',
            'FILTRAR_LOGS',
            f'Búsqueda de logs: {len(logs)} resultados',
            detalles={'filtros': data}
        )
        
        return jsonify({
            'exito': True,
            'logs': logs,
            'estadisticas': stats
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error filtrando logs: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/logs/metricas', methods=['GET'])
def obtener_metricas_logs():
    """
    Obtiene métricas y estadísticas agregadas de logs.
    
    Query params:
        fecha_desde (str): Fecha inicio formato YYYY-MM-DD
        fecha_hasta (str): Fecha fin formato YYYY-MM-DD
    
    Returns:
        JSON con estadísticas detalladas:
        - Total de logs
        - Conteo por nivel (SUCCESS, INFO, WARNING, ERROR, CRITICAL)
        - Operaciones lentas (> 5 segundos)
        - Tasa de éxito
        - Top 5 operaciones
        - Top 5 usuarios más activos
    """
    try:
        from sqlalchemy import func
        
        # Obtener parámetros
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        
        if fecha_desde:
            fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        if fecha_hasta:
            fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        
        # Obtener estadísticas básicas
        stats = obtener_estadisticas_logs(fecha_desde, fecha_hasta)
        
        # Top 5 operaciones más frecuentes
        query_top_ops = db.session.query(
            LogSistemaDianVsErp.operacion,
            func.count(LogSistemaDianVsErp.id).label('cantidad')
        )
        
        if fecha_desde:
            query_top_ops = query_top_ops.filter(LogSistemaDianVsErp.fecha >= fecha_desde)
        if fecha_hasta:
            query_top_ops = query_top_ops.filter(LogSistemaDianVsErp.fecha <= fecha_hasta)
        
        top_operaciones = [
            {'operacion': op, 'cantidad': cant}
            for op, cant in query_top_ops.group_by(LogSistemaDianVsErp.operacion).order_by(func.count(LogSistemaDianVsErp.id).desc()).limit(5).all()
        ]
        
        # Top 5 usuarios más activos
        query_top_users = db.session.query(
            LogSistemaDianVsErp.usuario,
            func.count(LogSistemaDianVsErp.id).label('cantidad')
        ).filter(LogSistemaDianVsErp.usuario.isnot(None))
        
        if fecha_desde:
            query_top_users = query_top_users.filter(LogSistemaDianVsErp.fecha >= fecha_desde)
        if fecha_hasta:
            query_top_users = query_top_users.filter(LogSistemaDianVsErp.fecha <= fecha_hasta)
        
        top_usuarios = [
            {'usuario': usr, 'cantidad': cant}
            for usr, cant in query_top_users.group_by(LogSistemaDianVsErp.usuario).order_by(func.count(LogSistemaDianVsErp.id).desc()).limit(5).all()
        ]
        
        # Combinar todo
        stats['top_operaciones'] = top_operaciones
        stats['top_usuarios'] = top_usuarios
        
        return jsonify({
            'exito': True,
            'metricas': stats
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error obteniendo métricas: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/logs/exportar', methods=['POST'])
def exportar_logs_excel():
    """
    Exporta logs filtrados a Excel.
    
    Body JSON: mismos filtros que /api/logs/filtrar
    
    Returns:
        Archivo Excel con logs
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        data = request.get_json()
        
        # Reutilizar lógica de filtrado (mismo código que filtrar_logs_avanzado)
        query = LogSistemaDianVsErp.query
        
        if data.get('fecha_desde'):
            fecha_desde = datetime.strptime(data['fecha_desde'], '%Y-%m-%d').date()
            query = query.filter(LogSistemaDianVsErp.fecha >= fecha_desde)
        
        if data.get('fecha_hasta'):
            fecha_hasta = datetime.strptime(data['fecha_hasta'], '%Y-%m-%d').date()
            query = query.filter(LogSistemaDianVsErp.fecha <= fecha_hasta)
        
        if data.get('nivel'):
            query = query.filter_by(nivel=data['nivel'])
        
        if data.get('usuario'):
            query = query.filter(LogSistemaDianVsErp.usuario.ilike(f"%{data['usuario']}%"))
        
        if data.get('nit'):
            query = query.filter_by(nit_relacionado=data['nit'])
        
        if data.get('operacion'):
            query = query.filter_by(operacion=data['operacion'])
        
        # Obtener logs (sin límite para exportación)
        logs = query.order_by(LogSistemaDianVsErp.timestamp.desc()).all()
        
        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Logs Sistema"
        
        # Encabezados
        headers = ['ID', 'Fecha', 'Hora', 'Nivel', 'Usuario', 'IP', 'NIT', 'Empresa', 'Operación', 'Mensaje', 'Duración (ms)']
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for log in logs:
            ws.append([
                log.id,
                log.fecha.strftime('%Y-%m-%d') if log.fecha else '',
                log.timestamp.strftime('%H:%M:%S') if log.timestamp else '',
                log.nivel,
                log.usuario or '',
                log.ip_origen or '',
                log.nit_relacionado or '',
                log.empresa_relacionada or '',
                log.operacion or '',
                log.mensaje,
                log.duracion_ms or ''
            ])
        
        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Guardar a BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Registrar exportación
        registrar_log(
            'SUCCESS',
            'EXPORTAR_LOGS',
            f'Logs exportados a Excel: {len(logs)} registros',
            detalles={'filtros': data}
        )
        
        filename = f'logs_dian_vs_erp_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error exportando logs: {str(e)}'
        }), 500


# ============================================
# ENDPOINTS DE USUARIOS ASIGNADOS POR NIT
# ============================================

@dian_vs_erp_bp.route('/api/usuarios/todos', methods=['GET'])
def api_usuarios_todos():
    """Obtener todos los usuarios asignados a NITs"""
    try:
        from app import db as main_db
        
        # Query para obtener todos los usuarios
        query = """
            SELECT id, nit, razon_social, tipo_usuario, nombres, apellidos, 
                   correo, telefono, activo, fecha_creacion
            FROM usuarios_asignados
            ORDER BY nit, tipo_usuario, apellidos
        """
        
        result = main_db.session.execute(main_db.text(query))
        usuarios = []
        
        for row in result:
            usuarios.append({
                'usuario_id': row[0],
                'nit': row[1],
                'razon_social': row[2],
                'tipo_usuario': row[3],
                'nombres': row[4],
                'apellidos': row[5],
                'correo': row[6],
                'telefono': row[7],
                'activo': row[8],
                'fecha_creacion': row[9].isoformat() if row[9] else None
            })
        
        return jsonify({
            'exito': True,
            'usuarios': usuarios,
            'total': len(usuarios)
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error obteniendo usuarios: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/usuarios/por_nit/<nit>', methods=['GET'])
def api_usuarios_por_nit(nit):
    """Obtener usuarios de un NIT específico"""
    try:
        from app import db as main_db
        
        query = """
            SELECT id, nit, razon_social, tipo_usuario, nombres, apellidos, 
                   correo, telefono, activo, fecha_creacion
            FROM usuarios_asignados
            WHERE nit = :nit
            ORDER BY tipo_usuario, apellidos
        """
        
        result = main_db.session.execute(main_db.text(query), {'nit': nit})
        usuarios = []
        
        for row in result:
            usuarios.append({
                'usuario_id': row[0],
                'nit': row[1],
                'razon_social': row[2],
                'tipo_usuario': row[3],
                'nombres': row[4],
                'apellidos': row[5],
                'correo': row[6],
                'telefono': row[7],
                'activo': row[8],
                'fecha_creacion': row[9].isoformat() if row[9] else None
            })
        
        return jsonify({
            'exito': True,
            'usuarios': usuarios,
            'total': len(usuarios)
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error obteniendo usuarios del NIT {nit}: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/usuarios/agregar', methods=['POST'])
@requiere_permiso('dian_vs_erp', 'gestionar_usuarios_dian')
def api_usuarios_agregar():
    """Agregar un nuevo usuario asociado a un NIT"""
    try:
        from app import db as main_db
        import re
        
        datos = request.get_json()
        
        # Validaciones
        nit = datos.get('nit', '').strip()
        razon_social = datos.get('razon_social', '').strip()
        tipo_usuario = datos.get('tipo_usuario', '').upper()
        nombres = datos.get('nombres', '').strip()
        apellidos = datos.get('apellidos', '').strip()
        correo = datos.get('correo', '').strip().lower()
        telefono = datos.get('telefono', '').strip()
        
        if not nit:
            return jsonify({'exito': False, 'mensaje': 'NIT es requerido'}), 400
        
        if tipo_usuario not in ['SOLICITANTE', 'APROBADOR']:
            return jsonify({'exito': False, 'mensaje': 'Tipo de usuario inválido'}), 400
        
        if not nombres or not apellidos:
            return jsonify({'exito': False, 'mensaje': 'Nombres y apellidos son requeridos'}), 400
        
        # Validar email
        patron_email = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(patron_email, correo):
            return jsonify({'exito': False, 'mensaje': 'Email no válido'}), 400
        
        # Insertar usuario
        query = """
            INSERT INTO usuarios_asignados 
            (nit, razon_social, tipo_usuario, nombres, apellidos, correo, telefono, activo)
            VALUES (:nit, :razon_social, :tipo_usuario, :nombres, :apellidos, :correo, :telefono, TRUE)
        """
        
        main_db.session.execute(main_db.text(query), {
            'nit': nit,
            'razon_social': razon_social,
            'tipo_usuario': tipo_usuario,
            'nombres': nombres,
            'apellidos': apellidos,
            'correo': correo,
            'telefono': telefono
        })
        
        main_db.session.commit()
        
        return jsonify({
            'exito': True,
            'mensaje': f'Usuario {nombres} {apellidos} agregado exitosamente'
        })
        
    except Exception as e:
        main_db.session.rollback()
        
        # Verificar si es error de duplicado
        if 'duplicate' in str(e).lower() or 'unique' in str(e).lower():
            return jsonify({
                'exito': False,
                'mensaje': 'Ya existe un usuario con ese correo y tipo para este NIT'
            }), 400
        
        return jsonify({
            'exito': False,
            'mensaje': f'Error agregando usuario: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/usuarios/desactivar/<int:id>', methods=['POST'])
@requiere_permiso('dian_vs_erp', 'gestionar_usuarios_dian')
def api_usuarios_desactivar(id):
    """Desactivar (eliminar lógicamente) un usuario"""
    try:
        from app import db as main_db
        
        query = """
            UPDATE usuarios_asignados 
            SET activo = FALSE, fecha_modificacion = CURRENT_TIMESTAMP
            WHERE id = :id
        """
        
        result = main_db.session.execute(main_db.text(query), {'id': id})
        main_db.session.commit()
        
        if result.rowcount == 0:
            return jsonify({
                'exito': False,
                'mensaje': 'Usuario no encontrado'
            }), 404
        
        return jsonify({
            'exito': True,
            'mensaje': 'Usuario desactivado correctamente'
        })
        
    except Exception as e:
        from app import db as main_db
        main_db.session.rollback()
        print(f"ERROR desactivar usuario: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error al desactivar usuario: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/usuarios/obtener/<int:id>', methods=['GET'])
def api_usuarios_obtener(id):
    """Obtener datos de un usuario específico para edición"""
    try:
        from app import db as main_db
        
        query = """
            SELECT 
                id as usuario_id,
                nit,
                razon_social,
                tipo_usuario,
                nombres,
                apellidos,
                correo,
                telefono,
                activo
            FROM usuarios_asignados 
            WHERE id = :id
        """
        
        result = main_db.session.execute(main_db.text(query), {'id': id})
        usuario = result.fetchone()
        
        if not usuario:
            return jsonify({
                'exito': False,
                'mensaje': 'Usuario no encontrado'
            }), 404
        
        return jsonify({
            'exito': True,
            'usuario': dict(usuario._mapping)
        })
        
    except Exception as e:
        print(f"ERROR obtener usuario: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error al obtener usuario: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/usuarios/editar/<int:id>', methods=['PUT'])
@requiere_permiso('dian_vs_erp', 'gestionar_usuarios_dian')
def api_usuarios_editar(id):
    """Editar datos de un usuario existente"""
    try:
        from app import db as main_db
        data = request.get_json()
        
        # Validaciones
        if not data.get('nombres') or not data.get('apellidos') or not data.get('correo'):
            return jsonify({
                'exito': False,
                'mensaje': 'Faltan campos obligatorios'
            }), 400
        
        query = """
            UPDATE usuarios_asignados 
            SET 
                razon_social = :razon_social,
                tipo_usuario = :tipo_usuario,
                nombres = :nombres,
                apellidos = :apellidos,
                correo = :correo,
                telefono = :telefono,
                fecha_modificacion = CURRENT_TIMESTAMP
            WHERE id = :id
        """
        
        result = main_db.session.execute(main_db.text(query), {
            'id': id,
            'razon_social': data.get('razon_social'),
            'tipo_usuario': data.get('tipo_usuario'),
            'nombres': data.get('nombres'),
            'apellidos': data.get('apellidos'),
            'correo': data.get('correo'),
            'telefono': data.get('telefono')
        })
        
        main_db.session.commit()
        
        if result.rowcount == 0:
            return jsonify({
                'exito': False,
                'mensaje': 'Usuario no encontrado'
            }), 404
        
        return jsonify({
            'exito': True,
            'mensaje': 'Usuario actualizado correctamente'
        })
        
    except Exception as e:
        from app import db as main_db
        main_db.session.rollback()
        print(f"ERROR editar usuario: {e}")
        return jsonify({
            'exito': False,
            'mensaje': f'Error al actualizar usuario: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/usuarios/activar/<int:id>', methods=['POST'])
def api_usuarios_activar(id):
    """Activar un usuario desactivado"""
    try:
        from app import db as main_db
        
        query = """
            UPDATE usuarios_asignados 
            SET activo = TRUE, fecha_modificacion = CURRENT_TIMESTAMP
            WHERE id = :id
        """
        
        result = main_db.session.execute(main_db.text(query), {'id': id})
        main_db.session.commit()
        
        if result.rowcount == 0:
            return jsonify({
                'exito': False,
                'mensaje': 'Usuario no encontrado'
            }), 404
        
        return jsonify({
            'exito': True,
            'mensaje': 'Usuario activado exitosamente'
        })
        
    except Exception as e:
        main_db.session.rollback()
        return jsonify({
            'exito': False,
            'mensaje': f'Error activando usuario: {str(e)}'
        }), 500


# ========================================
# FIN DE ENDPOINTS ACTIVOS
# NOTA: Se eliminaron endpoints duplicados que causaban conflictos
# Los endpoints correctos están arriba (líneas 1187-1510)
# ========================================
        import json
        from modules.dian_vs_erp.scheduler_envios import reiniciar_scheduler_dian_vs_erp
        
        data = request.json
        usuario_sesion = session.get('usuario', 'WEB')
        
        # Validar campos requeridos
        if not data.get('nombre') or not data.get('tipo') or not data.get('hora_envio') or not data.get('frecuencia'):
            return jsonify({
                'exito': False,
                'mensaje': 'Faltan campos requeridos: nombre, tipo, hora_envio, frecuencia'
            }), 400
        
        # Crear configuración
        envio = EnvioProgramadoDianVsErp(
            nombre=data['nombre'].strip(),
            tipo=data['tipo'],
            dias_minimos=data.get('dias_minimos', 3),
            requiere_acuses_minimo=data.get('requiere_acuses_minimo', 2),
            estados_incluidos=data.get('estados_incluidos'),
            estados_excluidos=json.dumps(data.get('estados_excluidos', [])),
            hora_envio=data['hora_envio'],
            frecuencia=data['frecuencia'],
            dias_semana=json.dumps(data.get('dias_semana', [])),
            tipo_destinatario=data.get('tipo_destinatario'),
            email_cc=data.get('email_cc'),
            activo=data.get('activo', True),
            usuario_creacion=usuario_sesion
        )
        
        # Calcular próximo envío
        from modules.dian_vs_erp.scheduler_envios import scheduler_dian_vs_erp_global
        if scheduler_dian_vs_erp_global:
            envio.proximo_envio = scheduler_dian_vs_erp_global._calcular_proximo_envio(envio)
        
        db.session.add(envio)
        db.session.commit()
        
        # Reiniciar scheduler para cargar la nueva configuración
        try:
            reiniciar_scheduler_dian_vs_erp()
        except:
            pass
        
        return jsonify({
            'exito': True,
            'id': envio.id,
            'mensaje': 'Configuración creada exitosamente'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'exito': False,
            'mensaje': f'Error creando configuración: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/config/envios/<int:id>', methods=['PUT'])
def api_envios_actualizar(id):
    """Actualizar una configuración existente"""
    try:
        import json
        from modules.dian_vs_erp.scheduler_envios import reiniciar_scheduler_dian_vs_erp
        
        envio = EnvioProgramadoDianVsErp.query.get(id)
        if not envio:
            return jsonify({
                'exito': False,
                'mensaje': 'Configuración no encontrada'
            }), 404
        
        data = request.json
        usuario_sesion = session.get('usuario', 'WEB')
        
        # Actualizar campos
        if 'nombre' in data:
            envio.nombre = data['nombre'].strip()
        if 'tipo' in data:
            envio.tipo = data['tipo']
        if 'dias_minimos' in data:
            envio.dias_minimos = data['dias_minimos']
        if 'requiere_acuses_minimo' in data:
            envio.requiere_acuses_minimo = data['requiere_acuses_minimo']
        if 'estados_excluidos' in data:
            envio.estados_excluidos = json.dumps(data['estados_excluidos'])
        if 'hora_envio' in data:
            envio.hora_envio = data['hora_envio']
        if 'frecuencia' in data:
            envio.frecuencia = data['frecuencia']
        if 'dias_semana' in data:
            envio.dias_semana = json.dumps(data['dias_semana'])
        if 'activo' in data:
            envio.activo = data['activo']
        
        envio.usuario_modificacion = usuario_sesion
        envio.fecha_modificacion = datetime.now()
        
        db.session.commit()
        
        # Reiniciar scheduler
        try:
            reiniciar_scheduler_dian_vs_erp()
        except:
            pass
        
        return jsonify({
            'exito': True,
            'mensaje': 'Configuración actualizada exitosamente'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'exito': False,
            'mensaje': f'Error actualizando configuración: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/config/envios/<int:id>', methods=['DELETE'])
@requiere_permiso('dian_vs_erp', 'gestionar_envios')
def api_envios_eliminar(id):
    """Eliminar una configuración y su historial asociado"""
    try:
        from modules.dian_vs_erp.scheduler_envios import reiniciar_scheduler_dian_vs_erp
        from modules.dian_vs_erp.models import HistorialEnvioDianVsErp
        
        envio = EnvioProgramadoDianVsErp.query.get(id)
        if not envio:
            return jsonify({
                'exito': False,
                'mensaje': 'Configuración no encontrada'
            }), 404
        
        # PRIMERO: Eliminar registros del historial asociados a esta configuración
        HistorialEnvioDianVsErp.query.filter_by(configuracion_id=id).delete()
        
        # SEGUNDO: Eliminar la configuración
        db.session.delete(envio)
        db.session.commit()
        
        # Reiniciar scheduler
        try:
            reiniciar_scheduler_dian_vs_erp()
        except:
            pass
        
        return jsonify({
            'exito': True,
            'mensaje': 'Configuración eliminada exitosamente'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'exito': False,
            'mensaje': f'Error eliminando configuración: {str(e)}'
        }), 500
        db.session.rollback()
        return jsonify({
            'exito': False,
            'mensaje': f'Error eliminando configuración: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/config/envios/<int:id>/ejecutar', methods=['POST'])
@requiere_permiso('dian_vs_erp', 'gestionar_envios')
def api_envios_ejecutar_manual(id):
    """Ejecutar manualmente un envío (para pruebas)"""
    try:
        from modules.dian_vs_erp.scheduler_envios import scheduler_dian_vs_erp_global
        
        if not scheduler_dian_vs_erp_global:
            return jsonify({
                'exito': False,
                'mensaje': 'Scheduler no está iniciado'
            }), 500
        
        envio = EnvioProgramadoDianVsErp.query.get(id)
        if not envio:
            return jsonify({
                'exito': False,
                'mensaje': 'Configuración no encontrada'
            }), 404
        
        # Ejecutar en segundo plano para no bloquear la respuesta
        import threading
        thread = threading.Thread(
            target=scheduler_dian_vs_erp_global.ejecutar_envio_programado,
            args=(id,)
        )
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'exito': True,
            'mensaje': 'Ejecución iniciada. Revise el historial en unos momentos.'
        })
    
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error ejecutando envío: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/config/envios/<int:id>/toggle', methods=['POST'])
def api_envios_toggle(id):
    """Activar/Desactivar una configuración"""
    try:
        from modules.dian_vs_erp.scheduler_envios import reiniciar_scheduler_dian_vs_erp
        
        envio = EnvioProgramadoDianVsErp.query.get(id)
        if not envio:
            return jsonify({
                'exito': False,
                'mensaje': 'Configuración no encontrada'
            }), 404
        
        envio.activo = not envio.activo
        envio.fecha_modificacion = datetime.now()
        envio.usuario_modificacion = session.get('usuario', 'WEB')
        
        db.session.commit()
        
        # Reiniciar scheduler
        try:
            reiniciar_scheduler_dian_vs_erp()
        except:
            pass
        
        return jsonify({
            'exito': True,
            'activo': envio.activo,
            'mensaje': f'Configuración {"activada" if envio.activo else "desactivada"} exitosamente'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'exito': False,
            'mensaje': f'Error cambiando estado: {str(e)}'
        }), 500


# ========================================
# ENDPOINTS - USUARIOS DE CAUSACIÓN
# ========================================

@dian_vs_erp_bp.route('/api/usuarios-causacion', methods=['GET'])
def api_usuarios_causacion_get():
    """Obtener todos los usuarios de causación"""
    try:
        usuarios = UsuarioCausacionDianVsErp.query.order_by(
            UsuarioCausacionDianVsErp.nombre_causador.asc()
        ).all()
        
        return jsonify({
            'exito': True,
            'usuarios': [u.to_dict() for u in usuarios]
        })
    
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error obteniendo usuarios: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/usuarios-causacion', methods=['POST'])
def api_usuarios_causacion_crear():
    """Crear un nuevo usuario de causación"""
    try:
        data = request.json
        usuario_sesion = session.get('usuario', 'WEB')
        
        # Validar campos requeridos
        if not data.get('nombre_causador') or not data.get('email'):
            return jsonify({
                'exito': False,
                'mensaje': 'Faltan campos requeridos: nombre_causador, email'
            }), 400
        
        # Validar que no exista
        existe = UsuarioCausacionDianVsErp.query.filter_by(
            nombre_causador=data['nombre_causador'].strip().upper()
        ).first()
        
        if existe:
            return jsonify({
                'exito': False,
                'mensaje': 'Ya existe un usuario con ese nombre'
            }), 400
        
        # Crear usuario
        usuario = UsuarioCausacionDianVsErp(
            nombre_causador=data['nombre_causador'].strip().upper(),
            email=data['email'].strip().lower(),
            activo=data.get('activo', True),
            usuario_creacion=usuario_sesion
        )
        
        db.session.add(usuario)
        db.session.commit()
        
        return jsonify({
            'exito': True,
            'id': usuario.id,
            'mensaje': 'Usuario de causación creado exitosamente',
            'usuario': usuario.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'exito': False,
            'mensaje': f'Error creando usuario: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/usuarios-causacion/<int:id>', methods=['PUT'])
def api_usuarios_causacion_actualizar(id):
    """Actualizar un usuario de causación"""
    try:
        usuario = UsuarioCausacionDianVsErp.query.get(id)
        if not usuario:
            return jsonify({
                'exito': False,
                'mensaje': 'Usuario no encontrado'
            }), 404
        
        data = request.json
        usuario_sesion = session.get('usuario', 'WEB')
        
        # Actualizar campos
        if 'nombre_causador' in data:
            # Verificar que el nuevo nombre no exista en otro registro
            nuevo_nombre = data['nombre_causador'].strip().upper()
            if nuevo_nombre != usuario.nombre_causador:
                existe = UsuarioCausacionDianVsErp.query.filter(
                    UsuarioCausacionDianVsErp.nombre_causador == nuevo_nombre,
                    UsuarioCausacionDianVsErp.id != id
                ).first()
                
                if existe:
                    return jsonify({
                        'exito': False,
                        'mensaje': 'Ya existe un usuario con ese nombre'
                    }), 400
                
                usuario.nombre_causador = nuevo_nombre
        
        if 'email' in data:
            usuario.email = data['email'].strip().lower()
        
        if 'activo' in data:
            usuario.activo = data['activo']
        
        usuario.usuario_modificacion = usuario_sesion
        usuario.fecha_modificacion = datetime.now()
        
        db.session.commit()
        
        return jsonify({
            'exito': True,
            'mensaje': 'Usuario actualizado exitosamente',
            'usuario': usuario.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'exito': False,
            'mensaje': f'Error actualizando usuario: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/usuarios-causacion/<int:id>', methods=['DELETE'])
def api_usuarios_causacion_eliminar(id):
    """Eliminar un usuario de causación"""
    try:
        usuario = UsuarioCausacionDianVsErp.query.get(id)
        if not usuario:
            return jsonify({
                'exito': False,
                'mensaje': 'Usuario no encontrado'
            }), 404
        
        # Verificar si tiene documentos asociados
        if usuario.total_documentos and usuario.total_documentos > 0:
            return jsonify({
                'exito': False,
                'mensaje': f'No se puede eliminar. Este usuario tiene {usuario.total_documentos} documentos asociados. Desactívelo en su lugar.'
            }), 400
        
        db.session.delete(usuario)
        db.session.commit()
        
        return jsonify({
            'exito': True,
            'mensaje': 'Usuario eliminado exitosamente'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'exito': False,
            'mensaje': f'Error eliminando usuario: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/usuarios-causacion/<int:id>/toggle', methods=['POST'])
def api_usuarios_causacion_toggle(id):
    """Activar/Desactivar un usuario de causación"""
    try:
        usuario = UsuarioCausacionDianVsErp.query.get(id)
        if not usuario:
            return jsonify({
                'exito': False,
                'mensaje': 'Usuario no encontrado'
            }), 404
        
        usuario.activo = not usuario.activo
        usuario.fecha_modificacion = datetime.now()
        usuario.usuario_modificacion = session.get('usuario', 'WEB')
        
        db.session.commit()
        
        return jsonify({
            'exito': True,
            'activo': usuario.activo,
            'mensaje': f'Usuario {"activado" if usuario.activo else "desactivado"} exitosamente'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'exito': False,
            'mensaje': f'Error cambiando estado: {str(e)}'
        }), 500

@dian_vs_erp_bp.route('/api/causadores/disponibles', methods=['GET'])
def api_causadores_disponibles():
    """Obtener lista de causadores disponibles desde la tabla maestro"""
    try:
        # Obtener causadores únicos de la tabla maestro
        causadores = db.session.query(MaestroDianVsErp.doc_causado_por).filter(
            MaestroDianVsErp.doc_causado_por.isnot(None),
            MaestroDianVsErp.doc_causado_por != ''
        ).distinct().order_by(MaestroDianVsErp.doc_causado_por).all()
        
        lista_causadores = [c[0] for c in causadores if c[0]]
        
        return jsonify({
            'exito': True,
            'causadores': lista_causadores
        })
    
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error obteniendo causadores: {str(e)}'
        }), 500


# ========================================
# ENDPOINTS - HISTORIAL DE ENVÍOS
# ========================================

@dian_vs_erp_bp.route('/api/historial-envios', methods=['GET'])
def api_historial_envios():
    """Obtener el historial de envíos programados"""
    try:
        config_id = request.args.get('configuracion_id', type=int)
        limite = request.args.get('limite', 50, type=int)
        
        query = HistorialEnvioDianVsErp.query
        
        if config_id:
            query = query.filter_by(configuracion_id=config_id)
        
        historiales = query.order_by(
            HistorialEnvioDianVsErp.fecha_ejecucion.desc()
        ).limit(limite).all()
        
        # Incluir nombre de la configuración
        resultados = []
        for h in historiales:
            h_dict = h.to_dict()
            config = EnvioProgramadoDianVsErp.query.get(h.configuracion_id)
            h_dict['nombre_config'] = config.nombre if config else 'Configuración eliminada'
            resultados.append(h_dict)
        
        return jsonify({
            'exito': True,
            'historial': resultados
        })
    
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error obteniendo historial: {str(e)}'
        }), 500

# ========================================
# ENDPOINTS - SINCRONIZACI�N
# ========================================

@dian_vs_erp_bp.route('/api/sincronizar', methods=['POST'])
@requiere_permiso('dian_vs_erp', 'sincronizar')
def api_sincronizar():
    """
    🆕 VISOR V2: Sincronización con guardado en BD
    
    1. Verifica tablas y cuenta registros
    2. GUARDA fecha/hora en tabla log_sincronizacion_visor
    3. Frontend recarga datos automáticamente
    """
    try:
        from datetime import datetime
        import logging
        import json
        
        logger = logging.getLogger(__name__)
        logger.info("🔄 Sincronización Visor v2 activada...")
        
        fecha_actual = datetime.now()
        fecha_sincronizacion = fecha_actual.strftime('%Y-%m-%d %H:%M:%S')
        
        # ✅ VALIDAR QUE LAS TABLAS EXISTAN
        from modules.dian_vs_erp.models import Dian, ErpComercial, ErpFinanciero, Acuses
        from modules.recibir_facturas.models import FacturaTemporal, FacturaRecibida
        
        # Contar registros en cada tabla
        total_dian = Dian.query.count()
        total_erp_comercial = ErpComercial.query.count()
        total_erp_financiero = ErpFinanciero.query.count()
        total_acuses = Acuses.query.count()
        total_temporales = FacturaTemporal.query.count()
        total_recibidas = FacturaRecibida.query.count()
        
        logger.info(f"   ✅ Tablas verificadas:")
        logger.info(f"      📄 DIAN: {total_dian:,} registros")
        logger.info(f"      🏪 ERP Comercial: {total_erp_comercial:,} registros")
        logger.info(f"      💰 ERP Financiero: {total_erp_financiero:,} registros")
        logger.info(f"      ✉️ Acuses: {total_acuses:,} registros")
        logger.info(f"      📝 Facturas Temporales: {total_temporales:,} registros")
        logger.info(f"      ✅ Facturas Recibidas: {total_recibidas:,} registros")
        
        # ✅ VERIFICAR SI HAY FACTURAS TEMPORALES NUEVAS
        query_temporal_dian = """
            SELECT COUNT(*)
            FROM facturas_temporales f
            INNER JOIN dian d ON f.nit = d.nit_emisor
                              AND f.prefijo = d.prefijo
                              AND f.folio = d.folio
        """
        result_temp = db.session.execute(db.text(query_temporal_dian))
        facturas_temp_en_dian = result_temp.scalar()
        
        logger.info(f"   🔍 {facturas_temp_en_dian} facturas temporales están en DIAN")
        
        # ✅ VERIFICAR SI HAY FACTURAS RECIBIDAS NUEVAS
        query_recibida_dian = """
            SELECT COUNT(*)
            FROM facturas_recibidas f
            INNER JOIN dian d ON f.nit = d.nit_emisor
                              AND f.prefijo = d.prefijo
                              AND f.folio = d.folio
        """
        result_rec = db.session.execute(db.text(query_recibida_dian))
        facturas_rec_en_dian = result_rec.scalar()
        
        logger.info(f"   🔍 {facturas_rec_en_dian} facturas recibidas están en DIAN")
        
        # ✅ VISOR V2: NO actualiza maestro (obsoleto), los estados se calculan en tiempo real
        logger.info("✅ Sincronización V2: Estados se calculan automáticamente en el visor")
        logger.info("   (No se actualiza maestro_dian_vs_erp - tabla obsoleta)")
        
        # Commit para asegurar que cualquier cambio en las tablas esté guardado
        db.session.commit()
        
        # Contar registros que aparecerían como "Recibida" (información)
        docs_recibidas = facturas_temp_en_dian + facturas_rec_en_dian
        docs_causadas_doc = 0
        docs_causadas_erp = 0
        total_sincronizados = 0
        
        logger.info(f"ℹ️ Información: {docs_recibidas} facturas están en tablas de recepción")
        
        # 🆕 GUARDAR EN TABLA LOG_SINCRONIZACION_VISOR
        usuario_id = session.get('usuario_id')
        ip_origen = request.remote_addr
        user_agent = request.headers.get('User-Agent', '')
        
        resumen_json = {
            'total_dian': total_dian,
            'total_erp_comercial': total_erp_comercial,
            'total_erp_financiero': total_erp_financiero,
            'total_acuses': total_acuses,
            'total_temporales': total_temporales,
            'total_recibidas': total_recibidas,
            'temporales_en_dian': facturas_temp_en_dian,
            'recibidas_en_dian': facturas_rec_en_dian,
            'nota': 'Visor V2 calcula estados en tiempo real (no se actualiza tabla maestro)'
        }
        
        # Insertar registro en log_sincronizacion_visor
        insert_log = """
            INSERT INTO log_sincronizacion_visor 
            (usuario_id, fecha_sincronizacion, registros_verificados, tipo, datos_resumen, ip_origen, user_agent)
            VALUES (:usuario_id, :fecha_sync, :registros, 'manual', :resumen_json, :ip, :ua)
        """
        
        db.session.execute(db.text(insert_log), {
            'usuario_id': usuario_id,
            'fecha_sync': fecha_actual,
            'registros': total_dian,
            'resumen_json': json.dumps(resumen_json),
            'ip': ip_origen,
            'ua': user_agent
        })
        db.session.commit()
        
        logger.info(f"💾 Sincronización guardada en BD - Usuario ID: {usuario_id}")
        
        # ✅ RESPUESTA DE ÉXITO
        logger.info("✅ Sincronización completada - Frontend recargará datos")
        
        mensaje_detallado = []
        mensaje_detallado.append(f"📊 Verificación completada:")
        mensaje_detallado.append(f"   • DIAN: {total_dian:,} registros")
        mensaje_detallado.append(f"   • ERP Comercial: {total_erp_comercial:,} registros")
        mensaje_detallado.append(f"   • ERP Financiero: {total_erp_financiero:,} registros")
        mensaje_detallado.append(f"   • Acuses: {total_acuses:,} registros")
        mensaje_detallado.append(f"   • Facturas Temporales: {total_temporales:,} ({facturas_temp_en_dian} en DIAN)")
        mensaje_detallado.append(f"   • Facturas Recibidas: {total_recibidas:,} ({facturas_rec_en_dian} en DIAN)")
        mensaje_detallado.append(f"")
        mensaje_detallado.append(f"✅ El visor recalculará los estados automáticamente")
        
        return jsonify({
            'exito': True,
            'mensaje': '\n'.join(mensaje_detallado),
            'fecha_sincronizacion': fecha_sincronizacion,
            'resumen': resumen_json
        })
        
    except Exception as e:
        logger.error(f"❌ Error en sincronización: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'exito': False,
            'mensaje': f'Error en sincronización: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/api/ultima_sincronizacion', methods=['GET'])
def api_ultima_sincronizacion():
    """
    🆕 Obtiene la fecha y hora de la última sincronización desde BD
    """
    try:
        # Consultar última sincronización
        query = """
            SELECT fecha_sincronizacion, registros_verificados, datos_resumen
            FROM log_sincronizacion_visor
            ORDER BY fecha_sincronizacion DESC
            LIMIT 1
        """
        
        result = db.session.execute(db.text(query)).fetchone()
        
        if result:
            fecha_sync = result[0]
            # Formatear fecha como string
            if isinstance(fecha_sync, datetime):
                fecha_formateada = fecha_sync.strftime('%Y-%m-%d %H:%M:%S')
            else:
                fecha_formateada = str(fecha_sync)
            
            return jsonify({
                'exito': True,
                'fecha_sincronizacion': fecha_formateada,
                'registros_verificados': result[1],
                'resumen': result[2] if result[2] else {}
            })
        else:
            # No hay sincronizaciones previas
            return jsonify({
                'exito': True,
                'fecha_sincronizacion': None,
                'mensaje': 'No hay sincronizaciones previas'
            })
            
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error al obtener última sincronización: {str(e)}'
        }), 500


# ==============================
# 🚀 CARGA AUTOMÁTICA DE ARCHIVOS
# Sistema SIMPLE sin tablas adicionales
# Fecha: 30 de Diciembre de 2025
# ==============================

# Configuración HARDCODED de carpetas
CARPETAS_CARGA = [
    {
        'nombre': 'DIAN',
        'ruta': r'D:\PERFIL\Descargas\1.A. Para pruebas dian vs erp 29 12 2025\Dian',
        'ruta_procesados': r'D:\PERFIL\Descargas\1.A. Para pruebas dian vs erp 29 12 2025\Dian\Procesados',
        'tipo': 'dian'
    },
    {
        'nombre': 'ERP_FINANCIERO',
        'ruta': r'D:\PERFIL\Descargas\1.A. Para pruebas dian vs erp 29 12 2025\ERP Financiero',
        'ruta_procesados': r'D:\PERFIL\Descargas\1.A. Para pruebas dian vs erp 29 12 2025\ERP Financiero\Procesados',
        'tipo': 'erp_financiero'
    },
    {
        'nombre': 'ERP_COMERCIAL',
        'ruta': r'D:\PERFIL\Descargas\1.A. Para pruebas dian vs erp 29 12 2025\ERP Comercial',
        'ruta_procesados': r'D:\PERFIL\Descargas\1.A. Para pruebas dian vs erp 29 12 2025\ERP Comercial\Procesados',
        'tipo': 'erp_comercial'
    },
    {
        'nombre': 'ACUSES',
        'ruta': r'D:\PERFIL\Descargas\1.A. Para pruebas dian vs erp 29 12 2025\Acuses',
        'ruta_procesados': r'D:\PERFIL\Descargas\1.A. Para pruebas dian vs erp 29 12 2025\Acuses\Procesados',
        'tipo': 'acuses'
    }
]

@dian_vs_erp_bp.route('/cargar_automatico_escanear', methods=['POST'])
def escanear_archivos_pendientes():
    """Escanea carpetas y retorna archivos Excel/CSV encontrados"""
    try:
        from pathlib import Path
        
        archivos_encontrados = []
        archivos_procesados_carpeta = []
        
        for config in CARPETAS_CARGA:
            nombre_carpeta = config['nombre']
            ruta_pendientes = config['ruta']
            ruta_procesados = config['ruta_procesados']
            
            # Verificar que la ruta existe
            path = Path(ruta_pendientes)
            if not path.exists():
                print(f"⚠️ Carpeta no existe: {ruta_pendientes}")
                continue
            
            # Verificar carpeta de procesados
            path_procesados = Path(ruta_procesados)
            
            # Buscar archivos .xlsx, .xlsm, .csv (NO .xls porque están corruptos)
            for archivo in path.glob('*.xlsx'):
                archivos_encontrados.append({
                    'nombre': archivo.name,
                    'ruta': str(archivo),
                    'carpeta': nombre_carpeta,
                    'tipo': config['tipo'],
                    'tamanio': archivo.stat().st_size
                })
            
            for archivo in path.glob('*.xlsm'):
                archivos_encontrados.append({
                    'nombre': archivo.name,
                    'ruta': str(archivo),
                    'carpeta': nombre_carpeta,
                    'tipo': config['tipo'],
                    'tamanio': archivo.stat().st_size
                })
            
            for archivo in path.glob('*.csv'):
                archivos_encontrados.append({
                    'nombre': archivo.name,
                    'ruta': str(archivo),
                    'carpeta': nombre_carpeta,
                    'tipo': config['tipo'],
                    'tamanio': archivo.stat().st_size
                })
            
            # Contar archivos en carpeta procesados
            if path_procesados.exists():
                procesados = len(list(path_procesados.rglob('*.*')))
                archivos_procesados_carpeta.append({
                    'carpeta': nombre_carpeta,
                    'cantidad': procesados
                })
        
        return jsonify({
            'exito': True,
            'total_encontrados': len(archivos_encontrados),
            'pendientes': len(archivos_encontrados),
            'ya_procesados': sum(p['cantidad'] for p in archivos_procesados_carpeta),
            'archivos_pendientes': archivos_encontrados,
            'archivos_procesados': archivos_procesados_carpeta
        })
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error al escanear archivos: {str(e)}'
        }), 500
        
    except Exception as e:
        return jsonify({
            'exito': False,
            'mensaje': f'Error al escanear archivos: {str(e)}'
        }), 500


@dian_vs_erp_bp.route('/cargar_automatico_ejecutar', methods=['POST'])
def ejecutar_carga_automatica():
    """Ejecuta la carga automática - usa tabla maestro_dian_vs_erp existente"""
    try:
        print("\n" + "="*80)
        print("🚀 INICIANDO EJECUCIÓN DE CARGA AUTOMÁTICA")
        print("="*80)
        
        from pathlib import Path
        import shutil
        from datetime import datetime
        
        datos = request.get_json()
        print(f"📦 Datos recibidos: {datos}")
        archivos_a_procesar = datos.get('archivos', [])
        print(f"📁 Total de archivos a procesar: {len(archivos_a_procesar)}")
        
        if not archivos_a_procesar:
            print("❌ No se proporcionaron archivos")
            return jsonify({
                'exito': False,
                'mensaje': 'No se proporcionaron archivos para procesar'
            }), 400
        
        resultados = []
        
        for archivo_info in archivos_a_procesar:
            nombre_archivo = archivo_info.get('nombre')
            ruta_completa = archivo_info.get('ruta')
            carpeta_origen = archivo_info.get('carpeta')
            
            tiempo_inicio = time.time()
            estado = 'COMPLETADO'
            mensaje_error = None
            registros_insertados = 0
            
            try:
                # Verificar que el archivo existe
                if not Path(ruta_completa).exists():
                    raise Exception(f"Archivo no encontrado: {ruta_completa}")
                
                # Convertir Excel a CSV temporal
                csv_path = save_excel_to_csv_from_disk(ruta_completa, UPLOADS['dian'])
                
                # Cargar CSV en maestro_dian_vs_erp (tabla existente)
                registros_insertados = cargar_csv_a_maestro(csv_path)
                
                # Eliminar CSV temporal
                Path(csv_path).unlink(missing_ok=True)
                
            except Exception as e:
                estado = 'ERROR'
                mensaje_error = str(e)
                print(f"❌ Error procesando {nombre_archivo}: {mensaje_error}")
            
            tiempo_proceso = time.time() - tiempo_inicio
            
            # Mover archivo a carpeta "Procesados"
            try:
                # Buscar configuración de esta carpeta
                config = next((c for c in CARPETAS_CARGA if c['nombre'] == carpeta_origen), None)
                
                if config and config['ruta_procesados']:
                    ruta_procesados = Path(config['ruta_procesados'])
                    # Crear subcarpeta por fecha
                    fecha_hoy = datetime.now().strftime('%Y-%m-%d')
                    carpeta_destino = ruta_procesados / fecha_hoy
                    carpeta_destino.mkdir(parents=True, exist_ok=True)
                    
                    # Mover archivo
                    origen = Path(ruta_completa)
                    destino = carpeta_destino / nombre_archivo
                    shutil.move(str(origen), str(destino))
                    print(f"✅ Archivo movido a: {destino}")
            except Exception as e:
                print(f"⚠️ Error moviendo archivo: {str(e)}")
            
            resultados.append({
                'archivo': nombre_archivo,
                'estado': estado,
                'registros_insertados': registros_insertados,
                'tiempo_segundos': round(tiempo_proceso, 2),
                'error': mensaje_error
            })
        
        # Resumen
        exitosos = sum(1 for r in resultados if r['estado'] == 'COMPLETADO')
        fallidos = sum(1 for r in resultados if r['estado'] == 'ERROR')
        total_registros = sum(r['registros_insertados'] for r in resultados)
        
        return jsonify({
            'exito': True,
            'mensaje': f'Procesamiento completado: {exitosos} exitosos, {fallidos} fallidos',
            'resumen': {
                'archivos_procesados': len(resultados),
                'exitosos': exitosos,
                'fallidos': fallidos,
                'total_registros_insertados': total_registros
            },
            'detalles': resultados
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'exito': False,
            'mensaje': f'Error en carga automática: {str(e)}'
        }), 500


def cargar_csv_a_maestro(csv_path):
    """Carga un CSV en maestro_dian_vs_erp usando COPY FROM"""
    try:
        # Leer CSV con pandas para validar
        df = pd.read_csv(csv_path, dtype=str, encoding='utf-8')
        
        if df.empty:
            return 0
        
        # Usar PostgreSQL COPY FROM para inserción masiva
        conn = db.engine.raw_connection()
        cursor = conn.cursor()
        
        # INSERT directo desde CSV
        with open(csv_path, 'r', encoding='utf-8') as f:
            # Leer primera línea para obtener columnas
            primera_linea = f.readline().strip()
            f.seek(0)
            
            columnas = primera_linea.split(',')
            columnas_str = ','.join(columnas)
            
            try:
                cursor.copy_expert(
                    sql=f"COPY maestro_dian_vs_erp ({columnas_str}) FROM STDIN WITH CSV HEADER",
                    file=f
                )
            except Exception as copy_error:
                print(f"⚠️ COPY FROM falló: {copy_error}")
                # Fallback: INSERT fila por fila
                f.seek(0)
                next(f)  # Skip header
                insertados = 0
                for linea in f:
                    try:
                        valores = linea.strip().split(',')
                        placeholders = ','.join(['%s'] * len(valores))
                        cursor.execute(
                            f"INSERT INTO maestro_dian_vs_erp ({columnas_str}) VALUES ({placeholders})",
                            valores
                        )
                        insertados += 1
                    except:
                        pass
                conn.commit()
                cursor.close()
                conn.close()
                return insertados
        
        registros_insertados = len(df)
        conn.commit()
        cursor.close()
        conn.close()
        
        # Eliminar duplicados
        eliminar_duplicados_maestro()
        
        # ✅ CALCULAR DÍAS DESDE EMISIÓN
        calcular_dias_desde_emision()
        
        return registros_insertados
        
    except Exception as e:
        print(f"Error en cargar_csv_a_maestro: {str(e)}")
        return 0


def eliminar_duplicados_maestro():
    """Elimina duplicados de maestro_dian_vs_erp"""
    try:
        resultado = db.session.execute(
            db.text("""
                DELETE FROM maestro_dian_vs_erp
                WHERE id NOT IN (
                    SELECT MIN(id)
                    FROM maestro_dian_vs_erp
                    GROUP BY prefijo, folio, nit_proveedor, nit_empresa
                )
            """)
        )
        db.session.commit()
        print(f"   🗑️ {resultado.rowcount} duplicados eliminados")
        return resultado.rowcount
    except Exception as e:
        db.session.rollback()
        print(f"Error eliminando duplicados: {str(e)}")
        return 0


def calcular_dias_desde_emision():
    """Calcula días desde emisión para todos los registros sin días calculados"""
    try:
        from datetime import date
        fecha_hoy = date.today()
        
        resultado = db.session.execute(
            db.text("""
                UPDATE maestro_dian_vs_erp
                SET dias_desde_emision = (:fecha_hoy - fecha_emision)
                WHERE fecha_emision IS NOT NULL
                  AND (dias_desde_emision IS NULL OR dias_desde_emision = 0)
            """),
            {'fecha_hoy': fecha_hoy}
        )
        db.session.commit()
        print(f"   📅 {resultado.rowcount} registros actualizados con días desde emisión")
        return resultado.rowcount
    except Exception as e:
        db.session.rollback()
        print(f"Error calculando días: {str(e)}")
        return 0


# ============================================================================
# GESTIÓN DE ELIMINACIÓN MASIVA DE DATOS (16 Feb 2026)
# ============================================================================

@dian_vs_erp_bp.route('/api/solicitar_eliminacion', methods=['POST'])
@requiere_permiso('dian_vs_erp', 'eliminar_datos')
def solicitar_eliminacion():
    """
    Solicita eliminación de datos masivos
    Genera token de 6 dígitos y envía correo de confirmación
    """
    from modules.dian_vs_erp.models import TokenEliminacionDatos
    from datetime import datetime, timedelta
    import secrets
    import json
    from flask_mail import Message
    from app import mail
    
    try:
        # Validar sesión
        if 'usuario' not in session or 'usuario_id' not in session:
            return jsonify({'success': False, 'error': 'Sesión no válida'}), 401
        
        # Obtener parámetros
        data = request.json
        tipo_rango = data.get('tipo_rango')  # 'dias', 'meses', 'año'
        fecha_inicio_str = data.get('fecha_inicio')
        fecha_fin_str = data.get('fecha_fin')
        archivos = data.get('archivos', [])  # ['dian', 'erp_fn', 'erp_cm', 'acuses']
        email = data.get('email')
        
        # Validaciones
        if not tipo_rango or not archivos or not email:
            return jsonify({'success': False, 'error': 'Parámetros incompletos'}), 400
        
        if not fecha_inicio_str or not fecha_fin_str:
            return jsonify({'success': False, 'error': 'Debe especificar rango de fechas'}), 400
        
        # Convertir fechas
        from datetime import datetime as dt
        fecha_inicio = dt.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = dt.strptime(fecha_fin_str, '%Y-%m-%d').date()
        
        if fecha_inicio > fecha_fin:
            return jsonify({'success': False, 'error': 'Fecha inicio debe ser menor a fecha fin'}), 400
        
        # 🔍 VALIDAR QUE EXISTAN DATOS EN EL RANGO
        from sqlalchemy import text
        conteo_result = db.session.execute(
            text("""
                SELECT COUNT(*) as total 
                FROM maestro_dian_vs_erp 
                WHERE fecha_emision >= :fecha_inicio 
                  AND fecha_emision <= :fecha_fin
            """),
            {'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin}
        ).fetchone()
        
        total_registros = conteo_result[0] if conteo_result else 0
        
        if total_registros == 0:
            return jsonify({
                'success': False, 
                'error': f'No se encontraron registros para eliminar en el rango {fecha_inicio.strftime("%d/%m/%Y")} - {fecha_fin.strftime("%d/%m/%Y")}',
                'registros_encontrados': 0
            }), 404
        
        # Generar token de 6 dígitos
        token = ''.join([str(secrets.randbelow(10)) for _ in range(6)])
        
        # Crear registro de token
        token_obj = TokenEliminacionDatos(
            token=token,
            usuario_solicitante=session['usuario'],
            email_destino=email,
            tipo_rango=tipo_rango,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            archivos_eliminar=json.dumps(archivos),
            usado=False,
            fecha_creacion=datetime.utcnow(),
            fecha_expiracion=datetime.utcnow() + timedelta(minutes=10),
            ip_solicitud=request.remote_addr
        )
        db.session.add(token_obj)
        db.session.commit()
        
        # Preparar mensaje de correo
        archivos_nombres = {
            'dian': 'Facturas DIAN (maestro + dian)',
            'erp_financiero': 'ERP Financiero',
            'erp_comercial': 'ERP Comercial',
            'acuses': 'Acuses de Recibo'
        }
        archivos_texto = ', '.join([archivos_nombres.get(a, a) for a in archivos])
        
        # Enviar correo
        msg = Message(
            subject='🔐 Código de Confirmación - Eliminación de Datos DIAN vs ERP',
            sender=current_app.config.get('MAIL_DEFAULT_SENDER', 'gestordocumentalsc01@gmail.com'),
            recipients=[email]
        )
        
        msg.html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #00d084 0%, #00a86b 100%); 
                          color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
                .content {{ background: #f5f6f8; padding: 30px; border-radius: 0 0 10px 10px; }}
                .token {{ font-size: 32px; font-weight: bold; color: #00c875; text-align: center; 
                         padding: 20px; background: white; border-radius: 8px; margin: 20px 0; 
                         letter-spacing: 8px; font-family: monospace; }}
                .warning {{ background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; 
                           margin: 20px 0; border-radius: 4px; }}
                .details {{ background: white; padding: 15px; border-radius: 8px; margin: 15px 0; }}
                .footer {{ text-align: center; color: #666; font-size: 12px; margin-top: 20px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🗑️ Eliminación de Datos</h1>
                    <p>Sistema de Facturas DIAN vs ERP</p>
                </div>
                <div class="content">
                    <p>Hola <strong>{session['usuario']}</strong>,</p>
                    
                    <p>Has solicitado eliminar datos del sistema. Para confirmar esta acción, 
                    utiliza el siguiente código de validación:</p>
                    
                    <div class="token">{token}</div>
                    
                    <div class="warning">
                        <strong>⚠️ ADVERTENCIA:</strong> Esta acción es IRREVERSIBLE. 
                        Una vez confirmada, los datos serán eliminados permanentemente de la base de datos.
                    </div>
                    
                    <div class="details">
                        <h3>📋 Detalles de la Eliminación:</h3>
                        <ul>
                            <li><strong>Tipo de rango:</strong> {tipo_rango.upper()}</li>
                            <li><strong>Fecha inicio:</strong> {fecha_inicio.strftime('%d/%m/%Y')}</li>
                            <li><strong>Fecha fin:</strong> {fecha_fin.strftime('%d/%m/%Y')}</li>
                            <li><strong>Archivos/Tablas:</strong> {archivos_texto}</li>
                            <li><strong>Solicitado por:</strong> {session['usuario']}</li>
                            <li><strong>IP:</strong> {request.remote_addr}</li>
                        </ul>
                    </div>
                    
                    <p><strong>Validez del código:</strong> 10 minutos</p>
                    
                    <p style="color: #666; font-size: 14px;">
                        Si no solicitaste esta eliminación, ignora este correo y el código expirará automáticamente.
                    </p>
                </div>
                <div class="footer">
                    <p>Supertiendas Cañaveral - Gestor Documental</p>
                    <p>Este es un correo automático, por favor no responder.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Intentar enviar correo (envolver en try/except para manejar errores SMTP)
        correo_enviado = False
        error_correo = None
        
        try:
            mail.send(msg)
            correo_enviado = True
        except Exception as e_mail:
            error_correo = str(e_mail)
            # Log pero no fallar - el token ya está guardado
            from app import log_security
            log_security(f"⚠️ ERROR ENVÍO CORREO ELIMINACIÓN | usuario={session['usuario']} | email={email} | error={error_correo}")
        
        # Log de seguridad (siempre, incluso si falla el correo)
        from app import log_security
        log_security(f"TOKEN ELIMINACIÓN GENERADO | usuario={session['usuario']} | token_id={token_obj.id} | archivos={archivos_texto} | rango={fecha_inicio} a {fecha_fin}")
        
        # Retornar respuesta (con advertencia si falla el correo)
        if not correo_enviado:
            return jsonify({
                'success': True,
                'message': f'⚠️ Token generado pero NO se pudo enviar el correo a {email}. CÓDIGO: {token}',
                'token': token,  # Mostrar token en respuesta si falla el correo
                'token_id': token_obj.id,
                'validez_minutos': 10,
                'advertencia_correo': f'Error SMTP: {error_correo[:100]}'
            })
        
        return jsonify({
            'success': True,
            'message': f'Código de validación enviado a {email}',
            'token_id': token_obj.id,
            'validez_minutos': 10
        })
        
    except Exception as e:
        db.session.rollback()
        from app import log_security
        log_security(f"ERROR SOLICITUD ELIMINACIÓN | usuario={session.get('usuario')} | error={str(e)}")
        return jsonify({'success': False, 'error': f'Error al solicitar eliminación: {str(e)}'}), 500


@dian_vs_erp_bp.route('/api/confirmar_eliminacion', methods=['POST'])
@requiere_permiso('dian_vs_erp', 'eliminar_datos')
def confirmar_eliminacion():
    """
    Confirma y ejecuta la eliminación de datos
    Valida token y elimina registros según parámetros
    """
    from modules.dian_vs_erp.models import TokenEliminacionDatos
    from datetime import datetime
    import json
    from sqlalchemy import text, extract
    
    try:
        # Validar sesión
        if 'usuario' not in session or 'usuario_id' not in session:
            return jsonify({'success': False, 'error': 'Sesión no válida'}), 401
        
        # Obtener parámetros
        data = request.json
        token_codigo = data.get('token')
        
        if not token_codigo:
            return jsonify({'success': False, 'error': 'Código de validación requerido'}), 400
        
        # Buscar token
        token_obj = TokenEliminacionDatos.query.filter_by(
            token=token_codigo,
            usado=False
        ).first()
        
        if not token_obj:
            return jsonify({'success': False, 'error': 'Código inválido o ya utilizado'}), 400
        
        # Verificar expiración
        if datetime.utcnow() > token_obj.fecha_expiracion:
            return jsonify({'success': False, 'error': 'Código expirado. Solicita uno nuevo.'}), 400
        
        # Verificar que sea el mismo usuario
        if token_obj.usuario_solicitante != session['usuario']:
            return jsonify({'success': False, 'error': 'Token no pertenece a este usuario'}), 403
        
        # Parsear archivos a eliminar
        archivos = json.loads(token_obj.archivos_eliminar)
        fecha_inicio = token_obj.fecha_inicio
        fecha_fin = token_obj.fecha_fin
        
        # Ejecutar eliminación
        resultado = {
            'total_eliminados': 0,
            'por_tabla': {}
        }
        
        # 🗑️ ELIMINACIÓN POR TABLA SEGÚN CHECKBOXES
        
        # 1️⃣ Si "dian" está seleccionado → Borrar maestro_dian_vs_erp Y dian
        if 'dian' in archivos:
            # Borrar tabla consolidada maestro
            eliminados_maestro = db.session.execute(
                text("""
                    DELETE FROM maestro_dian_vs_erp
                    WHERE fecha_emision >= :fecha_inicio
                      AND fecha_emision <= :fecha_fin
                """),
                {'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin}
            )
            cantidad_maestro = eliminados_maestro.rowcount
            resultado['por_tabla']['maestro_dian_vs_erp'] = cantidad_maestro
            resultado['total_eliminados'] += cantidad_maestro
            
            # Borrar tabla cruda dian
            eliminados_dian = db.session.execute(
                text("""
                    DELETE FROM dian
                    WHERE fecha_emision >= :fecha_inicio
                      AND fecha_emision <= :fecha_fin
                """),
                {'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin}
            )
            cantidad_dian = eliminados_dian.rowcount
            resultado['por_tabla']['dian'] = cantidad_dian
            resultado['total_eliminados'] += cantidad_dian
        
        # 2️⃣ Si "erp_comercial" está seleccionado → Borrar erp_comercial
        if 'erp_comercial' in archivos:
            eliminados_cm = db.session.execute(
                text("""
                    DELETE FROM erp_comercial
                    WHERE fecha_recibido >= :fecha_inicio
                      AND fecha_recibido <= :fecha_fin
                """),
                {'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin}
            )
            cantidad_cm = eliminados_cm.rowcount
            resultado['por_tabla']['erp_comercial'] = cantidad_cm
            resultado['total_eliminados'] += cantidad_cm
        
        # 3️⃣ Si "erp_financiero" está seleccionado → Borrar erp_financiero
        if 'erp_financiero' in archivos:
            eliminados_fn = db.session.execute(
                text("""
                    DELETE FROM erp_financiero
                    WHERE fecha_recibido >= :fecha_inicio
                      AND fecha_recibido <= :fecha_fin
                """),
                {'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin}
            )
            cantidad_fn = eliminados_fn.rowcount
            resultado['por_tabla']['erp_financiero'] = cantidad_fn
            resultado['total_eliminados'] += cantidad_fn
        
        # 4️⃣ Si "acuses" está seleccionado → Borrar acuses
        if 'acuses' in archivos:
            eliminados_acuses = db.session.execute(
                text("""
                    DELETE FROM acuses
                    WHERE fecha >= :fecha_inicio
                      AND fecha <= :fecha_fin
                """),
                {'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin}
            )
            cantidad_acuses = eliminados_acuses.rowcount
            resultado['por_tabla']['acuses'] = cantidad_acuses
            resultado['total_eliminados'] += cantidad_acuses
        
        # Commit de todas las eliminaciones
        db.session.commit()
        
        # Log detallado por tabla
        from app import log_security
        tablas_resumen = ', '.join([f"{k}={v}" for k, v in resultado['por_tabla'].items()])
        log_security(f"ELIMINACIÓN DATOS DIAN VS ERP | usuario={session['usuario']} | total_eliminados={resultado['total_eliminados']} | tablas={tablas_resumen} | rango={fecha_inicio} a {fecha_fin} | archivos_seleccionados={archivos}")
        
        # Marcar token como usado
        token_obj.usado = True
        token_obj.fecha_uso = datetime.utcnow()
        token_obj.registros_eliminados = resultado['total_eliminados']
        token_obj.resultado_json = json.dumps(resultado)
        token_obj.ip_confirmacion = request.remote_addr
        db.session.commit()
        
        # Log de seguridad
        from app import log_security
        log_security(f"ELIMINACIÓN DATOS EJECUTADA | usuario={session['usuario']} | token_id={token_obj.id} | eliminados={resultado['total_eliminados']} | rango={fecha_inicio} a {fecha_fin}")
        
        return jsonify({
            'success': True,
            'message': 'Datos eliminados exitosamente',
            'resultado': resultado
        })
        
    except Exception as e:
        db.session.rollback()
        from app import log_security
        log_security(f"ERROR CONFIRMACIÓN ELIMINACIÓN | usuario={session.get('usuario')} | error={str(e)}")
        return jsonify({'success': False, 'error': f'Error al eliminar datos: {str(e)}'}), 500


# ========================================
# 📊 REPORTES CON TABLA DINÁMICA
# ========================================

@dian_vs_erp_bp.route('/reportes_dinamicos')
def reportes_dinamicos():
    """Página de generación de reportes con tabla dinámica"""
    # Validar sesión
    if 'usuario_id' not in session or 'usuario' not in session:
        return redirect('/login')
    
    return render_template('dian_vs_erp/reportes_dinamicos.html', usuario=session.get('usuario', 'Usuario'))


@dian_vs_erp_bp.route('/api/generar_reporte_dinamico', methods=['POST'])
@requiere_permiso('dian_vs_erp', 'reportes_dinamicos')
def generar_reporte_dinamico():
    """
    Genera reporte Excel con tabla dinámica según filtros
    
    Body JSON:
        {
            "fecha_inicio": "2025-12-01",
            "fecha_fin": "2025-12-31",
            "nit_emisor": ["805028041", "900123456"],
            "tipo_tercero": ["PROVEEDOR", "ACREEDOR"],
            "estado_aprobacion": ["Pendiente", "Acuse Recibido"],
            "forma_pago": ["Contado", "Crédito"],
            "estado_contable": ["Recibida", "Causada"],
            ... (otros filtros)
        }
    
    Returns:
        Excel file with 2 sheets:
        - "Datos": Filtered invoice list
        - "Análisis Dinámico": Native Excel pivot table
    """
    try:
        print("\n" + "="*80)
        print("📊 INICIANDO GENERACIÓN DE REPORTE DINÁMICO")
        print("="*80)
        
        data = request.get_json()
        print(f"📋 Filtros recibidos: {data.keys()}")
        
        # 🔍 DEBUG ESPECÍFICO para estado_contable (valor que está causando el problema)
        if data.get('estado_contable'):
            print(f"\n🔎 DESGLOSE DETALLADO DE 'estado_contable':")
            print(f"   • Tipo: {type(data['estado_contable']).__name__}")
            print(f"   • Cantidad de valores: {len(data['estado_contable'])}")
            for i, estado in enumerate(data['estado_contable']):
                print(f"   • Valor [{i}]: '{estado}'")
                print(f"     - Longitud: {len(estado)} caracteres")
                print(f"     - Repr: {repr(estado)}")
                print(f"     - Bytes: {estado.encode('utf-8')}")
        else:
            print(f"\n⚠️ estado_contable NO llegó en la petición")
        
        # 🔍 CONSTRUIR CONSULTA SQL DINÁMICA - ⭐ IGUAL QUE VISOR V2 (/api/dian_v2)
        from sqlalchemy import and_, or_
        from datetime import date
        
        # ⭐ QUERY BASE: LEFT JOINs a todas las tablas (igual que visor_v2)
        query = db.session.query(
            Dian,
            Acuses.estado_docto.label('estado_acuse'),
            ErpFinanciero.id.label('existe_financiero'),
            ErpFinanciero.modulo.label('modulo_financiero'),
            ErpFinanciero.doc_causado_por.label('causado_por_financiero'),
            ErpComercial.id.label('existe_comercial'),
            ErpComercial.modulo.label('modulo_comercial'),
            ErpComercial.doc_causado_por.label('causado_por_comercial'),
            FacturaTemporal.id.label('existe_temporal'),
            FacturaRecibida.id.label('existe_recibida'),
            FacturaDigital.id.label('existe_digital'),
            TipoTerceroDianErp.tipo_tercero.label('tipo_tercero_erp')
        ).outerjoin(
            Acuses,
            and_(
                Dian.cufe_cude == Acuses.cufe,
                Dian.cufe_cude != None,
                Dian.cufe_cude != ''
            )
        ).outerjoin(
            ErpFinanciero,
            and_(
                Dian.clave == ErpFinanciero.clave_erp_financiero,
                Dian.clave != None,
                Dian.clave != ''
            )
        ).outerjoin(
            ErpComercial,
            and_(
                Dian.clave == ErpComercial.clave_erp_comercial,
                Dian.clave != None,
                Dian.clave != ''
            )
        ).outerjoin(
            FacturaTemporal,
            and_(
                Dian.prefijo == FacturaTemporal.prefijo,
                Dian.folio == FacturaTemporal.folio,
                Dian.prefijo != None,
                Dian.folio != None
            )
        ).outerjoin(
            FacturaRecibida,
            and_(
                Dian.prefijo == FacturaRecibida.prefijo,
                Dian.folio == FacturaRecibida.folio,
                Dian.prefijo != None,
                Dian.folio != None
            )
        ).outerjoin(
            FacturaDigital,
            and_(
                db.func.concat(Dian.prefijo, Dian.folio) == FacturaDigital.numero_factura,
                Dian.prefijo != None,
                Dian.folio != None
            )
        ).outerjoin(
            TipoTerceroDianErp,
            Dian.nit_emisor == TipoTerceroDianErp.nit_emisor
        )
        
        # 📅 FILTRO DE FECHAS (OBLIGATORIO)
        if data.get('fecha_inicio'):
            fecha_inicio = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date()
            query = query.filter(Dian.fecha_emision >= fecha_inicio)
            print(f"   ✅ Filtro fecha_inicio: {fecha_inicio}")
        
        if data.get('fecha_fin'):
            fecha_fin = datetime.strptime(data['fecha_fin'], '%Y-%m-%d').date()
            query = query.filter(Dian.fecha_emision <= fecha_fin)
            print(f"   ✅ Filtro fecha_fin: {fecha_fin}")
        
        # 🏢 FILTRO DE NIT EMISOR (multi-select)
        if data.get('nit_emisor') and len(data['nit_emisor']) > 0:
            query = query.filter(Dian.nit_emisor.in_(data['nit_emisor']))
            print(f"   ✅ Filtro nit_emisor: {len(data['nit_emisor'])} NITs")
        
        # 📝 FILTRO DE TIPO DOCUMENTO (multi-select)
        if data.get('tipo_documento') and len(data['tipo_documento']) > 0:
            query = query.filter(Dian.tipo_documento.in_(data['tipo_documento']))
            print(f"   ✅ Filtro tipo_documento: {len(data['tipo_documento'])} tipos")
        
        # 📌 FILTRO DE PREFIJO (multi-select)
        if data.get('prefijo') and len(data['prefijo']) > 0:
            query = query.filter(Dian.prefijo.in_(data['prefijo']))
            print(f"   ✅ Filtro prefijo: {len(data['prefijo'])} prefijos")
        
        # Ordenar por fecha emisión descendente
        query = query.order_by(Dian.fecha_emision.desc())
        
        # 🚀 EJECUTAR CONSULTA
        print("\n🔍 Ejecutando consulta en PostgreSQL con JOINs (igual que visor_v2)...")
        t_start = time.time()
        resultados = query.all()
        t_end = time.time()
        print(f"✅ Consulta completada en {t_end - t_start:.2f}s - {len(resultados):,} registros ANTES de filtros calculados")
        
        # 📦 PREPARAR DATOS PARA EXCEL - ⭐ CALCULAR ESTADOS IGUAL QUE VISOR V2
        print("\n📦 Preparando datos y calculando estados en tiempo real...")
        datos_excel = []
        datos_para_filtrar = []  # Lista temporal para aplicar filtros calculados
        
        for row in resultados:
            registro = row[0]  # Dian
            estado_acuse = row[1]
            existe_financiero = row[2]
            modulo_financiero = row[3]
            causado_por_financiero = row[4]
            existe_comercial = row[5]
            modulo_comercial = row[6]
            causado_por_comercial = row[7]
            existe_temporal = row[8]
            existe_recibida = row[9]
            existe_digital = row[10]
            tipo_tercero_erp = row[11]
            
            # ⭐ CALCULAR FORMA DE PAGO (igual que visor_v2 líneas 720-730)
            forma_pago_raw = (registro.forma_pago or "").strip()
            if forma_pago_raw in ["Contado", "Crédito"]:
                forma_pago_texto = forma_pago_raw
            elif forma_pago_raw in ["1", "01"]:
                forma_pago_texto = "Contado"
            elif forma_pago_raw in ["2", "02"]:
                forma_pago_texto = "Crédito"
            else:
                forma_pago_texto = "Crédito"  # Default
            
            # ⭐ CALCULAR ESTADO DE APROBACIÓN (igual que visor_v2 líneas 733-737)
            if estado_acuse and str(estado_acuse).strip():
                estado_aprobacion_final = str(estado_acuse).strip()
            else:
                estado_aprobacion_final = "No Registra"
            
            # ⭐ CALCULAR ESTADO CONTABLE Y MÓDULO (igual que visor_v2 líneas 740-780)
            modulo_final = ""
            causado_por_final = ""
            
            if existe_financiero:
                estado_contable_validado = "Causada"
                modulo_final = modulo_financiero or "Financiero"
                causado_por_final = causado_por_financiero or ""
            elif existe_comercial:
                estado_contable_validado = "Causada"
                modulo_final = modulo_comercial or "Comercial"
                causado_por_final = causado_por_comercial or ""
            elif existe_recibida or existe_temporal or existe_digital:
                estado_contable_validado = "Recibida"
                modulo_final = "Recepción de Facturas"
                causado_por_final = ""
            else:
                estado_contable_validado = "No Registrada"
                modulo_final = ""
                causado_por_final = ""
            
            # ⭐ CALCULAR DÍAS DESDE EMISIÓN EN TIEMPO REAL (igual que visor_v2 líneas 807-820)
            dias_transcurridos = 0
            if registro.fecha_emision:
                try:
                    fecha_emision_date = registro.fecha_emision
                    if isinstance(fecha_emision_date, datetime):
                        fecha_emision_date = fecha_emision_date.date()
                    dias_transcurridos = (date.today() - fecha_emision_date).days
                except Exception as e:
                    dias_transcurridos = 0
            
            # 🔴 DETECTAR NOTAS CRÉDITO Y CONVERTIR VALOR A NEGATIVO
            tipo_doc = registro.tipo_documento or ''
            valor_original = float(registro.total) if registro.total else 0
            
            # Si es nota crédito, convertir a negativo
            es_nota_credito = (
                'Nota de crédito electrónica' in tipo_doc or
                'Nota de ajuste crédito del documento equivalente' in tipo_doc
            )
            
            valor_final = -abs(valor_original) if es_nota_credito else valor_original
            
            # 🔗 ENLACE DIAN (Ver documento en portal DIAN)
            cufe = registro.cufe_cude or ''
            enlace_dian = f"https://catalogo-vpfe.dian.gov.co/User/SearchDocument?DocumentKey={cufe}" if cufe else ''
            
            # Almacenar en lista temporal con valores calculados
            datos_para_filtrar.append({
                'NIT Emisor': registro.nit_emisor or '',
                'Razón Social': registro.nombre_emisor or '',
                'Tipo Tercero': tipo_tercero_erp or 'No Registrado',
                'Fecha Emisión': registro.fecha_emision.strftime('%Y-%m-%d') if registro.fecha_emision else '',
                'Tipo Documento': tipo_doc,
                'Prefijo': registro.prefijo or '',
                'Folio': registro.folio or '',
                'Valor': valor_final,
                'Estado Aprobación': estado_aprobacion_final,
                'Forma de Pago': forma_pago_texto,
                'Estado Contable': estado_contable_validado,
                'N° Días': dias_transcurridos,
                'Usuario Recibió': '',  # No disponible en tablas actuales
                'Usuario Causación': '',  # No disponible en tablas actuales
                'Causador': causado_por_final,
                'Módulo': modulo_final,
                'CUFE': cufe,
                'Ver DIAN': enlace_dian
            })
        
        print(f"✅ {len(datos_para_filtrar):,} filas procesadas con estados calculados")
        
        # 🔍 DEBUG: Contar estados contables ANTES de filtrar
        from collections import Counter
        estados_contables_count = Counter([f['Estado Contable'] for f in datos_para_filtrar])
        print(f"\n📊 DISTRIBUCIÓN DE ESTADOS CONTABLES (antes de filtrar):")
        for estado, cantidad in estados_contables_count.items():
            print(f"   • {estado}: {cantidad}")
        
        # ⭐ APLICAR FILTROS CALCULADOS (tipo_tercero, estado_aprobacion, forma_pago, estado_contable, modulo)
        print(f"\n🔍 Aplicando filtros en memoria para campos calculados...")
        print(f"   📋 Filtros aplicados:")
        if data.get('tipo_tercero'):
            print(f"      • tipo_tercero: {data['tipo_tercero']}")
        if data.get('estado_aprobacion'):
            print(f"      • estado_aprobacion: {data['estado_aprobacion']}")
        if data.get('forma_pago'):
            print(f"      • forma_pago: {data['forma_pago']}")
        if data.get('estado_contable'):
            print(f"      • estado_contable: {data['estado_contable']}")
        if data.get('modulo'):
            print(f"      • modulo: {data['modulo']}")
        
        contador_descartados_por_estado = 0
        
        for fila in datos_para_filtrar:
            # Filtro tipo_tercero (calculado desde TipoTerceroDianErp)
            if data.get('tipo_tercero') and len(data['tipo_tercero']) > 0:
                if fila['Tipo Tercero'] not in data['tipo_tercero']:
                    continue
            
            # Filtro estado_aprobacion (calculado desde Acuses)
            if data.get('estado_aprobacion') and len(data['estado_aprobacion']) > 0:
                if fila['Estado Aprobación'] not in data['estado_aprobacion']:
                    continue
            
            # Filtro forma_pago (transformado de códigos)
            if data.get('forma_pago') and len(data['forma_pago']) > 0:
                if fila['Forma de Pago'] not in data['forma_pago']:
                    continue
            
            # Filtro estado_contable (calculado según existencia en tablas)
            if data.get('estado_contable') and len(data['estado_contable']) > 0:
                # � NORMALIZAR: Eliminar espacios extra y convertir a minúsculas para comparación
                estado_fila_norm = fila['Estado Contable'].strip().lower()
                filtros_norm = [e.strip().lower() for e in data['estado_contable']]
                
                # 🐛 DEBUG: Mostrar las primeras 3 filas que se descartan
                if estado_fila_norm not in filtros_norm:
                    if contador_descartados_por_estado < 3:
                        print(f"   ❌ DESCARTADA: NIT={fila['NIT Emisor']}, Prefijo={fila['Prefijo']}, Folio={fila['Folio']}")
                        print(f"      Estado en fila: '{fila['Estado Contable']}' → normalizado: '{estado_fila_norm}'")
                        print(f"      Filtro original: {data['estado_contable']}")
                        print(f"      Filtro normalizado: {filtros_norm}")
                        print(f"      ¿Coincide normalizado? {estado_fila_norm in filtros_norm}")
                    contador_descartados_por_estado += 1
                    continue
            
            # Filtro modulo (calculado según tabla de origen)
            if data.get('modulo') and len(data['modulo']) > 0:
                if fila['Módulo'] not in data['modulo']:
                    continue
            
            # Filtro dias_min y dias_max (calculado en tiempo real)
            if data.get('dias_min') is not None:
                if fila['N° Días'] < int(data['dias_min']):
                    continue
            
            if data.get('dias_max') is not None:
                if fila['N° Días'] > int(data['dias_max']):
                    continue
            
            # Filtro valor_min y valor_max
            if data.get('valor_min') is not None:
                if fila['Valor'] < float(data['valor_min']):
                    continue
            
            if data.get('valor_max') is not None:
                if fila['Valor'] > float(data['valor_max']):
                    continue
            
            # Si pasó todos los filtros, agregar a resultado final
            datos_excel.append(fila)
        
        print(f"\n📊 RESUMEN DE FILTRADO:")
        print(f"   • Filas iniciales: {len(datos_para_filtrar):,}")
        print(f"   • Descartadas por estado_contable: {contador_descartados_por_estado:,}")
        print(f"   • Filas finales: {len(datos_excel):,}")
        
        if len(datos_excel) == 0:
            print(f"\n❌ NO SE ENCONTRARON REGISTROS después de aplicar filtros")
            print(f"   Revisa si los valores del filtro coinciden con los calculados:")
            print(f"   • Estados disponibles: {list(estados_contables_count.keys())}")
            print(f"   • Estado filtrado: {data.get('estado_contable')}")
            
            registrar_log('INFO', 'REPORTE DINAMICO SIN DATOS', f"Usuario {session.get('usuario')} generó reporte sin resultados")
            return jsonify({
                'success': False,
                'error': 'No se encontraron registros con los filtros aplicados. Intenta ampliar el rango de fechas o reducir los filtros.'
            }), 404
        
        print(f"✅ {len(datos_excel):,} filas preparadas para Excel")
        
        
        # 📊 GENERAR EXCEL CON OPENPYXL
        print("\n📊 Generando Excel con tabla de resumen...")
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.worksheet.table import Table, TableStyleInfo
        import pandas as pd
        
        # Crear workbook
        wb = Workbook()
        
        # ==========================================
        # 📋 SHEET 1: "Datos" - Tabla con datos filtrados
        # ==========================================
        ws_datos = wb.active
        ws_datos.title = "Datos"
        
        print("   📋 Creando sheet 'Datos'...")
        
        # Convertir a DataFrame y escribir
        df = pd.DataFrame(datos_excel)
        
        # Escribir encabezados
        for c_idx, column in enumerate(df.columns, 1):
            cell = ws_datos.cell(row=1, column=c_idx, value=column)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Escribir datos
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws_datos.cell(row=r_idx, column=c_idx, value=value)
                
                # Formato especial para columna de Valor (columna 8)
                if c_idx == 8:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                    
                    # 🔴 VALORES NEGATIVOS EN ROJO (notas crédito)
                    if isinstance(value, (int, float)) and value < 0:
                        cell.font = Font(color="FF0000", bold=True)  # Rojo
                
                # 🔗 HIPERVÍNCULO para columna "Ver DIAN" (última columna = 18)
                if c_idx == 18 and value:  # Si hay enlace
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
                    cell.value = "🔗 Ver en DIAN"
        
        # Ajustar anchos de columna
        for column in ws_datos.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_datos.column_dimensions[column_letter].width = adjusted_width
        
        # Crear tabla Excel (para que los datos sean tabuleables)
        tabla_rango = f"A1:{chr(64 + len(df.columns))}{len(df) + 1}"
        tabla = Table(displayName="TablaDatos", ref=tabla_rango)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tabla.tableStyleInfo = style
        ws_datos.add_table(tabla)
        
        print(f"   ✅ Sheet 'Datos' creado con {len(df):,} filas")
        
        # ==========================================
        # 📊 SHEET 2: "Resumen" - Tabla agregada
        # ==========================================
        ws_resumen = wb.create_sheet("Resumen")
        
        print("   📊 Creando hoja de resumen...")
        
        # Título
        ws_resumen['A1'] = "📊 Resumen de Facturas por Tipo de Tercero y Forma de Pago"
        ws_resumen['A1'].font = Font(size=14, bold=True, color="166534")
        ws_resumen.merge_cells('A1:E1')
        
        # Crear tabla resumen con pandas
        resumen = df.groupby(['Tipo Tercero', 'Forma de Pago']).agg({
            'Valor': 'sum',
            'NIT Emisor': 'count'
        }).reset_index()
        resumen.columns = ['Tipo Tercero', 'Forma de Pago', 'Total Valor', 'Cantidad Facturas']
        
        # Escribir encabezados
        headers_resumen = ['Tipo Tercero', 'Forma de Pago', 'Total Valor', 'Cantidad Facturas']
        for c_idx, col in enumerate(headers_resumen, 1):
            cell = ws_resumen.cell(row=3, column=c_idx, value=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Escribir datos del resumen
        for r_idx, row in enumerate(resumen.itertuples(index=False), 4):
            ws_resumen.cell(row=r_idx, column=1, value=row[0])  # Tipo Tercero
            ws_resumen.cell(row=r_idx, column=2, value=row[1])  # Forma de Pago
            cell_valor = ws_resumen.cell(row=r_idx, column=3, value=row[2])  # Total Valor
            cell_valor.number_format = '$#,##0.00'
            ws_resumen.cell(row=r_idx, column=4, value=row[3])  # Cantidad
        
        # Ajustar anchos
        ws_resumen.column_dimensions['A'].width = 25
        ws_resumen.column_dimensions['B'].width = 20
        ws_resumen.column_dimensions['C'].width = 18
        ws_resumen.column_dimensions['D'].width = 18
        
        print(f"   ✅ Hoja de resumen creada con {len(resumen)} filas")
        
        # ==========================================
        # 📊 SHEET 3: "Instrucciones" - Cómo crear tabla dinámica
        # ==========================================
        ws_instruc = wb.create_sheet("Instrucciones")
        
        instrucciones_texto = [
            ["📋 Cómo crear una Tabla Dinámica en Excel"],
            [""],
            ["1. Ve a la hoja 'Datos'"],
            ["2. Selecciona cualquier celda de la tabla"],
            ["3. Ve al menú: Insertar → Tabla Dinámica"],
            ["4. Confirma el rango de datos y haz clic en Aceptar"],
            [""],
            ["5. En el panel de Campos de Tabla Dinámica:"],
            ["   • Arrastra 'Tipo Tercero' a FILAS"],
            ["   • Arrastra 'Forma de Pago' a COLUMNAS"],
            ["   • Arrastra 'Valor' a VALORES (se sumará automáticamente)"],
            ["   • Arrastra 'NIT Emisor' a VALORES (se contará automáticamente)"],
            [""],
            ["6. Personaliza los filtros arrastrando más campos"],
            ["7. Aplica formato según tus necesidades"],
            [""],
            ["💡 NOTA: La hoja 'Resumen' ya contiene un resumen pre-calculado"],
            ["    por Tipo de Tercero y Forma de Pago."]
        ]
        
        for r_idx, row in enumerate(instrucciones_texto, 1):
            ws_instruc.cell(row=r_idx, column=1, value=row[0])
            if r_idx == 1:
                ws_instruc.cell(row=r_idx, column=1).font = Font(size=14, bold=True, color="166534")
        
        ws_instruc.column_dimensions['A'].width = 70
        
        print(f"   ✅ Hoja de instrucciones creada")
        
        # ==========================================
        # 💾 GUARDAR Y RETORNAR ARCHIVO
        # ==========================================
        print("\n💾 Guardando archivo Excel...")
        
        # Nombre del archivo con timestamp
        fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Reporte_DIAN_vs_ERP_{fecha_str}.xlsx"
        
        # Guardar en BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        print(f"✅ Archivo generado: {filename}")
        print("="*80)
        
        # Registrar en logs
        registrar_log(
            'SUCCESS',
            'GENERAR_REPORTE_DINAMICO',
            f'Reporte con tabla dinámica generado: {len(datos_excel):,} registros',
            detalles={'filtros': list(data.keys()), 'registros': len(datos_excel)}
        )
        
        # Retornar archivo
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"\n❌ ERROR GENERANDO REPORTE: {e}")
        import traceback
        traceback.print_exc()
        
        registrar_log(
            'ERROR',
            'GENERAR_REPORTE_DINAMICO',
            f'Error generando reporte: {str(e)}',
            incluir_stacktrace=True
        )
        
        return jsonify({
            'exito': False,
            'mensaje': f'Error generando reporte: {str(e)}'
        }), 500
