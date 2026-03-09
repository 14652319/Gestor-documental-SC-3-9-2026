"""
Script muy simple - solo nombres de columnas
"""
import openpyxl

print("=" * 80)
print("INSPECCION DE COLUMNAS")
print("=" * 80)

# DIAN
print("\nARCHIVO DIAN:")
try:
    wb = openpyxl.load_workbook(r'uploads\dian\Dian.xlsx', read_only=True, data_only=True)
    ws = wb.active
    
    # Leer solo primera fila (headers)
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value))
    
    print(f"\nColumnas encontradas ({len(headers)}):")
    for i, h in enumerate(headers, 1):
        print(f"   {i}. {h}")
    
    # Buscar CUFE
    cufe_col_idx = None
    for i, h in enumerate(headers):
        if h and ('cufe' in h.lower() or 'cude' in h.lower()):
            cufe_col_idx = i
            print(f"\n>>> CUFE encontrado en columna {i+1}: '{h}'")
            break
    
    # Leer primer CUFE
    if cufe_col_idx is not None:
        print(f"\nPrimeros 3 valores de CUFE:")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=4), 1):
            valor = row[cufe_col_idx].value
            if valor:
                print(f"   {row_idx}. {str(valor)[:80]}")
    
    wb.close()
except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()

# ACUSES
print("\n" + "=" * 80)
print("ARCHIVO ACUSES:")
try:
    wb = openpyxl.load_workbook(r'uploads\acuses\acuses.xlsx', read_only=True, data_only=True)
    ws = wb.active
    
    # Leer solo primera fila (headers)
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value))
    
    print(f"\nColumnas encontradas ({len(headers)}):")
    for i, h in enumerate(headers, 1):
        print(f"   {i}. {h}")
    
    # Buscar CUFE
    cufe_col_idx = None
    for i, h in enumerate(headers):
        if h and ('cufe' in h.lower() or 'cude' in h.lower()):
            cufe_col_idx = i
            print(f"\n>>> CUFE encontrado en columna {i+1}: '{h}'")
            break
    
    # Buscar ESTADO
    estado_col_idx = None
    for i, h in enumerate(headers):
        if h and 'estado' in h.lower():
            estado_col_idx = i
            print(f">>> ESTADO encontrado en columna {i+1}: '{h}'")
            break
    
    # Leer primeros CUFEs
    if cufe_col_idx is not None:
        print(f"\nPrimeros 3 valores de CUFE:")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=4), 1):
            cufe_val = row[cufe_col_idx].value
            estado_val = row[estado_col_idx].value if estado_col_idx else 'N/A'
            if cufe_val:
                print(f"   {row_idx}. CUFE: {str(cufe_val)[:70]}")
                print(f"       Estado: {estado_val}")
    
    wb.close()
except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 80)
print("FIN")
print("=" * 80)
