"""
VERIFICAR DUPLICADOS EN EXCEL - SIN EMOJIS
"""
from openpyxl import load_workbook
from collections import Counter
import os

archivo = r'uploads\dian\Dian_23022026.xlsx'

print("\n" + "="*80)
print("VERIFICANDO DUPLICADOS EN EXCEL DIAN")
print("="*80)

if not os.path.exists(archivo):
    print("\nArchivo no encontrado")
    exit()

print("\nAbriendo Excel con openpyxl (solo lectura)...")

# Abrir Excel
wb = load_workbook(archivo, read_only=True, data_only=True)
ws = wb.active

print("Excel abierto correctamente")

# Leer header (primera fila)
headers = []
for cell in ws[1]:
    headers.append(str(cell.value).strip().lower())

print(f"Headers leidos: {len(headers)} columnas")

# Buscar columna CUFE
col_cufe_idx = None
for idx, header in enumerate(headers):
    if 'cufe' in header or 'cude' in header:
        col_cufe_idx = idx + 1  # openpyxl usa índice 1-based
        print(f"Columna CUFE encontrada: '{header}' en posicion {col_cufe_idx}")
        break

if not col_cufe_idx:
    print("\nNo se encontro columna CUFE")
    wb.close()
    exit()

print(f"\nLeyendo CUFEs de la columna {col_cufe_idx}...")

# Leer todos los CUFEs
cufes = []
fila_num = 0
for row in ws.iter_rows(min_row=2, min_col=col_cufe_idx, max_col=col_cufe_idx):
    fila_num += 1
    cufe = row[0].value
    if cufe:
        cufes.append(str(cufe).strip())
    
    if fila_num % 10000 == 0:
        print(f"   Leidas {fila_num:,} filas...")

wb.close()

print(f"Total CUFEs leidos: {len(cufes):,}")

# Análisis
unicos = len(set(cufes))
duplicados = len(cufes) - unicos

print(f"\n{'='*80}")
print("RESULTADO DEL ANALISIS")
print("="*80)
print(f"Total CUFEs: {len(cufes):,}")
print(f"CUFEs unicos: {unicos:,}")
print(f"Duplicados encontrados: {duplicados:,}")

if duplicados > 0:
    print(f"\n*** ATENCION: HAY {duplicados:,} CUFEs DUPLICADOS EN EL EXCEL ***")
    
    # Contar repeticiones
    contador = Counter(cufes)
    dups = [(cufe, count) for cufe, count in contador.items() if count > 1]
    dups.sort(key=lambda x: x[1], reverse=True)
    
    print(f"\nTop 10 CUFEs mas repetidos:")
    cufe_error = "929f7761de9ff5fd92865b32d3aabbd4e056589f9cb854c0cfc15a570564f657aba35f6131872b4523c43152f393c793"
    
    for i, (cufe, count) in enumerate(dups[:10]):
        marca = " <<< ESTE ES EL CUFE DEL ERROR!" if cufe == cufe_error else ""
        print(f"   {i+1}. {cufe[:50]}... -> {count} veces{marca}")
    
    print("\n" + "="*80)
    print("CONCLUSION:")
    print("="*80)
    print("El Excel tiene CUFEs duplicados.")
    print("NO PUEDES CARGAR este archivo sin limpiarlo primero.")
    print("\nSOLUCIONES:")
    print("1. Eliminar manualmente las facturas duplicadas del Excel")
    print("2. Pedir al proveedor un Excel SIN duplicados")
    print("3. Usar filtros de Excel para eliminar duplicados")
    print("="*80)
    
else:
    print(f"\n*** PERFECTO: NO HAY CUFEs DUPLICADOS EN EL EXCEL ***")
    print("\nSi las tablas estan vacias Y el Excel no tiene duplicados,")
    print("entonces el problema esta en el CODIGO de insercion.")
    print("\nPosibles causas:")
    print("- El codigo esta insertando cada registro DOS VECES")
    print("- Hay un loop que repite la insercion")
    print("- El proceso se ejecuto dos veces simultaneamente")

print("\n")
