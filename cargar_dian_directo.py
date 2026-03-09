"""
Script para cargar archivo DIAN directamente a PostgreSQL
Sin dependencia de pyarrow - usa openpyxl + psycopg2
"""
import os
import sys
from pathlib import Path
import psycopg2
from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal

print("=" * 80)
print("CARGA DIRECTA DE ARCHIVO DIAN A POSTGRESQL")
print("=" * 80)

# Ruta al archivo DIAN
archivo_dian = r"C:\Users\Usuario\Downloads\Ricardo - copia\Dian.xlsx"

if not os.path.exists(archivo_dian):
    print(f"❌ ERROR: No se encontró el archivo:")
    print(f"   {archivo_dian}")
    sys.exit(1)

print(f"✅ Archivo encontrado: {Path(archivo_dian).name}")
print(f"   Tamaño: {os.path.getsize(archivo_dian) / 1024:.2f} KB")

# Conectar a PostgreSQL
try:
    conn = psycopg2.connect(
        host="localhost",
        database="gestor_documental",
        user="gestor_user",
        password="Gest0rP@ssw0rd2024"
    )
    conn.autocommit = False
    cursor = conn.cursor()
    print("✅ Conexión a PostgreSQL exitosa")
except Exception as e:
    print(f"❌ ERROR al conectar a PostgreSQL: {e}")
    sys.exit(1)

# Leer archivo Excel con openpyxl
try:
    print("\n📖 Leyendo archivo Excel...")
    wb = load_workbook(archivo_dian, read_only=True, data_only=True)
    ws = wb.active
    
    # Obtener encabezados (primera fila)
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else "")
    
    print(f"✅ Columnas encontradas: {len(headers)}")
    print(f"   {headers[:5]}...")  # Mostrar primeras 5
    
    # Leer datos
    filas = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if idx > 102:  # Solo las primeras 100 filas de datos (+ header + 1)
            break
        if row[0]:  # Si la primera columna tiene datos
            filas.append(row)
    
    total_filas = len(filas)
    print(f"✅ Total de filas a cargar: {total_filas}")
    
except Exception as e:
    print(f"❌ ERROR al leer Excel: {e}")
    conn.close()
    sys.exit(1)

# Mapeo de columnas (ajustar según tu estructura real)
# Esto es un ejemplo - debes ajustarlo a tu archivo real
COLUMNAS_ESPERADAS = [
    'empresa_ruc', 'nit', 'razon_social', 'tipo_documento', 'prefijo', 
    'folio', 'fecha_factura', 'fecha_vencimiento', 'valor_total',
    'subtotal', 'iva', 'ica', 'estado_dian', 'cufe', 'clase_documento',
    'periodo_factura', 'fecha_recepcion', 'observaciones'
]

# Normalizar nombres de columnas
def normalizar_columna(nombre):
    """Normaliza nombres de columna"""
    # Remove special chars, lowercase, replace spaces with underscore
    import re
    nombre = nombre.lower().strip()
    nombre = re.sub(r'[áäà]', 'a', nombre)
    nombre = re.sub(r'[éëè]', 'e', nombre)
    nombre = re.sub(r'[íïì]', 'i', nombre)
    nombre = re.sub(r'[óöò]', 'o', nombre)
    nombre = re.sub(r'[úü]', 'u', nombre)
    nombre = re.sub(r'[^a-z0-9_\s]', '', nombre)
    nombre = re.sub(r'\s+', '_', nombre)
    return nombre

headers_norm = [normalizar_columna(h) for h in headers]
print(f"\n📋 Columnas normalizadas:")
for i, (orig, norm) in enumerate(zip(headers[:10], headers_norm[:10])):
    print(f"   {i+1}. '{orig}' → '{norm}'")

# Insertar en base de datos
try:
    print(f"\n🚀 Insertando {total_filas} filas en tabla 'dian'...")
    
    # Verificar si la tabla existe
    cursor.execute("""
        SELECT EXISTS (
            SELECT FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_name = 'dian'
        );
    """)
    tabla_existe = cursor.fetchone()[0]
    
    if not tabla_existe:
        print("❌ ERROR: La tabla 'dian' no existe en la base de datos")
        print("   Ejecuta primero: python inicializar_dian_vs_erp.py")
        conn.close()
        sys.exit(1)
    
    # Obtener columnas de la tabla real
    cursor.execute("""
        SELECT column_name, data_type 
        FROM information_schema.columns 
        WHERE table_name = 'dian' 
        ORDER BY ordinal_position;
    """)
    columnas_db = cursor.fetchall()
    print(f"\n✅ Tabla 'dian' tiene {len(columnas_db)} columnas:")
    for col, tipo in columnas_db[:10]:
        print(f"   - {col} ({tipo})")
    
    # Crear mapeo de columnas Excel → DB
    mapeo = {}
    for i, header_norm in enumerate(headers_norm):
        for col_db, _ in columnas_db:
            if header_norm == col_db or header_norm.replace('_', '') == col_db.replace('_', ''):
                mapeo[i] = col_db
                break
    
    print(f"\n📌 Mapeo de columnas (Excel → DB):")
    for idx_excel, col_db in sorted(mapeo.items())[:15]:
        print(f"   Excel col {idx_excel} ('{headers[idx_excel]}') → DB '{col_db}'")
    
    # Preparar INSERT dinámico
    columnas_insert = list(mapeo.values())
    placeholders = ', '.join(['%s'] * len(columnas_insert))
    columnas_str = ', '.join(columnas_insert)
    
    query = f"""
        INSERT INTO dian ({columnas_str})
        VALUES ({placeholders})
    """
    
    print(f"\n⚙️  Query preparada:")
    print(f"   INSERT INTO dian ({columnas_str[:80]}...)")
    
    # Insertar filas
    insertadas = 0
    errores = 0
    
    for i, fila in enumerate(filas, start=1):
        try:
            # Extraer valores según mapeo
            valores = []
            for idx_excel in sorted(mapeo.keys()):
                valor = fila[idx_excel] if idx_excel < len(fila) else None
                
                # Convertir tipos
                if valor is not None:
                    if isinstance(valor, datetime):
                        valores.append(valor)
                    elif isinstance(valor, (int, float, Decimal)):
                        valores.append(valor)
                    else:
                        valores.append(str(valor).strip() if valor else None)
                else:
                    valores.append(None)
            
            # Insertar
            cursor.execute(query, valores)
            insertadas += 1
            
            if insertadas % 10 == 0:
                print(f"   Procesadas: {insertadas}/{total_filas} filas...")
                
        except Exception as e:
            errores += 1
            print(f"   ⚠️  Error en fila {i}: {str(e)[:100]}")
            if errores > 10:
                print("   ❌ Demasiados errores, abortando...")
                raise
    
    # Commit
    conn.commit()
    print(f"\n✅ ÉXITO: {insertadas} filas insertadas correctamente")
    if errores > 0:
        print(f"   ⚠️  {errores} filas con errores (omitidas)")
    
    # Verificar inserción
    cursor.execute("SELECT COUNT(*) FROM dian")
    total_db = cursor.fetchone()[0]
    print(f"\n📊 Total registros en tabla 'dian': {total_db}")
    
    # Mostrar últimos registros insertados
    cursor.execute("""
        SELECT nit, razon_social, prefijo, folio, valor_total, fecha_factura
        FROM dian 
        ORDER BY id DESC 
        LIMIT 5
    """)
    print(f"\n📋 Últimos 5 registros insertados:")
    for row in cursor.fetchall():
        print(f"   {row}")
    
except Exception as e:
    conn.rollback()
    print(f"\n❌ ERROR durante la inserción: {e}")
    import traceback
    traceback.print_exc()
finally:
    cursor.close()
    conn.close()
    print("\n✅ Conexión cerrada")

print("\n" + "=" * 80)
print("PROCESO COMPLETADO")
print("=" * 80)
